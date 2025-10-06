import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, date, time as dtime
from openpyxl import load_workbook
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException
import time
import subprocess

try: 
    # --- file formpelapor.py
    # --- Setup Selenium untuk Brave ---
    options = Options()
    options.debugger_address = "127.0.0.1:9222"
    service = Service("chromedriver.exe")
    driver = webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 20)

    # --- Load nomor inklusi dari Excel ---
    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb.active
    val_no_inklusi = str(ws['K3'].value).strip()

    # --- Fungsi konversi tanggal menjadi dd-mm-YYYY (robust) ---
    def format_tanggal_ddmmyyyy(value):
        if value is None:
            return ""
        # kalau sudah bertipe datetime.datetime atau datetime.date
        if isinstance(value, (datetime, date)):
            return value.strftime("%d-%m-%Y")
        s = str(value).strip()
        if not s:
            return ""
        # coba beberapa format string umum
        possible = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d %B %Y", "%d %b %Y"]
        for fmt in possible:
            try:
                dt = datetime.strptime(s, fmt)
                return dt.strftime("%d-%m-%Y")
            except Exception:
                pass
        # fallback: jika isi berupa angka (mungkin serial excel) -> kembalikan as-is,
        # tapi biasanya openpyxl sudah mengembalikan datetime untuk sel berformat tanggal.
        return s

    def set_text(driver, wait, id_, value):
        if value is None or str(value).strip() == "":
            return False
        try:
            el = wait.until(EC.visibility_of_element_located((By.ID, id_)))
            el.click()
            time.sleep(0.1)
            el.clear()
            time.sleep(0.1)
            el.send_keys(str(value))
            print(f"→ {id_} = {value}")
            return True
        except Exception as e:
            print(f"! failed set_text {id_}: {e}")
            try:
                # fallback pake javascript set value supaya lebih pasti masuk
                el = driver.find_element(By.ID, id_)
                driver.execute_script(
                    "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input'));", el, str(value)
                )
                print(f"→ {id_} set via JS fallback = {value}")
                return True
            except Exception as e2:
                print(f"! JS fallback failed for {id_}: {e2}")
                return False

    # --- Load data dari Excel ---
    val_no_inklusi  = ws["K3"].value   # itemid_58832
    val_inisial     = ws["C3"].value   # itemid_58833
    val_provinsi       = ws["M3"].value   # form_group_58835 (radio value)
    val_tgl_lapor       = ws["N3"].value   # itemid_58836 
    val_hasPengobatan   = ws["O3"].value # checkbox

    print("Start to fill-in Form Pelapor")
    print(f"📘 No. inklusi from Excel: {val_no_inklusi}")
    print("Data from Excel:", val_no_inklusi, val_inisial, val_provinsi, val_tgl_lapor, val_hasPengobatan)

    # --- Connect ke Brave yang sudah dibuka dengan --remote-debugging-port=9222 ---
    # options = webdriver.ChromeOptions()
    # options.debugger_address = "127.0.0.1:9222"  # pastikan Brave dijalankan dengan remote debugging
    # driver = webdriver.Chrome(options=options)

    # --- Isi form ---

    # Text input itemid_58832
    # set_text(driver, wait, "itemid_58832", val_no_inklusi)
    # set_text(driver, wait, "itemid_58833", val_inisial)

    # Radio button form_group_58835
    try:
        if val_provinsi is not None and str(val_provinsi).strip() != "":
            css = f"#form_group_58835 input[type='radio'][value='{val_provinsi}']"
            radio = driver.find_element(By.CSS_SELECTOR, css)
            # gunakan JS click untuk menghindari intercept
            driver.execute_script("arguments[0].click();", radio)
            print("→ form_group_58835 =", val_provinsi)
    except Exception as e:
        print("❌ FAILED to select radio form_group_58835:", e)

    # Datepicker itemid_58836 -> format dd-mm-YYYY
    try:
        if val_tgl_lapor:
            formatted = format_tanggal_ddmmyyyy(val_tgl_lapor)
            field = driver.find_element(By.ID, "itemid_58836")
            # set via JS agar datepicker tidak mengoverride; dispatch input agar framework tangkap
            driver.execute_script(
                "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input'));",
                field, formatted
            )
            time.sleep(0.2)
            # shift focus supaya datepicker tertutup — pakai TAB
            field.send_keys(Keys.TAB)
            print("→ itemid_58836 =", formatted)
    except Exception as e:
        print("❌ FAILED to set itemid_58836:", e)

    # Checkbox hasPengobatan
    try:
        checkbox = driver.find_element(By.ID, "hasPengobatan")
        should_check = False
        if val_hasPengobatan is not None:
            hv = str(val_hasPengobatan).strip().lower()
            if hv in ["1", "ya", "true", "x", "yes"]:
                should_check = True
        # set checkbox sesuai flag
        if should_check and not checkbox.is_selected():
            driver.execute_script("arguments[0].click();", checkbox)
        if (not should_check) and checkbox.is_selected():
            driver.execute_script("arguments[0].click();", checkbox)
        print("→ hasPengobatan =", should_check)
    except Exception as e:
        print("❌ FAILED to set hasPengobatan checkbox:", e)

    print("✅ Copy-paste INFORMASI PELAPOR: DONE.")
    #
    #
    #
    # >>>> OPSI 1 : otomatis save, lanjut ke script berikutnya <<<<<
    # >>>> DIPAKAI HANYA JIKA BENAR-BENAR YAKIN DATA DI EXCEL BERSIH <<<<
    #
    # try:
    #     WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "btn-submit"))).click()
    #     print("→ Clicked btn-submit")
    # except Exception as e:
    #     print("❌ FAILED to click btn-submit:", e)

    # print("✅ Save FORM: DONE.")
    #
    #
    #
    #
    #
    #
    # >>>> OPSI 2 : tunggu user ENTER di Keyboard <<<<<
    #
    # intentionally do NOT call driver.quit()
    # PAUSE: tunggu Anda klik SAVE manual di browser lalu tekan ENTER di terminal
    input("👉 Klik tombol SAVE di browser, lalu tekan ENTER di terminal untuk lanjut...")
    #
    #
    #
    #
    #
    #
    # >>>> OPSI 3 : tidak perlu tekan ENTER di Keyboard <<<<<
    # >>>> setelah user klik tombol Submit, barulah script lanjut secara otomatis <<<<
    #
    # Tunggu sampai JavaScript var xSaved berisi "Data Berhasil Disimpan"
    # print("⌛ Waiting for SUBMIT button is clicked (dan xSaved = 'Data Berhasil Disimpan')...")
    # while True:
    #     try:
    #         xSaved_value = driver.execute_script("return window.xSaved;")
    #         if xSaved_value == "Data Berhasil Disimpan":
    #             print("✅ Saving succeessfully! Continue automatically...")
    #             break
    #     except:
    #         pass
    #     time.sleep(0.5)
    #
    #
    # Loop ke file openpasien.py
    jawaban = input("➡️  Continue to FORM INFORMASI PASIEN? (Y/N): ").strip().lower()

    if jawaban == 'y':
        print("▶️  Continue to next form...")
        # lanjutkan script
    elif jawaban == 'n':
        print("⏹️ Process stop by user.")
        exit()
    else:
        print("⚠️  Unrecognized. Please tap Y or N on your keyboard.")

    # >>>> Lanjut ke file openpasien.py secara otomatis <<<<
    subprocess.run(["python", "formpasien.py"])

except KeyboardInterrupt:
    print("\n⏹️ Kill process by user.")

except Exception as e:
    print(f"\n❌ An unhandled exception occurred: {e}")
