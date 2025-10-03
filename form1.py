import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, date, time as dtime
from openpyxl import load_workbook
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException
import time
import subprocess

try: 
    # --- file: form1.py 
    # Setup Brave (attach ke existing session)
    #
    options = Options()
    options.debugger_address = "127.0.0.1:9222"  # koneksi ke Brave yang sudah terbuka
    service = Service(r"chromedriver.exe")
    driver = webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 60)

    # --- Load nomor inklusi dari Excel ---
    wb = load_workbook('data.xlsx')
    ws = wb.active

    val_no_inklusi = ws['K3'].value


    # === LOAD DATA DARI EXCEL ===
    val_puskesmas = ws['B3'].value  # puskesmas
    val_inisial = ws['C3'].value  # inisial_nama
    val_tgllahir = ws['D3'].value  # tanggalLahir_s
    val_jeniskelamin = ws['E3'].value  # jenis_kelamin (value radio: 1/2)
    val_tglscreening = ws['F3'].value  # tanggalScreening_s
    val_no_inklusi = ws['K3'].value

    print(f"📘 No. inklusi from Excel: {val_no_inklusi}")
    print("Data from Excel:", val_puskesmas, val_inisial, val_tgllahir, val_jeniskelamin, val_tglscreening)

    # === Mapping bulan Indonesia ===
    bulan_id = {
        1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
        5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
        9: "September", 10: "Oktober", 11: "November", 12: "Desember"
    }

    # === Fungsi format tanggal ke dd MMMM yyyy (Indonesia) ===
    def format_tanggal(tgl_excel):
        if isinstance(tgl_excel, datetime):
            hari = tgl_excel.day
            bulan = bulan_id[tgl_excel.month]
            tahun = tgl_excel.year
            return f"{hari:02d} {bulan} {tahun}"
        else:
            return str(tgl_excel)  # fallback kalau bukan datetime

    # === Format tanggal lahir dan screening ===
    tgl_lahir = format_tanggal(val_tgllahir)
    tgl_screening = format_tanggal(val_tglscreening)

    # === ISI DROPDOWN PUSKESMAS ===
    select_puskesmas = Select(wait.until(EC.presence_of_element_located((By.ID, "puskesmas"))))
    select_puskesmas.select_by_visible_text(val_puskesmas)
    print("→ puskesmas =", val_puskesmas)


    # === ISI INISIAL NAMA ===
    inisial_input = driver.find_element(By.ID, "inisial_nama")
    inisial_input.clear()
    inisial_input.send_keys(val_inisial)
    print("→ inisial_nama =", val_inisial)

    # === ISI TANGGAL LAHIR ===
    tgl_lahir_input = driver.find_element(By.ID, "tanggalLahir_s")
    driver.execute_script("arguments[0].removeAttribute('readonly')", tgl_lahir_input)
    tgl_lahir_input.clear()
    tgl_lahir_input.send_keys(tgl_lahir + Keys.TAB)
    print("→ tanggalLahir_s =", tgl_lahir)

    time.sleep(1)  # beri waktu datepicker tertutup

    # === PILIH RADIO BUTTON (1 = Laki-laki, 2 = Perempuan) ===
    radio_buttons = driver.find_elements(By.NAME, "jenis_kelamin")
    found = False
    for radio in radio_buttons:
        if radio.get_attribute("value") == str(val_jeniskelamin):
            driver.execute_script("arguments[0].click();", radio)
            found = True
            break
            print("→ jenis_kelamin =", val_jeniskelamin)
    if not found:
        raise Exception(f"Radio button dengan value '{val_jeniskelamin}' tidak ditemukan.")


    # === ISI TANGGAL SCREENING ===
    tgl_screening_input = driver.find_element(By.ID, "tanggalScreening_s")
    driver.execute_script("arguments[0].removeAttribute('readonly')", tgl_screening_input)
    tgl_screening_input.clear()
    tgl_screening_input.send_keys(tgl_screening + Keys.TAB)
    time.sleep(1)  # beri waktu datepicker tertutup
    print("→ tanggalScreening_s =", tgl_screening)

    print("✅ Copy-paste DATA RESPONDEN: DONE.")

    #
    #
    # >>>> OPSI 1 : otomatis save, lanjut ke script berikutnya <<<<<
    #
    # try:
    #     WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "btn-add-respondent"))).click()
    #     print("→ Clicked btn-add-respondent")
    # except Exception as e:
    #     print("❌ FAILED to click btn-add-respondent:", e)

    # print("✅ Save FORM: DONE.")
    #
    #
    #
    #
    #
    #
    # >>>> OPSI 2 : tunggu user ENTER di Keyboard <<<<<
    #
    # intentionally do NOT call driver.quit()
    # PAUSE: tunggu Anda klik SAVE manual di browser lalu tekan ENTER di terminal
    # input("👉 Click button SIMPAN at browser, then tap ENTER at terminal to continue...")
    #
    #
    #
    #
    #
    #
    # >>>> OPSI 3 : tidak perlu tekan ENTER di Keyboard <<<<<
    # >>>> setelah user klik tombol Submit/Save, lalu klik manual tombol Isi <<<<
    #
    btn_simpan = WebDriverWait(driver, 15).until(
        EC.element_to_be_clickable((By.ID, "btn-add-respondent"))
    )
    btn_simpan.click()
    time.sleep(3)

    try:
        # Tunggu sampai elemen dengan class toast-success muncul dan mengandung teks tertentu
        WebDriverWait(driver, 15).until(
            EC.text_to_be_present_in_element(
                (By.CLASS_NAME, "toast-success"),
                "Tambah Responden Berhasil"
            )
        )
        print("✅ Success notification is loaded: Tambah Responden Berhasil.")
    except:
        print("❌ Success notification is not loaded in time frame.")

    #
    #

    # Loop ke file openform2.py
    jawaban = input("➡️  Continue to FORM INKLUSI/EKSKLUSI? (Y/N): ").strip().lower()

    if jawaban == 'y':
        print("▶️  Continue to next form...")
        # lanjutkan script
    elif jawaban == 'n':
        print("⏹️ Process stop by user.")
        exit()
    else:
        print("⚠️  Unrecognized. Please tap Y or N on your keyboard.")

    # >>>> Lanjut ke openform2.py secara otomatis <<<<
    subprocess.run(["python", "openform2.py"])

except KeyboardInterrupt:
    print("\n⏹️ Kill process by user.")

except Exception as e:
    print(f"\n❌ An unhandled exception occurred: {e}")
