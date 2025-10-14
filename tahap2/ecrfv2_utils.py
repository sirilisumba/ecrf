import openpyxl
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import time
import datetime
import winsound

EXCEL_PATH = "data2.xlsx"
EXCEL_ROW = 3

def create_driver(debugger_address="127.0.0.1:9222", chromedriver_path="chromedriver.exe", wait_long=30, wait_short=10):
    options = Options()
    options.debugger_address = debugger_address
    service = Service(chromedriver_path)
    driver = webdriver.Chrome(service=service, options=options)
    
    wait_long_obj = WebDriverWait(driver, wait_long)
    wait_short_obj = WebDriverWait(driver, wait_short)
    
    return driver, wait_long_obj, wait_short_obj

def is_row_empty(ws, row_num):
    for cell in ws[row_num]:
        if cell.value is not None and str(cell.value).strip() != "":
            return False
    return True

def play_sound():
    duration = 500  # ms
    freq = 1000  # Hz
    for _ in range(3):
        winsound.Beep(freq, duration)
        time.sleep(0.2)

def load_excel_data(path=EXCEL_PATH, sheet_name=None, row_num=EXCEL_ROW):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    if is_row_empty(ws, row_num):
        play_sound()
        raise Exception("Baris 3 KOSONG, copy data dulu!")

    data = {
        "val_puskesmas"             : ws['B3'].value,
        "val_inisial"               : ws['C3'].value,
        "val_tgllahir"              : ws['D3'].value,
        "val_jeniskelamin"          : ws['E3'].value,
        "val_tglscreening"          : ws['F3'].value,
        "val_radio1"                : ws["G3"].value,
        "val_radio2"                : ws["H3"].value,
        "val_radio3"                : ws["I3"].value,
        "val_tgl_inklusi"           : ws["J3"].value,
        "val_no_inklusi"            : ws["K3"].value,
        "val_keterangan"            : ws["L3"].value,
        "val_provinsi"              : ws["M3"].value,
        "val_tgl_lapor"             : ws["N3"].value,
        "val_hasPengobatan"         : ws["O3"].value,
        "val_usia_thn"              : ws["P3"].value,
        "val_usia_bln"              : ws["Q3"].value,
        "val_usia_hr"               : ws["R3"].value,
        "val_jenis_vaksin"          : ws["S3"].value,
        "val_manufaktur"            : ws["T3"].value,
        "val_no_batch"              : ws["U3"].value,
        "val_dosis"                 : ws["V3"].value,
        "val_tgl_vaksin"            : ws["W3"].value,
        "val_wkt_vaksin"            : ws["X3"].value,
        "val_tempat_vaksin"         : ws["Y3"].value,
        "val_vaksin_lain"           : ws["Z3"].value,
        "val_vaksin_lain1"          : ws["AA3"].value,
        "val_vaksin_lain2"          : ws["AB3"].value,
        "val_vaksin_lain3"          : ws["AC3"].value,
        "val_haspengobatan"         : ws["O3"].value,
        "val_kategori"              : ws["AD3"].value,
        "val_lokal"                 : ws["AE3"].value,
        "val_nyeri"                 : ws["AF3"].value,
        "val_merah"                 : ws["AG3"].value,
        "val_tebal"                 : ws["AH3"].value,
        "val_bengkak"               : ws["AI3"].value,
        "val_lokal_lain"            : ws["AJ3"].value,
        "val_lokal_lain1"           : ws["AK3"].value,
        "val_lokal_lain1_nama"      : ws["AL3"].value,
        "val_lokal_lain2"           : ws["AM3"].value,
        "val_lokal_lain2_nama"      : ws["AN3"].value,
        "val_lokal_lain3"           : ws["AO3"].value,
        "val_lokal_lain3_nama"      : ws["AP3"].value,
        "val_lokal_lain4"           : ws["AQ3"].value,
        "val_lokal_lain4_nama"      : ws["AR3"].value,
        "val_sistemik"              : ws["AS3"].value,
        "val_demam"                 : ws["AT3"].value,
        "val_rewel"                 : ws["AU3"].value,
        "val_nangis"                : ws["AV3"].value,
        "val_sistemik_lain"         : ws["AW3"].value,
        "val_sistemik_lain_1"       : ws["AX3"].value,
        "val_sistemik_lain_1_nama"  : ws["AY3"].value,
        "val_sistemik_lain_2"       : ws["AZ3"].value,
        "val_sistemik_lain_2_nama"  : ws["BA3"].value,
        "val_sistemik_lain_3"       : ws["BB3"].value,
        "val_sistemik_lain_3_nama"  : ws["BC3"].value,
        "val_sistemik_lain_4"       : ws["BD3"].value,
        "val_sistemik_lain_4_nama"  : ws["BE3"].value,
        "val_kondisi_akhir"         : ws["BF3"].value,
    }
    return data

# def isi_dropdown(driver, wait_short, element_id, option_text):
#     try:
#         # Tunggu sampai dropdown bisa diklik
#         dropdown_elem = wait_short.until(EC.element_to_be_clickable((By.ID, element_id)))
#         select_obj = Select(dropdown_elem)

#         # Pilih opsi berdasarkan teks
#         select_obj.select_by_visible_text(option_text)
#         print(f"→ Dropdown '{element_id}' diisi dengan '{option_text}'")

#     except Exception as e:
#         raise Exception(f"❌ GAGAL isi dropdown '{element_id}': {e}")

def isi_dropdown(driver, wait_short, element_id, option_text, timeout=5, delay=0.3):
    """Isi dropdown <select> biasa secara aman & tahan error."""
    if not option_text or str(option_text).strip() == "":
        print(f"⚠️  Lewati dropdown '{element_id}': value kosong.")
        return False

    try:
        elem = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.ID, element_id))
        )
        select_obj = Select(elem)

        # Ambil semua opsi untuk pencocokan fleksibel (case-insensitive)
        all_options = [opt.text.strip() for opt in select_obj.options]
        match = next((opt for opt in all_options if opt.lower() == option_text.lower()), None)

        if not match:
            print(f"❌ Opsi '{option_text}' tidak ditemukan di dropdown '{element_id}'.")
            print(f"   Pilihan tersedia: {all_options}")
            return False

        select_obj.select_by_visible_text(match)
        time.sleep(delay)

        # Verifikasi hasil
        selected = select_obj.first_selected_option.text.strip()
        if selected.lower() == match.lower():
            print(f"✅ Dropdown '{element_id}' diisi dengan '{selected}'")
            return True
        else:
            print(f"⚠️  Verifikasi gagal: '{selected}' bukan '{option_text}'")
            return False

    except Exception as e:
        print(f"❌ GAGAL isi dropdown '{element_id}': {e}")
        return False


# def save_form1(driver, button_id, toast_success_keyword=None, timeout_click=3, timeout_toast=5, delay_after_click=1):
#     try:
#         print(f"⏳ Menunggu tombol '{button_id}' bisa diklik...")
#         btn = WebDriverWait(driver, timeout_click).until(
#             EC.element_to_be_clickable((By.ID, button_id))
#         )

#         driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
#         time.sleep(0.3)
#         try:
#             btn.click()
#         except Exception:
#             print("⚠️ Klik biasa gagal, coba klik lewat JavaScript")
#             driver.execute_script("arguments[0].click();", btn)

#         print(f"🖱️ Tombol '{button_id}' diklik.")
#         time.sleep(delay_after_click)

#         if toast_success_keyword:
#             toast = WebDriverWait(driver, timeout_toast).until(
#                 EC.visibility_of_element_located((By.CLASS_NAME, "toast-success"))
#             )
#             toast_text = toast.text.strip()
#             if toast_success_keyword.lower() not in toast_text.lower():
#                 raise Exception(f"Isi toast-success tidak sesuai: '{toast_text}'")
#             print(f"✅ Form disimpan: {toast_text}")
#         else:
#             print("✅ Form disimpan (tanpa validasi toast).")

#     except TimeoutException:
#         play_sound()
#         raise Exception(f"❌ Timeout: Tombol '{button_id}' atau toast-success tidak muncul.")
#     except Exception as e:
#         play_sound()
#         raise Exception(f"❌ Gagal menyimpan form '{button_id}': {e}")

def save_form1(driver, button_id, toast_success_keyword=None, timeout_click=2, timeout_toast=3, delay_after_click=0.5):
    try:
        print(f"⏳ Menunggu tombol '{button_id}' bisa diklik...")
        btn = WebDriverWait(driver, timeout_click).until(
            EC.element_to_be_clickable((By.ID, button_id))
        )

        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
        time.sleep(0.2)
        try:
            btn.click()
        except Exception:
            print("⚠️ Klik biasa gagal, coba klik lewat JavaScript")
            driver.execute_script("arguments[0].click();", btn)

        print(f"🖱️ Tombol '{button_id}' diklik.")
        time.sleep(delay_after_click)

        if toast_success_keyword:
            toast = WebDriverWait(driver, timeout_toast).until(
                EC.visibility_of_element_located((By.CLASS_NAME, "toast-success"))
            )
            toast_text = toast.text.strip()
            if toast_success_keyword.lower() not in toast_text.lower():
                raise Exception(f"Isi toast-success tidak sesuai: '{toast_text}'")
            print(f"✅ Form disimpan: {toast_text}")
        else:
            print("✅ Form disimpan (tanpa validasi toast).")

    except TimeoutException:
        play_sound()
        raise Exception(f"❌ Timeout: Tombol '{button_id}' atau toast-success tidak muncul.")
    except Exception as e:
        play_sound()
        raise Exception(f"❌ Gagal menyimpan form '{button_id}': {e}")


# def save_form(driver, button_id, toast_success_keyword=None, timeout=5):
#     try:
#         print(f"⏳ Menunggu tombol '{button_id}' bisa diklik...")
#         btn = WebDriverWait(driver, timeout).until(
#             EC.element_to_be_clickable((By.ID, button_id))
#         )
#         driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
#         time.sleep(0.2)
#         try:
#             btn.click()
#         except Exception:
#             print("⚠️ Klik biasa gagal, coba klik lewat JavaScript")
#             driver.execute_script("arguments[0].click();", btn)
#         print(f"🖱️ Tombol '{button_id}' diklik.")

#         if toast_success_keyword:
#             toast = WebDriverWait(driver, timeout).until(
#                 EC.visibility_of_element_located((By.CLASS_NAME, "toast-success"))
#             )
#             time.sleep(0.3)
#             toast_text = toast.text.strip()
#             if toast_success_keyword.lower() not in toast_text.lower():
#                 raise Exception(f"Isi toast-success tidak sesuai: '{toast_text}'")
#             print(f"✅ Form disimpan: {toast_text}")
#         else:
#             print(f"✅ Form disimpan (tanpa validasi toast).")

#     except TimeoutException as e:
#         raise Exception(f"❌ Timeout: elemen tidak muncul ({e})")
#     except Exception as e:
#         raise Exception(f"❌ Gagal menyimpan form '{button_id}': {e}")

def save_form(driver, button_id, toast_success_keyword=None, timeout_click=3, timeout_toast=5):
    try:
        print(f"⏳ Menunggu tombol '{button_id}' bisa diklik...")
        btn = WebDriverWait(driver, timeout_click).until(
            EC.element_to_be_clickable((By.ID, button_id))
        )

        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
        time.sleep(0.2)

        try:
            btn.click()
        except Exception:
            print("⚠️ Klik biasa gagal, coba lewat JavaScript")
            driver.execute_script("arguments[0].click();", btn)

        print(f"🖱️ Tombol '{button_id}' diklik.")

        if toast_success_keyword:
            # Tunggu toast muncul, tapi tidak terlalu lama
            toast = WebDriverWait(driver, timeout_toast).until(
                EC.visibility_of_element_located((By.CLASS_NAME, "toast-success"))
            )
            toast_text = toast.text.strip()
            if toast_success_keyword.lower() not in toast_text.lower():
                raise Exception(f"Isi toast tidak sesuai: '{toast_text}'")
            print(f"✅ Form disimpan: {toast_text}")
        else:
            print("✅ Form disimpan (tanpa validasi toast).")

    except TimeoutException:
        raise Exception(f"⏰ Timeout: Tombol '{button_id}' atau toast-success tidak muncul tepat waktu.")
    except Exception as e:
        raise Exception(f"❌ Gagal menyimpan form '{button_id}': {e}")


# def set_text(driver, id_, value, timeout=1.5):
#         if value is None or str(value).strip() == "":
#             return False
#         try:
#             el = WebDriverWait(driver, timeout).until(
#                 EC.element_to_be_clickable((By.ID, id_))
#             )
#         except TimeoutException:
#             print(f"❌ FAILED element {id_} not found (timeout after {timeout}s)")
#             return False
#         try:
#             el.clear()
#             el.send_keys(str(value))
#             print(f"→ {id_} = {value}")
#             return True
#         except Exception as e:
#             print(f"❌ FAILED set_text {id_}: {e}")
#             return False

def set_text(driver, id_, value, timeout=1.5):
    if value is None or str(value).strip() == "":
        return False

    try:
        # Tunggu sampai elemen terlihat dan bisa diklik
        el = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.ID, id_))
        )

        # Scroll agar elemen terlihat (menghindari intercept)
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
        time.sleep(0.1)

        # Pastikan tidak readonly atau disabled
        driver.execute_script(
            "arguments[0].removeAttribute('readonly'); arguments[0].removeAttribute('disabled');",
            el
        )

        # Bersihkan isi lama dan ketik ulang
        el.clear()
        el.send_keys(str(value))
        print(f"→ {id_} = {value}")
        return True

    except TimeoutException:
        print(f"❌ Timeout: elemen '{id_}' tidak ditemukan dalam {timeout}s.")
        return False

    except Exception as e:
        # Coba klik via JavaScript kalau send_keys gagal karena overlay
        try:
            print(f"⚠️ Gagal input langsung ke '{id_}' ({e}), coba pakai JS...")
            driver.execute_script(
                f"document.getElementById('{id_}').value = arguments[0];", str(value)
            )
            print(f"✅ JS set value '{id_}' = {value}")
            return True
        except Exception as e2:
            print(f"❌ Gagal set_text '{id_}' via JS juga: {e2}")
            return False

def format_tanggal_ddmmyyyy(date_obj):
    if not date_obj:
        return ""
    if isinstance(date_obj, datetime.datetime):
        return date_obj.strftime("%d-%m-%Y")
    elif isinstance(date_obj, str):
        return date_obj 
    try:
        return date_obj.strftime("%d-%m-%Y")
    except Exception:
        return str(date_obj)
    
# def isi_datepicker(driver, wait_short, field_id, tanggal_obj, timeout=15, delay=0.2):
#     if not tanggal_obj:
#         print(f"⚠️  Skip field {field_id} karena tanggal kosong.")
#         return
#     try:
#         formatted = format_tanggal_ddmmyyyy(tanggal_obj)
#         field = wait_short.until(EC.element_to_be_clickable((By.ID, field_id)))
#         driver.execute_script(
#             "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input'));",
#             field, formatted
#         )
#         time.sleep(delay)
#         field.send_keys(Keys.TAB)
#         print(f"→ {field_id} = {formatted}")
        
#     except Exception as e:
#         print(f"❌ GAGAL set {field_id}: {e}")

def isi_datepicker(driver, wait_short, field_id, tanggal_obj, timeout=3, delay=0.2):
    if not tanggal_obj:
        print(f"⚠️ Skip {field_id}: tanggal kosong.")
        return False

    formatted = format_tanggal_ddmmyyyy(tanggal_obj)

    try:
        field = wait_short.until(EC.element_to_be_clickable((By.ID, field_id)))

        # Scroll biar visible
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", field)
        time.sleep(0.1)

        # Pastikan field aktif dan tidak readonly
        driver.execute_script(
            "arguments[0].removeAttribute('readonly'); arguments[0].removeAttribute('disabled');",
            field
        )

        # Bersihkan isi lama dan isi ulang via JS (lebih stabil dari send_keys langsung)
        driver.execute_script("arguments[0].value = '';", field)
        driver.execute_script(
            "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input'));",
            field, formatted
        )
        time.sleep(delay)

        # Trigger blur agar datepicker menutup dan nilai tersimpan
        field.send_keys(Keys.TAB)
        time.sleep(0.2)

        # Validasi ulang: ambil value di DOM, pastikan sesuai
        value_now = driver.execute_script("return arguments[0].value;", field)
        if value_now.strip() != formatted:
            # Coba set ulang sekali lagi (kadang datepicker override saat blur)
            driver.execute_script(
                "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input'));",
                field, formatted
            )
            field.send_keys(Keys.TAB)
            time.sleep(0.2)
            value_now = driver.execute_script("return arguments[0].value;", field)

        if value_now.strip() == formatted:
            print(f"→ {field_id} = {formatted}")
            return True
        else:
            print(f"⚠️ {field_id}: nilai belum tersimpan dengan benar (current={value_now})")
            return False

    except TimeoutException:
        print(f"❌ Timeout: elemen {field_id} tidak ditemukan dalam {timeout}s.")
        return False
    except Exception as e:
        print(f"❌ GAGAL set {field_id}: {e}")
        return False


# def isi_time(driver, wait_short, field_id, time_obj, timeout=15, delay=0.2):
#     if not time_obj:
#         print(f"⚠️  Skip field {field_id} karena waktu kosong.")
#         return
#     try:
#         # Format waktu 24 jam: "HH:mm"
#         formatted = time_obj.strftime("%H:%M")

#         field = wait_short.until(EC.element_to_be_clickable((By.ID, field_id)))
#         driver.execute_script(
#             "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input'));",
#             field, formatted
#         )
#         time.sleep(delay)
#         field.send_keys(Keys.TAB)  # supaya trigger event, tutup timepicker, dsb

#         print(f"🕒 Field '{field_id}' diisi dengan waktu: {formatted}")
        
#     except Exception as e:
#         print(f"❌ GAGAL isi timepicker '{field_id}': {e}")

def isi_time(driver, wait_short, field_id, time_obj, timeout=3, delay=0.2):
    """Isi field timepicker dengan format HH:mm secara stabil dan cepat."""
    if not time_obj:
        print(f"⚠️ Skip {field_id}: waktu kosong.")
        return False

    formatted = time_obj.strftime("%H:%M")

    try:
        field = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.ID, field_id))
        )

        # Scroll agar terlihat dan pastikan tidak readonly
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", field)
        driver.execute_script(
            "arguments[0].removeAttribute('readonly'); arguments[0].removeAttribute('disabled');",
            field
        )

        # Kosongkan lalu isi nilai baru via JS agar langsung valid di front-end
        driver.execute_script("arguments[0].value = '';", field)
        driver.execute_script(
            "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input'));",
            field, formatted
        )

        time.sleep(delay)
        field.send_keys(Keys.TAB)  # tutup timepicker dan trigger blur/input

        # Validasi ulang value-nya
        current = driver.execute_script("return arguments[0].value;", field)
        if current.strip() != formatted:
            # Coba set ulang jika frontend ngereset nilainya
            driver.execute_script(
                "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input'));",
                field, formatted
            )
            field.send_keys(Keys.TAB)
            time.sleep(0.1)
            current = driver.execute_script("return arguments[0].value;", field)

        if current.strip() == formatted:
            print(f"🕒 {field_id} = {formatted}")
            return True
        else:
            print(f"⚠️ {field_id}: gagal menyimpan nilai (current={current})")
            return False

    except TimeoutException:
        print(f"❌ Timeout: elemen {field_id} tidak muncul dalam {timeout}s.")
        return False
    except Exception as e:
        print(f"❌ Gagal isi timepicker '{field_id}': {e}")
        return False


def wait_clickable(driver, by, locator, timeout=5):
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((by, locator))
        )
        return element
    except TimeoutException:
        print(f"❌ Element {locator} tidak clickable setelah {timeout} detik")
        return None
def should_check(val):
        if val is None:
            return None
        hv = str(val).strip().lower()
        return hv in ("1", "yes", "ya", "true", "x", "y")
def set_checkbox(driver, element_id, should_check_value):
    cb = wait_clickable(driver, By.ID, element_id)
    if not cb:
        print(f"! Checkbox {element_id} tidak ditemukan atau tidak clickable")
        return False
    want = should_check(should_check_value)
    if want is True and not cb.is_selected():
        driver.execute_script("arguments[0].click();", cb)
        print(f"→ {element_id} checked")
    elif want is False and cb.is_selected():
        driver.execute_script("arguments[0].click();", cb)
        print(f"→ {element_id} unchecked")
    else:
        print(f"→ {element_id} left as-is (matches desired state)")
    return True

def buka_form(driver, wait, nomor_inklusi, spinner_class="spinner", chevron_class="fa-chevron-down", timeout_click=1):
    try:
        print("⏳ ...menunggu data dimuat dan spinner hilang...")
        # wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, spinner_class)))
        wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".spinner.loading")))
        print("✅ Loading is done.")
    except TimeoutException:
        print("⚠️ Timeout loading selesai. Continue...")

    try:
        print(f"🔍 Mencari nomor_inklusi di tabel: '{nomor_inklusi}'")
        xpath_tr = f"//tr[td[text()='{nomor_inklusi}']]"
        tr_elem = wait.until(EC.presence_of_element_located((By.XPATH, xpath_tr)))
        print("✅ Baris <TR> ditemukan berdasarkan nomor inklusi.")

        # Cari ikon chevron di dalam kolom terakhir
        chevron_xpath = f".//td[last()]//i[contains(@class, '{chevron_class}')]"
        chevron = tr_elem.find_element(By.XPATH, chevron_xpath)
        print(f"🔽 Chevron '{chevron_class}' ditemukan, coba diklik...")

        # Tunggu sampai chevron bisa diklik
        wait.until(EC.element_to_be_clickable((By.XPATH, chevron_xpath)))

        # Scroll ke chevron dan klik
        driver.execute_script("arguments[0].scrollIntoView(true);", chevron)
        time.sleep(0.3)
        driver.execute_script("arguments[0].click();", chevron)
        print("✅ Chevron diklik untuk expand.")

        time.sleep(timeout_click)  # beri waktu agar konten dimuat

    except Exception as e:
        print("❌ GAGAL menemukan atau klik chevron berdasarkan nomor inklusi.")
        print(e)
        return False

    try:
        tombol_isi_xpath = "//tr[.//td[contains(., 'Informasi Pasien')]]//button[contains(., 'ISI')]"
        tombol_isi = wait.until(EC.presence_of_element_located((By.XPATH, tombol_isi_xpath)))
        print("✅ Tombol 'ISI' ketemu")
        print("   OnClick attr:", tombol_isi.get_attribute("onclick"))

        driver.execute_script("arguments[0].scrollIntoView(true);", tombol_isi)
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", tombol_isi)
        print("🖱️ Tombol 'ISI' dklik.")

        return True

    except Exception as e:
        print("❌ GAGAL menemukan atau klik tombol 'ISI'.")
        print(e)
        return False

# def set_radio(driver, group_id, value, timeout=10):
#     try:
#         if value is None or str(value).strip() == "":
#             print(f"⚠️  Lewati set_radio: value kosong untuk group {group_id}")
#             return False

#         css = f"#{group_id} input[type='radio'][value='{value}']"
#         el = WebDriverWait(driver, timeout).until(
#             EC.element_to_be_clickable((By.CSS_SELECTOR, css))
#         )

#         driver.execute_script("arguments[0].click();", el)
#         print(f"✅ {group_id} = {value}")
#         return True

#     except Exception as e:
#         print(f"❌ GAGAL memilih radio '{group_id}' = '{value}': {e}")
#         return False

def set_radio(driver, group_id, value, timeout=3):
    """Pilih radio button dengan ID grup tertentu dan value sesuai."""
    if value is None or str(value).strip() == "":
        print(f"⚠️ Lewati set_radio: value kosong untuk group {group_id}")
        return False

    try:
        css_selector = f"#{group_id} input[type='radio'][value='{value}']"
        el = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located(("css selector", css_selector))
        )

        # Scroll agar terlihat dan hilangkan kemungkinan terblokir
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
        driver.execute_script(
            "arguments[0].removeAttribute('disabled'); arguments[0].removeAttribute('readonly');",
            el
        )

        # Klik lewat JS untuk hindari overlay atau label yang menutup
        driver.execute_script("arguments[0].click();", el)

        # Validasi hasil (apakah radio benar-benar terpilih)
        checked = driver.execute_script("return arguments[0].checked;", el)
        if checked:
            print(f"✅ {group_id} = {value}")
            return True
        else:
            print(f"⚠️ {group_id}: klik tidak berhasil, radio belum terpilih.")
            return False

    except TimeoutException:
        print(f"❌ Timeout: radio {group_id} dengan value '{value}' tidak ditemukan dalam {timeout}s.")
        return False
    except Exception as e:
        print(f"❌ GAGAL memilih radio '{group_id}' = '{value}': {e}")
        return False


# def isi_radio_button(driver, group_name, value):
#     radio_buttons = driver.find_elements(By.NAME, group_name)
#     found = False
#     for radio in radio_buttons:
#         if radio.get_attribute("value") == str(value):
#             driver.execute_script("arguments[0].click();", radio)
#             print(f"→ {group_name} = {value}")
#             found = True
#             break
#     if not found:
#         raise Exception(f"❌ Radio button dengan value '{value}' di group '{group_name}' tidak ditemukan.")

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

def isi_radio_button(driver, group_name, value, timeout=3):
    """Isi radio button berdasarkan atribut 'name' dan value target."""
    try:
        # Tunggu sampai minimal satu radio muncul
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.NAME, group_name))
        )
        radio_buttons = driver.find_elements(By.NAME, group_name)
        found = False

        for radio in radio_buttons:
            if radio.get_attribute("value") == str(value):
                # Scroll agar terlihat
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", radio)
                # Pastikan tidak disabled
                driver.execute_script(
                    "arguments[0].removeAttribute('disabled'); arguments[0].removeAttribute('readonly');",
                    radio
                )
                # Klik via JS (lebih aman)
                driver.execute_script("arguments[0].click();", radio)

                # Validasi hasil klik
                checked = driver.execute_script("return arguments[0].checked;", radio)
                if checked:
                    print(f"✅ {group_name} = {value}")
                    found = True
                    break
                else:
                    print(f"⚠️ Klik pada {group_name} value='{value}' tidak berhasil, radio belum terpilih.")

        if not found:
            raise Exception(f"❌ Radio button dengan value '{value}' di group '{group_name}' tidak ditemukan atau gagal diklik.")
        return True

    except TimeoutException:
        print(f"❌ Timeout: radio group '{group_name}' belum muncul setelah {timeout}s.")
        return False
    except Exception as e:
        print(f"❌ GAGAL isi radio '{group_name}' = '{value}': {e}")
        return False


# Mapping bulan Indonesia
bulan_id = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
    5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
    9: "September", 10: "Oktober", 11: "November", 12: "Desember"
}
def format_tanggal_indo(tgl_excel):
    if isinstance(tgl_excel, datetime.datetime):
        hari = tgl_excel.day
        bulan = bulan_id[tgl_excel.month]
        tahun = tgl_excel.year
        return f"{hari:02d} {bulan} {tahun}"
    else:
        return str(tgl_excel)
# def isi_date_indo(driver, wait_short, element_id, tgl_excel):
#     tanggal_str = format_tanggal_indo(tgl_excel)
#     try:
#         input_elem = wait_short.until(EC.element_to_be_clickable((By.ID, element_id)))
#         driver.execute_script("arguments[0].removeAttribute('readonly')", input_elem)
#         input_elem.clear()
#         input_elem.send_keys(tanggal_str + Keys.TAB)
#         time.sleep(1)
#         print(f"→ {element_id} = {tanggal_str}")
#         return True
#     except Exception as e:
#         print(f"❌ GAGAL mengisi datepicker {element_id}: {e}")
#         return False
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

def isi_date_indo(driver, wait_short, element_id, tgl_excel, delay=0.3, timeout=5):
    """Isi input tanggal format Indonesia (dd-mm-yyyy) secara aman & stabil."""
    if not tgl_excel:
        print(f"⚠️  Lewati {element_id}: tanggal kosong.")
        return False

    try:
        tanggal_str = format_tanggal_indo(tgl_excel)
    except Exception as e:
        print(f"❌ Format tanggal tidak valid untuk {element_id}: {e}")
        return False

    try:
        field = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.ID, element_id))
        )

        # Pastikan input siap diisi
        driver.execute_script("""
            arguments[0].removeAttribute('readonly');
            arguments[0].removeAttribute('disabled');
            arguments[0].scrollIntoView({block: 'center'});
        """, field)

        # Clear dengan cara lebih aman (pakai JS untuk framework tertentu)
        driver.execute_script("arguments[0].value = '';", field)
        time.sleep(delay)

        # Isi nilai baru
        field.send_keys(tanggal_str)
        field.send_keys(Keys.TAB)
        time.sleep(delay)

        # Verifikasi hasil
        current_val = field.get_attribute("value")
        if tanggal_str in current_val:
            print(f"✅ {element_id} = {tanggal_str}")
            return True
        else:
            print(f"⚠️  Gagal verifikasi isi {element_id}: value sekarang '{current_val}'")
            return False

    except Exception as e:
        print(f"❌ GAGAL mengisi datepicker {element_id}: {e}")
        return False


