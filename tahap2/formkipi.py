import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, date, time as dtime
from openpyxl import load_workbook
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
import time
import subprocess

try: 

    # --- file formkipi.py
    # --- Setup Selenium untuk Brave ---
    options = Options()
    options.debugger_address = "127.0.0.1:9222"
    service = Service("chromedriver.exe")
    driver = webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 20)

    # --- Load nomor inklusi dari Excel ---
    wb = openpyxl.load_workbook('data2.xlsx')
    ws = wb.active
    val_no_inklusi = str(ws['K3'].value).strip()

    # ---------- Helpers ----------
    def format_tanggal_ddmmyyyy(value):
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return value.strftime("%d-%m-%Y")
        s = str(value).strip()
        if not s:
            return ""
        candidates = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d %B %Y", "%d %b %Y"]
        for fmt in candidates:
            try:
                return datetime.strptime(s, fmt).strftime("%d-%m-%Y")
            except:
                pass
        return s

    def format_time_hhmm(value):
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%H:%M")
        if isinstance(value, dtime):
            return value.strftime("%H:%M")
        s = str(value).strip()
        if not s:
            return ""
        candidates = ["%H:%M", "%H.%M", "%I:%M %p", "%H%M"]
        for fmt in candidates:
            try:
                return datetime.strptime(s, fmt).strftime("%H:%M")
            except:
                pass
        return s

    def should_check(val):
        if val is None:
            return None
        hv = str(val).strip().lower()
        return hv in ("1", "yes", "ya", "true", "x", "y")

    def safe_find(driver, locator, by=By.ID, timeout=3):
        try:
            return WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((by, locator))
            )
        except:
            return None

    def set_text(driver, id_, value, timeout=3):
        if value is None or str(value).strip() == "":
            return False
        try:
            el = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((By.ID, id_))
            )
        except TimeoutException:
            print(f"❌ FAILED element {id_} not found (timeout after {timeout}s)")
            return False
        try:
            el.clear()
            el.send_keys(str(value))
            print(f"→ {id_} = {value}")
            return True
        except Exception as e:
            print(f"❌ FAILED set_text {id_}: {e}")
            return False

    def set_radio(driver, group_id, value):
        if value is None or str(value).strip() == "":
            return False
        css = f"#{group_id} input[type='radio'][value='{value}']"
        el = safe_find(driver, css, By.CSS_SELECTOR, timeout=2)
        if not el:
            print(f"❌ FAILED radio {group_id} value {value} not found")
            return False
        try:
            driver.execute_script("arguments[0].click();", el)
            print(f"→ {group_id} = {value}")
            return True
        except Exception as e:
            print(f"❌ FAILED set_radio {group_id}: {e}")
            return False

    def set_date(driver, id_, value):
        if value is None or str(value).strip() == "":
            return False
        formatted = format_tanggal_ddmmyyyy(value)
        el = safe_find(driver, id_, By.ID, timeout=3)
        if not el:
            print(f"❌ FAILED date element {id_} not found")
            return False
        try:
            # set via JS + dispatch events so frameworks pick it up
            driver.execute_script(
                "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input')); arguments[0].dispatchEvent(new Event('change'));",
                el, formatted
            )
            time.sleep(0.12)
            el.send_keys(Keys.TAB)  # close datepicker
            print(f"→ {id_} = {formatted}")
            return True
        except Exception as e:
            print(f"❌ FAILED set_date {id_}: {e}")
            return False

    def set_time(driver, id_, value):
        if value is None or str(value).strip() == "":
            return False
        formatted = format_time_hhmm(value)
        el = safe_find(driver, id_, By.ID, timeout=3)
        if not el:
            print(f"❌ FAILED time element {id_} not found")
            return False
        try:
            driver.execute_script(
                "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input')); arguments[0].dispatchEvent(new Event('change'));",
                el, formatted
            )
            time.sleep(0.12)
            el.send_keys(Keys.TAB)  # close timepicker
            print(f"→ {id_} = {formatted}")
            return True
        except Exception as e:
            print(f"❌ FAILED set_time {id_}: {e}")
            return False

    # ---------- Load Excel AO2..BB2 ----------
    val_no_inklusi = ws["K3"].value   # itemid_58646 (text)      >> 1. no inklusi
    val_inisial = ws["C3"].value   # itemid_58647 (text)      >> 2. inisial

    val_kategori = ws["AD3"].value   # form_group_58650 (radio) >> 3. Kategori KIPI (serius/non-serius):
    val_lokal = ws["AE3"].value   # form_group_58766 (radio) >> 3A.Reaksi Lokal:
    val_nyeri = ws["AF3"].value   # form_group_58767 (radio) >>  3A1. Nyeri lokal:
    val_merah = ws["AG3"].value   # form_group_58768 (radio) >>  3A2. Kemerahan:
    val_tebal = ws["AH3"].value   # form_group_58769 (radio) >>  3A3. Penebalan:
    val_bengkak = ws["AI3"].value   # form_group_58770 (radio) >>  3A4. Pembengkakan:
    val_lokal_lain = ws["AJ3"].value   # form_group_58782 (radio) >> 3A5. Lain-lain:
    val_lokal_lain1 = ws["AK3"].value   # form_group_59084 (radio) >>  3A51. Lain-lain 1:
    val_lokal_lain1_nama = ws["AL3"].value   # itemid_59088 (text)          >> 3A511. Lain-lain 1 (nama)
    val_lokal_lain2 = ws["AM3"].value   # form_group_59085 (radio) >>  3A52. Lain-lain 2:
    val_lokal_lain2_nama = ws["AN3"].value   # itemid_59092 (text)          >> 3A521. Lain-lain 2 (nama)
    val_lokal_lain3 = ws["AO3"].value   # form_group_59086 (radio) >>  3A53. Lain-lain 3:
    val_lokal_lain3_nama = ws["AP3"].value   # itemid_59095 (text)          >> 3A531. Lain-lain 3 (nama)
    val_lokal_lain4 = ws["AQ3"].value   # form_group_59087 (radio) >>  3A54. Lain-lain 4:
    val_lokal_lain4_nama = ws["AR3"].value   # itemid_59098 (text)          >> 3A541. Lain-lain 4 (nama)

    val_sistemik = ws["AS3"].value   # form_group_58781 (radio) >> 3B. Reaksi Sistemik:
    val_demam = ws["AT3"].value   # form_group_58795 (radio) >>  3B1. Demam
    val_rewel = ws["AU3"].value   # form_group_58796 (radio) >> 3B2. Rewel:
    val_nangis = ws["AV3"].value   # form_group_58797 (radio) >> 3B3. Nangis:
    val_sistemik_lain = ws["AW3"].value   # form_group_58807 (radio) >> 3B4. Lain-lain:
    val_sistemik_lain_1 = ws["AX3"].value   # form_group_59129 (radio) >>  3B41. Lain-lain 1:
    val_sistemik_lain_1_nama = ws["AY3"].value   # itemid_59113 (text)          >> 3B411. Lain-lain 1 (nama)
    val_sistemik_lain_2 = ws["AZ3"].value   # form_group_59130 (radio) >>  3B42. Lain-lain 2:
    val_sistemik_lain_2_nama = ws["BA3"].value   # itemid_59116 (text)          >> 3B421. Lain-lain 2 (nama)
    val_sistemik_lain_3 = ws["BB3"].value   # form_group_59131 (radio) >>  3B43. Lain-lain 3:
    val_sistemik_lain_3_nama = ws["BC3"].value   # itemid_59119 (text)          >> 3B431. Lain-lain 3 (nama)
    val_sistemik_lain_4 = ws["BD3"].value   # form_group_59132 (radio) >>  3B44. Lain-lain 4:
    val_sistemik_lain_4_nama = ws["BE3"].value   # itemid_59122 (text)          >> 3B441. Lain-lain 4 (nama)

    val_kondisi_akhir = ws["BF3"].value   # form_group_58827 (radio) >>      3C. Sembuh
    
    val_haspengobatan = ws["O3"].value  # hasPengobatan (checkbox) >> 4. hasPengobatan


    print("Start to fill-in Form KIPI")
    print(f"📘 No. inklusi from Excel: {val_no_inklusi}")
    print("Data from Excel:", val_no_inklusi, val_inisial, val_kategori, val_lokal, val_nyeri, val_merah, val_tebal, val_bengkak, 
    val_lokal_lain, val_lokal_lain1, val_lokal_lain1_nama, val_lokal_lain2, val_lokal_lain2_nama, 
    val_lokal_lain3, val_lokal_lain3_nama, val_lokal_lain4, val_lokal_lain4_nama, val_sistemik, val_demam, val_rewel, val_nangis, 
    val_sistemik_lain, val_sistemik_lain_1, val_sistemik_lain_1_nama, val_sistemik_lain_2, val_sistemik_lain_2_nama, 
    val_sistemik_lain_3, val_sistemik_lain_3_nama, val_sistemik_lain_4, val_sistemik_lain_4_nama, val_kondisi_akhir, val_haspengobatan)      

    # ---------- Fill form ----------
    # text fields
    set_text(driver, "itemid_58646", val_no_inklusi)  # no_inklusi
    set_text(driver, "itemid_58647", val_inisial)  # inisial

    # main radio group 58650 >>> Kategori KIPI Serius atau Non Serius: 
    if set_radio(driver, "form_group_58650", val_kategori):
        # only handle nested block if value == 2
        try:
            if str(val_kategori).strip() == "2": # 
                time.sleep(0.35)  # wait UI to render

                # AR2 -> form_group_58766 >>> REAKSI LOKAL
                if set_radio(driver, "form_group_58766", val_lokal):
                    if str(val_lokal).strip() == "1":
                        time.sleep(0.25)
                        # AS2 -> form_group_58767 >>> NYERI LOKAL
                        set_radio(driver, "form_group_58767", val_nyeri)
                        set_radio(driver, "form_group_58768", val_merah)
                        set_radio(driver, "form_group_58769", val_tebal)
                        set_radio(driver, "form_group_58770", val_bengkak)
                        # BI2 -> form_group_58782 >>> LAIN-LAIN
                        if set_radio(driver, "form_group_58782", val_lokal_lain):
                            if str(val_lokal_lain).strip() == "1":
                                time.sleep(0.25)
                                # BJ2 -> form_group_59084 >>> lain-lain 1
                                if set_radio(driver, "form_group_59084", val_lokal_lain1):
                                    if str(val_lokal_lain1).strip() == "1":
                                        set_text(driver, "itemid_59088", val_lokal_lain1_nama) # BK2 (nama)
                                # BO2 -> form_group_59085 >>> lain-lain 2
                                if set_radio(driver, "form_group_59085", val_lokal_lain2):
                                    if str(val_lokal_lain2).strip() == "1":
                                        set_text(driver, "itemid_59092", val_lokal_lain2_nama) # BP2 (nama)
                                # BT2 -> form_group_59086 >>> lain-lain 3
                                if set_radio(driver, "form_group_59086", val_lokal_lain3):
                                    if str(val_lokal_lain3).strip() == "1":
                                        set_text(driver, "itemid_59095", val_lokal_lain3_nama) # BU2 (nama)
                                # BY2 -> form_group_59087 >>> lain-lain 4
                                if set_radio(driver, "form_group_59087", val_lokal_lain4):
                                    if str(val_lokal_lain4).strip() == "1":
                                        set_text(driver, "itemid_59098", val_lokal_lain4_nama) # BZ2 (nama)
                # CD2 -> form_group_58781 >>> SISTEMIK
                if set_radio(driver, "form_group_58781", val_sistemik):
                    if str(val_sistemik).strip() == "1":
                        time.sleep(0.25)
                        # CE2 -> form_group_58795 >>> DEMAM
                        set_radio(driver, "form_group_58795", val_demam)
                        set_radio(driver, "form_group_58796", val_rewel)
                        set_radio(driver, "form_group_58797", val_nangis)
                        # CQ2 -> form_group_58807 >>> LAIN-LAIN
                        if set_radio(driver, "form_group_58807", val_sistemik_lain):
                            if str(val_sistemik_lain).strip() == "1":
                                time.sleep(0.25)
                                # CR2 -> form_group_59129 >>> lain-lain 1
                                if set_radio(driver, "form_group_59129", val_sistemik_lain_1):
                                    if str(val_sistemik_lain_1).strip() == "1":
                                        set_text(driver, "itemid_59113", val_sistemik_lain_1_nama) # CS2 (nama)
                                # CW2 -> form_group_59130 >>> lain-lain 2
                                if set_radio(driver, "form_group_59130", val_sistemik_lain_2):
                                    if str(val_sistemik_lain_2).strip() == "1":
                                        set_text(driver, "itemid_59116", val_sistemik_lain_2_nama) # CX2 (nama)
                                # DB2 -> form_group_59131 >>> lain-lain 3
                                if set_radio(driver, "form_group_59131", val_sistemik_lain_3):
                                    if str(val_sistemik_lain_3).strip() == "1":
                                        set_text(driver, "itemid_59119", val_sistemik_lain_3_nama) # DC2 (nama)
                                # DG2 -> form_group_59132 >>> lain-lain 4
                                if set_radio(driver, "form_group_59132", val_sistemik_lain_4):
                                    if str(val_sistemik_lain_4).strip() == "1":
                                        set_text(driver, "itemid_59122", val_sistemik_lain_4_nama) # DH2 (nama)                
                # DL2 -> form_group_58827 >> SEMBUH
                set_radio(driver, "form_group_58827", val_kondisi_akhir)
        except Exception as e:
            print("❌ FAILED error handling nested radios:", e)
    else:
        print("→ main radio form_group_58650 not set (or empty)")

    # checkbox hasPengobatan (BB2)
    cb = safe_find(driver, "hasPengobatan", By.ID, timeout=2)
    if cb:
        want = should_check(val_haspengobatan)
        if want is True and not cb.is_selected():
            driver.execute_script("arguments[0].click();", cb)
            print("→ hasPengobatan checked")
        elif want is False and cb.is_selected():
            driver.execute_script("arguments[0].click();", cb)
            print("→ hasPengobatan unchecked")
        else:
            print("→ hasPengobatan left as-is")
    else:
        print("❌ FAILED checkbox hasPengobatan not found")

    print("✅ Copy-paste KIPI: DONE.")


    #
    #
    #
    # >>>> OPSI 1 : otomatis save, lanjut ke script berikutnya <<<<<
    #
    # try:
    #     WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "btn-submit"))).click()
    #     print("→ Clicked btn-submit")
    # except Exception as e:
    #     print("❌ FAILED to click btn-submit:", e)

    # print("✅ Save FORM: DONE.")
    #
    #
    #
    #
    #
    #
    # >>>> OPSI 2 : tunggu user ENTER di Keyboard <<<<<
    #
    # intentionally do NOT call driver.quit()
    # PAUSE: tunggu Anda klik SAVE manual di browser lalu tekan ENTER di terminal
    input("👉 Klik tombol SAVE di browser, lalu tekan ENTER di terminal untuk lanjut...")
    #
    #
    #
    #
    #
    #
    # >>>> OPSI 3 : tidak perlu tekan ENTER di Keyboard <<<<<
    # >>>> setelah user klik tombol Submit, otomatis lanjut <<<<
    #
    # Tunggu sampai JavaScript var xSaved berisi "Data Berhasil Disimpan"
    # print("⌛ Waiting for SUBMIT button is clicked (dan xSaved = 'Data Berhasil Disimpan')...")
    # while True:
    #     try:
    #         xSaved_value = driver.execute_script("return window.xSaved;")
    #         if xSaved_value == "Data Berhasil Disimpan":
    #             print("✅ Saving succeessfully! Continue automatically to deleting row 2 in Excel.")
    #             break
    #     except:
    #         pass
    #     time.sleep(0.5)
    #
    #
    #
    #
    #
    #

    # >>>> Lanjut ke file deleterow.py secara otomatis <<<<
    print("▶️  Continue deleting row...")
    subprocess.run(["python", "deleterow.py"])


    jawaban = input("➡️  Continue to next data? (Y/N): ").strip().lower()

    if jawaban == 'y':
        print("▶️  Continue next data...")
        # lanjutkan script
    elif jawaban == 'n':
        print("⏹️ Process stop by user.")
        exit()
    else:
        print("⚠️  Unrecognized. Please tap Y or N on your keyboard.")


    # >>>> Lanjut ke file ecrf1.py secara otomatis <<<<
    print("▶️  Continue next nomor inklusi...")
    subprocess.run(["python", "form1.py"])


except KeyboardInterrupt:
    print("\n⏹️ Kill process by user.")

except Exception as e:
    print(f"\n❌ An unhandled exception occurred: {e}")

