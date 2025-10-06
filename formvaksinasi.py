import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, date, time as dtime
from openpyxl import load_workbook
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException
import time
import subprocess

try: 

    # --- file formvaksinasi.py
    # --- Setup Selenium untuk Brave ---
    options = Options()
    options.debugger_address = "127.0.0.1:9222"
    service = Service("chromedriver.exe")
    driver = webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 20)

    # --- Load nomor inklusi dari Excel ---
    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb.active
    val_no_inklusi = str(ws['K3'].value).strip()

    # ---------------- helper ----------------
    def set_text(driver, wait, id_, value):
        if value is None or str(value).strip() == "":
            return False
        try:
            el = wait.until(EC.visibility_of_element_located((By.ID, id_)))
            el.click()
            time.sleep(0.1)
            el.clear()
            time.sleep(0.1)
            el.send_keys(str(value))
            print(f"→ {id_} = {value}")
            return True
        except Exception as e:
            print(f"❌ FAILED failed set_text {id_}: {e}")
            try:
                # fallback pake javascript set value supaya lebih pasti masuk
                el = driver.find_element(By.ID, id_)
                driver.execute_script(
                    "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input'));", el, str(value)
                )
                print(f"→ {id_} set via JS fallback = {value}")
                return True
            except Exception as e2:
                print(f"❌ FAILED JS fallback failed for {id_}: {e2}")
                return False

    def format_tanggal_ddmmyyyy(value):
        """Return string dd-mm-YYYY or empty string."""
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return value.strftime("%d-%m-%Y")
        s = str(value).strip()
        if not s:
            return ""
        # try several common formats
        candidates = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d %B %Y", "%d %b %Y"]
        for fmt in candidates:
            try:
                dt = datetime.strptime(s, fmt)
                return dt.strftime("%d-%m-%Y")
            except Exception:
                pass
        # fallback: return original string
        return s

    def format_time_hhmm(value):
        """Return string HH:MM or empty string."""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%H:%M")
        # openpyxl may return time as datetime.time
        if isinstance(value, dtime):
            return value.strftime("%H:%M")
        s = str(value).strip()
        if not s:
            return ""
        candidates = ["%H:%M", "%H.%M", "%I:%M %p", "%H%M"]
        for fmt in candidates:
            try:
                dt = datetime.strptime(s, fmt)
                return dt.strftime("%H:%M")
            except Exception:
                pass
        return s

    def should_check(val):
        if val is None: return None
        hv = str(val).strip().lower()
        return hv in ("1","yes","ya","true","x","y")

    # ---------------- load Excel ----------------
    # mapping AA..AM
    val_no_inklusi      = ws["K3"].value   # itemid_60292 > No Inklusi
    val_inisial         = ws["C3"].value   # itemid_60293 >> Inisial
    val_jenis_vaksin    = ws["S3"].value   # form_group_60294 (radio)
    val_manufaktur      = ws["T3"].value   # form_group_60295 (radio)
    val_no_batch        = ws["U3"].value   # itemid_60296
    val_dosis = ws["V3"].value   # form_group_60297 (radio)
    val_tgl_vaksin = ws["W3"].value   # itemid_60298 (datepicker dd-mm-yyyy)
    val_wkt_vaksin = ws["X3"].value   # itemid_60299 (time hh:mm)
    val_tempat_vaksin = ws["Y3"].value   # form_group_60301 (radio)
    val_vaksin_lain = ws["Z3"].value   # form_group_60302 (radio controlling conditional)
    val_vaksin_lain1 = ws["AA3"].value   # conditional itemid_60303
    val_vaksin_lain2 = ws["AB3"].value   # conditional itemid_60304
    val_vaksin_lain3 = ws["AC3"].value   # conditional itemid_60305

    # optional checkbox from AN2 (if present). If AN2 empty, leave checkbox unchanged.
    val_haspengobatan = None
    try:
        val_haspengobatan = ws["O3"].value
    except Exception:
        val_haspengobatan = None

    print("Start to fill-in Form Vaksinasi")
    print(f"📘 No. inklusi from Excel: {val_no_inklusi}")
    print("Data from Excel:", val_no_inklusi, val_inisial, val_jenis_vaksin, val_manufaktur, val_no_batch, val_dosis, val_tgl_vaksin, val_wkt_vaksin, val_tempat_vaksin, val_vaksin_lain, val_vaksin_lain1, val_vaksin_lain2, val_vaksin_lain3, val_haspengobatan)

    driver.implicitly_wait(1)

    # ---------------- fill form ----------------
    def safe_find(id_or_selector, by=By.ID, timeout=3):
        try:
            if timeout:
                return WebDriverWait(driver, timeout).until(
                    EC.presence_of_element_located((by, id_or_selector))
                )
            else:
                return driver.find_element(by, id_or_selector)
        except Exception:
            return None

    # text inputs
    set_text(driver, wait, "itemid_60292", val_no_inklusi)
    set_text(driver, wait, "itemid_60293", val_inisial)

    # radios
    def set_radio(group_id, value):
        try:
            if value is None or str(value).strip() == "":
                return False
            css = f"#{group_id} input[type='radio'][value='{value}']"
            el = safe_find(css, By.CSS_SELECTOR, timeout=2)
            if el:
                driver.execute_script("arguments[0].click();", el)
                print(f"→ {group_id} = {value}")
                return True
            else:
                print(f"! radio {group_id} with value {value} not found")
                return False
        except Exception as e:
            print(f"! radio {group_id} error:", e)
            return False

    set_radio("form_group_60294", val_jenis_vaksin)
    set_radio("form_group_60295", val_manufaktur)

    # itemid_60296 (text)
    try:
        if val_no_batch is not None and str(val_no_batch).strip() != "":
            el = safe_find("itemid_60296", By.ID)
            if el:
                el.clear()
                el.send_keys(str(val_no_batch))
                print("→ itemid_60296 =", val_no_batch)
    except Exception as e:
        print("❌ FAILED itemid_60296 error:", e)

    set_radio("form_group_60297", val_dosis)

    # datepicker itemid_60298 -> dd-mm-YYYY
    try:
        if val_tgl_vaksin is not None and str(val_tgl_vaksin).strip() != "":
            formatted = format_tanggal_ddmmyyyy(val_tgl_vaksin)
            el = safe_find("itemid_60298", By.ID)
            if el:
                # set via JS + dispatch events then tab to close picker
                driver.execute_script(
                    "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input')); arguments[0].dispatchEvent(new Event('change'));",
                    el, formatted
                )
                time.sleep(0.15)
                el.send_keys(Keys.TAB)
                print("→ itemid_60298 =", formatted)
    except Exception as e:
        print("❌ FAILED itemid_60298 error:", e)

    # timepicker itemid_60299 -> HH:MM
    try:
        if val_wkt_vaksin is not None and str(val_wkt_vaksin).strip() != "":
            formatted_time = format_time_hhmm(val_wkt_vaksin)
            el = safe_find("itemid_60299", By.ID)
            if el:
                driver.execute_script(
                    "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input')); arguments[0].dispatchEvent(new Event('change'));",
                    el, formatted_time
                )
                time.sleep(0.15)
                el.send_keys(Keys.TAB)
                print("→ itemid_60299 =", formatted_time)
    except Exception as e:
        print("❌ FAILED itemid_60299 error:", e)

    # radio form_group_60301
    set_radio("form_group_60301", val_tempat_vaksin)

    # radio form_group_60302 (conditional)
    try:
        applied = False
        if val_vaksin_lain is not None and str(val_vaksin_lain).strip() != "":
            if set_radio("form_group_60302", val_vaksin_lain):
                applied = True
                # give UI a moment to show/hide conditional inputs
                time.sleep(0.4)
                if str(val_vaksin_lain).strip() == "1":
                    # fill AK2..AM2 conditional fields
                    if val_vaksin_lain1 is not None and str(val_vaksin_lain1).strip() != "":
                        el = safe_find("itemid_60303", By.ID)
                        if el:
                            el.clear()
                            el.send_keys(str(val_vaksin_lain1))
                            print("→ itemid_60303 =", val_vaksin_lain1)
                    if val_vaksin_lain2 is not None and str(val_vaksin_lain2).strip() != "":
                        el = safe_find("itemid_60304", By.ID)
                        if el:
                            el.clear()
                            el.send_keys(str(val_vaksin_lain2))
                            print("→ itemid_60304 =", val_vaksin_lain2)
                    if val_vaksin_lain3 is not None and str(val_vaksin_lain3).strip() != "":
                        el = safe_find("itemid_60305", By.ID)
                        if el:
                            el.clear()
                            el.send_keys(str(val_vaksin_lain3))
                            print("→ itemid_60305 =", val_vaksin_lain3)
                else:
                    print("→ form_group_60302 != 1; conditional fields left empty/hidden")
        if not applied:
            print("→ form_group_60302 not applied (no value or not found)")
    except Exception as e:
        print("! form_group_60302 error:", e)

    # checkbox (optional, read from AN2 if present)
    try:
        chk_val = val_haspengobatan
        if chk_val is not None:
            cb = safe_find("hasPengobatan", By.ID)
            if cb:
                want = should_check(chk_val)
                if want is True and not cb.is_selected():
                    driver.execute_script("arguments[0].click();", cb)
                    print("→ hasPengobatan checked")
                elif want is False and cb.is_selected():
                    driver.execute_script("arguments[0].click();", cb)
                    print("→ hasPengobatan unchecked")
                else:
                    print("→ hasPengobatan left as-is (matches desired state)")
            else:
                print("! hasPengobatan element not found")
        else:
            print("→ No checkbox value provided in AN2; checkbox left unchanged")
    except Exception as e:
        print("❌ FAILED hasPengobatan error:", e)

    print("✅ Copy-paste DATA VAKSINASI: DONE.")

    # PAUSE: tunggu Anda klik SAVE manual di browser lalu tekan ENTER di terminal
    input("👉 Klik tombol SAVE di browser, lalu tekan ENTER di terminal untuk lanjut...")
    #
    #
    #
    #
    #
    #
    # >>>> OPSI 3 : tidak perlu tekan ENTER di Keyboard <<<<<
    # >>>> setelah user klik tombol Submit, otomatis lanjut <<<<
    #
    # Tunggu sampai JavaScript var xSaved berisi "Data Berhasil Disimpan"
    # print("⌛ Waiting for SUBMIT button is clicked (dan xSaved = 'Data Berhasil Disimpan')...")
    # while True:
    #     try:
    #         xSaved_value = driver.execute_script("return window.xSaved;")
    #         if xSaved_value == "Data Berhasil Disimpan":
    #             print("✅ Saving succeessfully! Continue automatically...")
    #             break
    #     except:
    #         pass
    #     time.sleep(0.5)
    #
    #
    # Loop ke file openkipi.py
    jawaban = input("➡️  Continue to FORM KIPI? Jika ya, buka dulu formnya di website (Y/N): ").strip().lower()

    if jawaban == 'y':
        print("▶️  Continue to next form...")
        # lanjutkan script
    elif jawaban == 'n':
        print("⏹️ Process stop by user.")
        exit()
    else:
        print("⚠️  Unrecognized. Please tap Y or N on your keyboard.")

    # >>>> Lanjut ke file openkipi.py secara otomatis <<<<
    subprocess.run(["python", "formkipi.py"])

except KeyboardInterrupt:
    print("\n⏹️ Kill process by user.")

except Exception as e:
    print(f"\n❌ An unhandled exception occurred: {e}")


