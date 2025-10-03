import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, date, time as dtime
from openpyxl import load_workbook
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException
import time
import subprocess

try: 

    # --- file: form2.py 
    # Setup Brave (attach ke existing session)
    #
    options = Options()
    options.debugger_address = "127.0.0.1:9222"  # koneksi ke Brave yang sudah terbuka
    service = Service(r"chromedriver.exe")
    driver = webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 60)

    # --- Load nomor inklusi dari Excel ---
    wb = load_workbook('data.xlsx')
    ws = wb.active

    val_no_inklusi = ws['K3'].value

    # tunggu sampai form dengan id 'surveyform' muncul, maksimal 30 detik
    # print("⏳ Waiting for button 'ISI' is clicked and modal 'Inklusi/Eksklusi' load...")
    # wait = WebDriverWait(driver, 9999)
    # modal = wait.until(EC.visibility_of_element_located((By.ID, "surveyform")))
    # print("✅ Modal is loaded! Continue to copy-paste form Inklusi/Eksklusi...")
    #

    # Ambil data dari row 2, kolom G, H, I
    val_radio1 = ws["G3"].value  # nilai = value radio (misalnya "1" atau "2")
    val_radio2 = ws["H3"].value
    val_radio3 = ws["I3"].value


    print(f"📘 No. inklusi from Excel: {val_no_inklusi}")
    print("Data from Excel:", val_radio1, val_radio2, val_radio3)
    # --- Isi Radio Button ---

    # Radio button 1
    if val_radio1:
        radio_elem1 =driver.find_element(
            By.CSS_SELECTOR, f"#form_group_60312 input[type='radio'][value='{val_radio1}']"
        )
        radio_elem1.click()
    # label_text1 = radio_elem1.find_element(By.XPATH, "./parent::label").text
    print("→ form_group_60312 =", val_radio1)

    # Radio button 2
    if val_radio2:
        radio_elem2 = driver.find_element(
            By.CSS_SELECTOR, f"#form_group_60313 input[type='radio'][value='{val_radio2}']"
        )
        radio_elem2.click()
    # label_text2 = radio_elem2.find_element(By.XPATH, "./parent::label").text
    print("→ form_group_60313 =", val_radio2)

    # Radio button 3
    if val_radio3:
        radio_elem3 = driver.find_element(
            By.CSS_SELECTOR, f"#form_group_60314 input[type='radio'][value='{val_radio3}']"
        )
        radio_elem3.click()
    # label_text3 = radio_elem3.find_element(By.XPATH, "./parent::label").text
    print("→ form_group_60314 =", val_radio3)

    print("✅ Copy-paste INKLUSI/EKSKLUSI: DONE.")
    #
    #
    # >>>> OPSI 1 : otomatis save, lanjut ke script berikutnya <<<<<
    #
    # try:
    #     WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "btn-submit"))).click()
    #     print("→ Clicked btn-submit")
    # except Exception as e:
    #     print("❌ FAILED to click btn-submit:", e)

    # print("✅ Save FORM: DONE.")
    #
    #
    #
    #
    #
    #
    # >>>> OPSI 2 : tunggu user ENTER di Keyboard <<<<<
    #
    # intentionally do NOT call driver.quit()
    # PAUSE: tunggu Anda klik SAVE manual di browser lalu tekan ENTER di terminal
    input("👉 Klik tombol SAVE di browser, lalu tekan ENTER di terminal untuk lanjut...")
    #
    #
    #
    #
    #
    #
    # >>>> OPSI 3 : tidak perlu tekan ENTER di Keyboard <<<<<
    # >>>> setelah user klik tombol Submit/Save, lalu klik manual tombol Isi <<<<
    #
    # print("⌛ Waiting for SIMPAN button is clicked and notification = 'Responden Lolos'...")
    # try:
    #     wait.until(EC.presence_of_element_located(
    #         (By.XPATH, "//*[contains(text(),'Tambah Responden Berhasil')]")
    #     ))
    #     print("✅ Saving succeessfully! Continue automatically...")
    # except TimeoutException:
    #     print("⚠️ Timeout: Notification not found. Continue manual or check error.")
    
    #
    #
    # Loop ke file bukaisi3.py
    jawaban = input("➡️  Continue to FORM TAMBAH NOMOR INKLUSI? (Y/N): ").strip().lower()

    if jawaban == 'y':
        print("▶️  Continue to next form...")
        # lanjutkan script
    elif jawaban == 'n':
        print("⏹️ Process stop by user.")
        exit()
    else:
        print("⚠️  Unrecognized. Please tap Y or N on your keyboard.")

    # >>>> Lanjut ke file openform3.py secara otomatis <<<<
    subprocess.run(["python", "openform3.py"])

except KeyboardInterrupt:
    print("\n⏹️ Kill process by user.")

except Exception as e:
    print(f"\n❌ An unhandled exception occurred: {e}")
