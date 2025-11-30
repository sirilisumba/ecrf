import subprocess
import sys
import openpyxl
import re
import time, locale, os, winsound
from openpyxl.utils import column_index_from_string
from datetime import datetime, time as dtime
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
import winsound

EXCEL_PATH = "data.xlsx"
EXCEL_ROW = 3
LOG_PATH = 'log.xlsx'

# === Konstanta Bulan Bahasa Indonesia ===
MONTHS_ID = {
    "januari": 1,
    "februari": 2,
    "maret": 3,
    "april": 4,
    "mei": 5,
    "juni": 6,
    "juli": 7,
    "agustus": 8,
    "september": 9,
    "oktober": 10,
    "november": 11,
    "desember": 12
}


# --- Buat Selenium Driver ---
def create_driver(debugger_address="127.0.0.1:9222", chromedriver_path="chromedriver.exe", wait_long=30, wait_short=10):
    options = Options()
    options.debugger_address = debugger_address
    service = Service(chromedriver_path)
    driver = webdriver.Chrome(service=service, options=options)
    
    wait_long_obj = WebDriverWait(driver, wait_long)
    wait_short_obj = WebDriverWait(driver, wait_short)
    
    return driver, wait_long_obj, wait_short_obj

# --- Buat Selenium Driver  ---
# def create_driver(debugger_address="127.0.0.1:9222", wait_long=30, wait_short=10):
#     options = Options()
#     options.debugger_address = debugger_address

#     print(f"🔗 Menghubungkan ke Brave di {debugger_address} ...")
#     driver = webdriver.Chrome(options=options)
#     print("✅ Selenium berhasil terhubung ke Brave!")

#     wait_long_obj = WebDriverWait(driver, wait_long)
#     wait_short_obj = WebDriverWait(driver, wait_short)
#     return driver, wait_long_obj, wait_short_obj

def create_driver1():
    possible_paths = [
        r"C:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe",
        r"C:\Program Files (x86)\BraveSoftware\Brave-Browser\Application\brave.exe",
        r"D:\Program Files\BraveSoftware\Brave-Browser\Application\brave.exe",
    ]
    user_data_dir = r"C:\Users\Admin\AppData\Local\BraveSoftware\Brave-Browser\User Data"
    profile_name = "Default"  # bisa diganti sesuai profil Brave kamu (lihat di folder User Data)

    brave_path = next((p for p in possible_paths if os.path.exists(p)), None)
    if not brave_path:
        raise FileNotFoundError("❌ Brave Browser tidak ditemukan di lokasi umum.")

    # Jalankan Brave otomatis dengan remote debugging
    user_data_dir = os.path.join(os.getcwd(), "brave_profile")
    os.makedirs(user_data_dir, exist_ok=True)

    subprocess.Popen([
        brave_path,
        "--remote-debugging-port=9222",
        f"--user-data-dir={user_data_dir}",
        f"--profile-directory={profile_name}"
    ])

    # Tunggu browser siap
    time.sleep(3)

    # Koneksikan Selenium ke instance Brave yang baru dibuka
    options = Options()
    options.binary_location = brave_path
    options.debugger_address = "127.0.0.1:9222"

    service = Service("chromedriver.exe")
    driver = webdriver.Chrome(service=service, options=options)

    wait_long = WebDriverWait(driver, 30)
    wait_short = WebDriverWait(driver, 10)

    print(f"✅ Terhubung ke Brave browser di: {brave_path}")

    # 🔹 Arahkan langsung ke URL target
    target_url = "https://aplikasi.biofarma.co.id/eCRF"
    driver.get(target_url)
    print(f"🌐 Membuka URL: {target_url}")

    return driver, wait_long, wait_short

# --- Cek apakah row kosong ---
def is_row_empty(ws, row_num):
    for cell in ws[row_num]:
        if cell.value is not None and str(cell.value).strip() != "":
            return False
    return True


# --- Bunyi beep kalau error ---
def play_sound():
    freqs = [700, 880, 1040]    # naik -> terdengar 'positif'
    dur = 140                   # ms per nada
    gap = 0.11                  # detik antar nada
    for f in freqs:
        winsound.Beep(f, dur)
        time.sleep(gap)

def play_success():
    pattern = [
        (520, 300),   # nada pertama: agak panjang
        (420, 300),   # turun -> memberi nuansa peringatan
        (360, 220)    # turun lagi, lebih pendek
    ]
    gap = 0.13
    for freq, dur in pattern:
        winsound.Beep(freq, dur)
        time.sleep(gap)
    winsound.Beep(520, 90)

def play_error():
    duration = 440  # ms
    freq = 1000  # Hz
    for _ in range(3):
        winsound.Beep(freq, duration)
        time.sleep(0.2)

def wait_excel_closed():
    """Menunggu sampai semua file Excel (data dan log) tertutup."""
    while True:
        locked_files = []  # daftar file yang masih kebuka

        # Cek file data.xlsx
        try:
            with open(EXCEL_PATH, 'a'):
                pass
        except PermissionError:
            locked_files.append(EXCEL_PATH)

        # Cek file log.xlsx
        try:
            with open(LOG_PATH, 'a'):
                pass
        except PermissionError:
            locked_files.append(LOG_PATH)

        # Kalau dua-duanya aman, keluar dari loop
        if not locked_files:
            break

        # Kalau ada yang masih kebuka
        play_error()
        print("⚠️ File berikut masih terbuka di Excel:")
        for f in locked_files:
            print(f"   - {f}")
        input("➡️ Tutup semua file di atas, lalu tekan ENTER untuk lanjut...\n")


# --- Load Excel row tertentu ---
def load_excel_data(path=EXCEL_PATH, sheet_name=None, row_num=EXCEL_ROW):
    while True:
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet_name] if sheet_name else wb.active
        nomor_inklusi = ws[f'K{row_num}'].value

        if not is_row_empty(ws, row_num):
            return wb, ws, nomor_inklusi  # ✅ data ketemu, keluar dari loop

        wb.close()
        play_sound()
        print(f"⛔ Baris {row_num} kosong. Copy data dulu ke Excel sebelum lanjut.")
        ans = input("👉 Jika sudah tekan ENTER untuk lanjut, atau ketik 'stop' untuk keluar: ").strip().lower()
        if ans == "stop":
            print("⏹️ Proses dihentikan user.")
            sys.exit(0)

def load_excel_log():
    log_wb = openpyxl.load_workbook(LOG_PATH)
    log_ws = log_wb.active
    return log_wb, log_ws

def get_mapping(section=None):
    mapping_all = {
        "responden": {
            1: {"id": "puskesmas", "col": "B"},
            2: {"id": "inisial_nama", "col": "C"},
            3: {"id": "tanggalLahir_s", "col": "D"},
            4: {"radio": "jenis_kelamin", "col": "E"},
            5: {"id": "tanggalScreening_s", "col": "F"},
            6: {"id": "tanggal_inklusi_s", "col": "J"},
            7: {"id": "nomor_inklusi", "col": "K"},
            8: {"force": True, "col": "L"},
        },
        "Inklusi/Eksklusi": {
            60312: {"radio": "form_group_60312", "col": "G"},
            60313: {"radio": "form_group_60313", "col": "H"},
            60314: {"radio": "form_group_60314", "col": "I"},
        },
        "Informasi Pelapor": {
            58832: {"id": "itemid_58832", "col": "K"},
            58833: {"id": "itemid_58833", "col": "C"},
            58835: {"radio": "form_group_58835", "col": "M"},
            58836: {"id": "itemid_58836", "col": "N"},
        },
        "Informasi Pasien": {
            58337: {"id": "itemid_58337", "col": "K"},
            58338: {"id": "itemid_58338", "col": "C"},
            58340: {"radio": "form_group_58340", "col": "E"},
            59060: {"id": "itemid_59060", "col": "D"},
            59061: {"id": "itemid_59061", "col": "P"},
            59062: {"id": "itemid_59062", "col": "Q"},
            59063: {"id": "itemid_59063", "col": "R"},
        },
        "Data Vaksinasi": {
            60292: {"id": "itemid_60292", "col": "K"},
            60293: {"id": "itemid_60293", "col": "C"},
            60294: {"radio": "form_group_60294", "col": "S"},
            60295: {"radio": "form_group_60295", "col": "T"},
            60296: {"id": "itemid_60296", "col": "U"},
            60297: {"radio": "form_group_60297", "col": "V"},
            60298: {"id": "itemid_60298", "col": "W"},
            60299: {"id": "itemid_60299", "col": "X"},
            60301: {"radio": "form_group_60301", "col": "Y"},
            60302: {"radio": "form_group_60302", "col": "Z"},
            60303: {"id": "itemid_60303", "col": "AA"},
            60304: {"id": "itemid_60304", "col": "AB"},
            60305: {"id": "itemid_60305", "col": "AC"},
        },
        "KIPI": {
            58646: {"id": "itemid_58646", "col": "K"},          # no inklusi
            58647: {"id": "itemid_58647", "col": "C"},          # inisial
            58650: {"radio": "form_group_58650", "col": "AD"},  # kategori
            58766: {"radio": "form_group_58766", "col": "AE"},  # lokal?
            58767: {"radio": "form_group_58767", "col": "AF"},  # nyeri lokal?
            58768: {"radio": "form_group_58768", "col": "AG"},  # kemerahan
            58769: {"radio": "form_group_58769", "col": "AH"},  # penebalan
            58770: {"radio": "form_group_58770", "col": "AI"},  # pembengkakan
            58782: {"radio": "form_group_58782", "col": "AJ"},  # Reaksi Lokal Lain?
            59084: {"radio": "form_group_59084", "col": "AK"},  # Lokal lain 1 ?
            59088: {"id": "itemid_59088", "col": "AL"},         # nama
            59085: {"radio": "form_group_59085", "col": "AM"},  # Lokal lain 2 ?
            59092: {"id": "itemid_59092", "col": "AN"},         # nama
            59086: {"radio": "form_group_59086", "col": "AO"},  # Lokal lain 3 ?
            59095: {"id": "itemid_59095", "col": "AP"},         # nama
            59087: {"radio": "form_group_59087", "col": "AQ"},  # Lokal lain 4 ?
            59098: {"id": "itemid_59098", "col": "AR"},         # nama
            58781: {"radio": "form_group_58781", "col": "AS"},  # Sistemik?
            58795: {"radio": "form_group_58795", "col": "AT"},  # Demam?
            58796: {"radio": "form_group_58796", "col": "AU"},  # Rewel?
            58797: {"radio": "form_group_58797", "col": "AV"},  # menangis persisten?
            58807: {"radio": "form_group_58807", "col": "AW"},  # Sistemik Lain?
            59129: {"radio": "form_group_59129", "col": "AX"},  # Sistemik lain 1 ?
            59113: {"id": "itemid_59113", "col": "AY"},         # nama
            59130: {"radio": "form_group_59130", "col": "AZ"},  # Sistemik lain 2 ?
            59116: {"id": "itemid_59116", "col": "BA"},         # nama
            59131: {"radio": "form_group_59131", "col": "BB"},  # Sistemik lain 3 ?
            59119: {"id": "itemid_59119", "col": "BC"},         # nama
            59132: {"radio": "form_group_59132", "col": "BD"},  # Sistemik lain 4 ?
            59122: {"id": "itemid_59122", "col": "BEF"},         # nama
            58827: {"radio": "form_group_58827", "col": "BF"},  # Status akhir
        },
    }

    if section:
        # return mapping_all.get(section, {})
        return mapping_all[section]
    return mapping_all

def tulis_verif(log_ws, log_wb, nomor):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Tambahkan header jika sheet kosong
    if log_ws.max_row == 0:
        log_ws.append(["Timestamp", "Section", "Nomor Inklusi", "Pesan"])  # row 1
        log_ws.append([])  # row 2 kosong

    # Cari row terakhir yang berisi data
    last_row = log_ws.max_row
    while last_row > 2 and not log_ws.cell(row=last_row, column=1).value:
        last_row -= 1

    # Ambil nomor terakhir dari kolom 3
    last_nomor = str(log_ws.cell(row=last_row, column=3).value or "").strip()

    # Jika nomor terakhir sama, hentikan fungsi
    if last_nomor == str(nomor):
        print(f"⚠️ Nomor {nomor} sudah tercatat sebelumnya, skip tulis_verif().")
        return

    # Jika belum ada, lanjut tulis baru
    log_row = last_row + 1
    pesan2 = "Verifikasi Done"

    log_ws.cell(row=log_row, column=1, value=timestamp)
    log_ws.cell(row=log_row, column=2, value="")
    log_ws.cell(row=log_row, column=3, value=nomor)
    log_ws.cell(row=log_row, column=4, value=pesan2)

    log_wb.save(LOG_PATH)
    print(f"📝 Log updated: {nomor} - {pesan2}")

# --- Tulis log mismatch pake timestamp ke bawah ---
# def tulis_log(nomor, section, pesan):
#     while True:
#         try:
#             with open(LOG_PATH, 'a'):
#                 pass
#             break
#         except PermissionError:
#             play_error()
#             print(f"⚠️ File {LOG_PATH} sedang terbuka. Tutup Excel dulu.")
#             input("➡️ Tekan Enter setelah menutup Excel...")
#             time.sleep(0.5)

#     timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

#     if os.path.exists(LOG_PATH):
#         log_wb = openpyxl.load_workbook(LOG_PATH)
#         log_ws = log_wb.active
#     else:
#         log_wb = openpyxl.Workbook()
#         log_ws = log_wb.active

#     # Tambahkan header jika sheet kosong
#     if log_ws.max_row == 0:
#         log_ws.append(["Timestamp", "Section", "Nomor Inklusi", "Pesan"])  # row 1
#         log_ws.append([])  # row 2 kosong

#     # Tentukan row log baru mulai dari row 3
#     log_row = 3
#     while log_ws.cell(row=log_row, column=1).value:
#         log_row += 1

#     # Tulis log
#     log_ws.cell(row=log_row, column=1, value=timestamp)
#     log_ws.cell(row=log_row, column=2, value=section)
#     log_ws.cell(row=log_row, column=3, value=nomor)
#     log_ws.cell(row=log_row, column=4, value="; ".join(pesan) if isinstance(pesan, list) else str(pesan))
#     log_wb.save(LOG_PATH)
#     print(f"📝 Log updated ({section}): {nomor} - {pesan}")

def tulis_log(log_ws, log_wb, nomor, section, pesan):

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Tentukan row log baru mulai dari row 3
    log_row = 3
    while log_ws.cell(row=log_row, column=1).value:
        log_row += 1

    # Tulis log
    log_ws.cell(row=log_row, column=1, value=timestamp)
    log_ws.cell(row=log_row, column=2, value=section)
    log_ws.cell(row=log_row, column=3, value=nomor)
    log_ws.cell(row=log_row, column=4, value="; ".join(pesan) if isinstance(pesan, list) else str(pesan))
    log_wb.save(LOG_PATH)
    print(f"📝 Log updated ({section}): {nomor} - {pesan}")

# --- Buka form 1 (fa-file) ---
def buka_form1(driver, wait_long, nomor_inklusi, max_retries=3):
    xpath_tr = f"//tr[td[text()='{nomor_inklusi}']]"
    xpath_btn = xpath_tr + "//button[i[contains(@class, 'fa-file')]]"
    try:
        print("⏳ Menunggu data dimuat dan spinner hilang...")
        wait_long.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".spinner, .loading")))
        time.sleep(0.5)
        print("✅ Loading selesai, lanjut buka form...")
    except Exception:
        print("⚠️ Timeout spinner, lanjut proses...")


    for attempt in range(1, max_retries+1):
        try:
            print(f"🔎 Mencari baris untuk {nomor_inklusi}, percobaan {attempt}...")
            
            # Cari ulang baris dan tombol setiap kali (hindari stale)
            tr_elem = wait_long.until(EC.presence_of_element_located((By.XPATH, xpath_tr)))
            btn = wait_long.until(EC.element_to_be_clickable((By.XPATH, xpath_btn)))

            # Scroll ke tengah layar
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
            time.sleep(0.3)

            # Klik dengan JS biar lebih kuat
            driver.execute_script("arguments[0].click();", btn)
            print(f"✅ Tombol fa-file berhasil diklik (percobaan {attempt})")

            # Tunggu modal muncul
            wait_long.until(EC.visibility_of_element_located((By.ID, "myModalTambahResponden")))
            print("✅ Modal myModalTambahResponden sudah muncul")
            return True

        except (StaleElementReferenceException, TimeoutException) as e:
            print(f"⚠️ Stale/Timeout di percobaan {attempt}: {e}")
            time.sleep(1)  # kasih jeda lalu coba ulang
            continue
        except Exception as e:
            print(f"❌ Error lain di percobaan {attempt}: {e}")
            break

    print("❌ Gagal buka form setelah beberapa percobaan")
    play_error()
    return False

def buka_form2(driver, wait_long, nomor_inklusi, nama_form, max_retries=3):
    for attempt in range(1, max_retries + 1):
        try:
            print(f"\n🔎 [Percobaan {attempt}] Buka form {nama_form} untuk {nomor_inklusi}")

            # Pastikan spinner/loading sudah hilang
            wait_long.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".spinner, .loading")))
            time.sleep(0.5)
            print("✅ Loading selesai, lanjut proses...")

            # Temukan baris utama
            xpath_tr = f"//tr[td[normalize-space(text())='{nomor_inklusi}']]"
            # tr_elem = wait_long.until(EC.presence_of_element_located((By.XPATH, xpath_tr)))

            # Klik chevron untuk expand
            chevron_xpath = xpath_tr + "//td[last()]//i[contains(@class,'fa-chevron-down')]"
            chevron = wait_long.until(EC.element_to_be_clickable((By.XPATH, chevron_xpath)))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", chevron)
            time.sleep(0.3)
            driver.execute_script("arguments[0].click();", chevron)
            print("✅ Chevron diklik, menunggu detail_row muncul...")

            # Tunggu detail_row muncul
            wait_long.until(EC.presence_of_element_located((
                By.XPATH,
                "//tr[starts-with(@class,'detail_row')]//div[contains(@id,'accordion')]"
            )))
            time.sleep(0.5)

            # Cari tombol "Lihat"
            tombol_xpath = f"//tr[starts-with(@class,'detail_row')]//tr[td[contains(normalize-space(.), '{nama_form}')]]//button[contains(@class, 'btn_lihat_uji_klinis')]"
            tombol_lihat = wait_long.until(EC.element_to_be_clickable((By.XPATH, tombol_xpath)))

            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", tombol_lihat)
            time.sleep(0.3)
            driver.execute_script("arguments[0].click();", tombol_lihat)
            print("✅ Tombol 'Lihat' berhasil diklik")

            # Tunggu modal muncul
            wait_long.until(EC.visibility_of_element_located((By.ID, "surveyform")))
            print("✅ Modal surveyform sudah muncul\n")
            return True

        except (StaleElementReferenceException, TimeoutException) as e:
            print(f"⚠️ Percobaan {attempt} gagal ({type(e).__name__}), coba ulang...")
            time.sleep(1)
            continue
        except Exception as e:
            print(f"❌ Error lain di percobaan {attempt}: {e}")
            break

    print(f"❌ Gagal buka form {nama_form} setelah {max_retries} percobaan\n")
    play_error()
    return False



# --- Locale Indonesia ---
try:
    locale.setlocale(locale.LC_TIME, "id_ID.utf8")
except:
    try:
        locale.setlocale(locale.LC_TIME, "Indonesian_indonesia")
    except:
        print("⚠️ Locale Indonesia tidak tersedia, pakai default")

# --- Normalisasi tanggal ---
def normalize_date_flex(date_str):
    try:
        s = date_str.strip()
        # Coba beberapa format umum
        for fmt in ("%d %B %Y", "%-d %B %Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(s, fmt)
                return f"{dt.day} {dt.strftime('%B %Y')}"  # tanpa leading zero
            except:
                continue
        return s
    except:
        return date_str.strip()

def cek_dan_klik_verified(driver, wait_short, element_id, should_be, field_data):
    try:
        element_id = str(element_id)
        icon = wait_short.until(EC.element_to_be_clickable((By.ID, f"verified_{element_id}")))

        # Ambil status verified dari class
        class_attr = icon.get_attribute("class") or ""
        current_val = "1" if "color_21B" in class_attr else "0"

        # === LOGIKA SESUAI KETENTUAN ===
        if should_be == "1":
            # actual == expected
            if current_val == "1":
                print(f"ℹ️ verified_{element_id} sudah bernilai 1 (sesuai), skip ({field_data})")
                return True
            else:
                driver.execute_script("arguments[0].scrollIntoView(true);", icon)
                time.sleep(0.2)
                driver.execute_script("arguments[0].click();", icon)
                print(f"✅ verified_{element_id} diubah dari 0 → 1 ({field_data})")
                return True

        elif should_be == "0":
            # actual != expected
            if current_val == "1":
                driver.execute_script("arguments[0].scrollIntoView(true);", icon)
                time.sleep(0.2)
                driver.execute_script("arguments[0].click();", icon)
                print(f"🟡 verified_{element_id} diubah dari 1 → 0 ({field_data})")
                return True
            else:
                print(f"ℹ️ verified_{element_id} sudah bernilai 0 (sesuai), skip ({field_data})")
                return True

        else:
            print(f"⚠️ Nilai should_be tidak valid: {should_be}")
            return False

    except Exception as e:
        print(f"⚠️ Gagal akses atau klik verified_{element_id}: {e} ({field_data})")
        return False

# --- Ambil semua label sekaligus (cache) ---
def build_label_cache(driver, mapping):
    cache = {}
    for key, m in mapping.items():
        try:
            if "id" in m:
                label_elem = driver.find_element(By.CSS_SELECTOR, f"label[for='{m['id']}']")
                text = label_elem.text.strip()
                cache[key] = text if text else m['id']
            elif "radio" in m:
                label_elem = driver.find_element(By.CSS_SELECTOR, f"#form_group_{key} label")
                text = label_elem.text.strip()
                cache[key] = text if text else m['radio']
            else:
                cache[key] = m.get("id") or m.get("radio") or "unknown"
        except:
            cache[key] = m.get("id") or m.get("radio") or "unknown"
    return cache

# --- helper normalizer ---
def normalize_cell(val):
    if val is None:
        return ""
    if isinstance(val, datetime.datetime):
        return val.strftime("%d-%m-%Y")
    if isinstance(val, datetime.time):
        return val.strftime("%H:%M")
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val.is_integer():
            return str(int(val))
        return str(val)
    return str(val).strip()

def normalize_attr(val):
    if val is None:
        return ""
    return str(val).strip()

# --- Fungsi bantu tunggu field terisi ---
def wait_for_field_value(driver, wait_short, field_id, timeout=5):
    end_time = time.time() + timeout
    val = ""
    while time.time() < end_time:
        elem = driver.find_element(By.ID, field_id)
        val = (
            elem.get_attribute("value-data") 
            or elem.get_attribute("value")
            or ""
        ).strip()
        if val:
            return val
        time.sleep(0.1)
    return val

def parse_date_indonesia(s):
    if not s:
        return None
    s = s.strip()
    parts = s.split()
    if len(parts) == 3:
        try:
            day = int(parts[0])
            month_name = parts[1].lower()
            year = int(parts[2])
            month = MONTHS_ID.get(month_name)
            if month:
                return datetime.date(year, month, day)
        except Exception:
            pass
    # fallback: try python's strptime (may fail if locale bukan id)
    try:
        dt = datetime.strptime(s, "%d %B %Y")
        return dt.date()
    except Exception:
        return None

def read_verified_state(driver, idx):
    """Baca state verified_idx dari DOM. Return '1' atau '0' (string)."""
    try:
        el = driver.find_element(By.ID, f"verified_{idx}")
        class_attr = (el.get_attribute("class") or "").lower()
        # heuristic: class contains indicator for active/colored state
        if "color_21b" in class_attr or "active" in class_attr or "checked" in class_attr or "verified" in class_attr:
            return "1"
        # sometimes there's a value attribute
        val = el.get_attribute("value")
        if val in ("1", "0"):
            return str(val)
    except Exception:
        pass
    return "0"

def click_verified_idx(driver, wait_short, idx):
    """Klik icon verified_idx (JS click). Return True kalau klik dilakukan, else False."""
    try:
        icon = wait_short.until(EC.element_to_be_clickable((By.ID, f"verified_{idx}")))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", icon)
        time.sleep(0.08)
        driver.execute_script("arguments[0].click();", icon)
        time.sleep(0.18)  # beri waktu UI update
        return True
    except Exception as e:
        print(f"⚠️ Gagal klik verified_{idx}: {e}")
        return False

# --- Fungsi bantu ambil label field ---
def get_label_text(driver, key, m):
    try:
        if "id" in m:
            label_elem = driver.find_element(By.CSS_SELECTOR, f"label[for='{m['id']}']")
            label_text = label_elem.text.strip()
            return label_text if label_text else m['id']
        elif "radio" in m:
            label_elem = driver.find_element(By.CSS_SELECTOR, f"#form_group_{key} label")
            label_text = label_elem.text.strip()
            return label_text if label_text else m['radio']
        else:
            return m.get("id") or m.get("radio") or "unknown"
    except:
        return m.get("id") or m.get("radio") or "unknown"

def normalize_date_only(value):
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y")
    if isinstance(value, dtime):
        return ""
    for fmt in ("%d-%m-%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value.strip(), fmt).strftime("%d-%m-%Y")
        except:
            continue
    return value.strip()

def normalize_time_only(value):
    if not value:
        return ""
    if isinstance(value, dtime):
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(value.strip(), fmt).strftime("%H:%M")
        except:
            continue
    return value.strip()

def get_field_actual_value(driver, m):
    dataid = m.get("dataid") or m.get("radio") or m.get("id") or ""
    typeid = (m.get("typeid") or "").lower()
    dataid = str(dataid).replace("itemid_", "").replace("form_group_", "").strip()

    if not typeid:
        try:
            group_div = driver.find_element(By.ID, f"form_group_{dataid}")
            typeid = (group_div.get_attribute("typeid") or "").lower()
        except:
            typeid = ""

    try:
        if typeid == "radio":
            try:
                elem = driver.find_element(
                    By.XPATH,
                    f"//div[@id='form_group_{dataid}']//input[@type='radio' and @checked]"
                )
                return elem.get_attribute("value").strip()
            except:
                return ""

        elem = driver.find_element(
            By.XPATH,
            f"//div[@id='form_group_{dataid}']//*[starts-with(@id,'itemid_') or name()='textarea' or name()='select']"
        )

        tag = elem.tag_name.lower()
        if tag == "input":
            input_type = elem.get_attribute("type") or ""
            if input_type == "checkbox":
                return "1" if elem.is_selected() else "0"
            return (elem.get_attribute("value") or "").strip()
        elif tag == "textarea":
            return (elem.get_attribute("value") or "").strip()
        elif tag == "select":
            try:
                selected = elem.find_element(By.XPATH, ".//option[@selected]")
                return (selected.get_attribute("value") or "").strip()
            except:
                return ""
        else:
            return (elem.get_attribute("value") or "").strip()
    except Exception as e:
        print(f"⚠️ Tidak bisa ambil nilai actual untuk {dataid} ({typeid}): {e}")
        return ""

def cek_dan_klik_verified_smart(driver, wait_short, key, field_name, expected, actual, current_val, field_type="text"):
    if field_type == "date":
        expected_norm = normalize_date_only(expected)
        actual_norm = normalize_date_only(actual)
    elif field_type == "time":
        expected_norm = normalize_time_only(expected)
        actual_norm = normalize_time_only(actual)
    else:
        expected_norm = str(expected).strip()
        actual_norm = str(actual).strip()

    # --- Cari icon verified dulu (digunakan untuk comment) ---
    try:
        icon = wait_short.until(EC.presence_of_element_located((By.ID, f"verified_{key}")))
    except Exception as e:
        print(f"⚠️ Icon verified_{key} tidak ditemukan: {e}")
        icon = None

    # --- Kalau actual ≠ expected, tulis comment dulu ---
    if actual_norm != expected_norm and icon:
        # Cek apakah sudah ada comment
        if sudah_ada_comment1(driver, key):
            print(f"💬 Field {field_name}: sudah ada comment, skip buka modal.")
            return current_val  # ⬅️ penting! hentikan fungsi di sini supaya gak lanjut
        else:
            play_sound()
            try:
                # cari tombol comment di div yang sama
                div_container = icon.find_element(By.XPATH, "./ancestor::div[contains(@class,'col-sm-2')]")
                comment_btn = div_container.find_element(By.XPATH, f".//button[contains(@onclick, 'commenting({key},')]")
                
                # tunggu tombol siap diklik
                wait_short.until(EC.element_to_be_clickable((By.XPATH, f".//button[contains(@onclick, 'commenting({key},')]")))
                
                driver.execute_script("arguments[0].scrollIntoView(true);", comment_btn)
                time.sleep(0.2)
                comment_btn.click()
                
                textarea = wait_short.until(EC.visibility_of_element_located((By.ID, "comment_text")))
                time.sleep(0.3)
                textarea.clear()
                # 💡 Tentukan isi comment berdasarkan kondisi actual
                if actual_norm == "" or actual_norm is None:
                    comment_text = "Wajib diisi."
                else:
                    comment_text = "Data yg di-input tidak sesuai dengan nilai di Excel atau aturan penulisan."
                print(f"Expected: {expected_norm} >> Actual: {actual}")
                textarea.send_keys(comment_text)
                print(f"📝 Comment untuk field {field_name} sudah diisi.")

                # Tunggu user manual
                input("📋 Cek comment di browser. Setelah yakin, tekan [Enter] untuk lanjut script...")
                print("✅ Lanjutkan script...\n")

                # Tunggu modal tertutup
                WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((By.ID, "myModalAddComment")))
                print("✅ Modal comment tertutup, lanjutkan script...\n")

            except Exception as e:
                print(f"⚠️ Gagal menulis comment untuk field {field_name}: {e}")


    # --- Tentukan nilai verified_ ---
    should_be = "1" if expected_norm == actual_norm else "0"
    print(f"🔎 Field {field_name}: expected={expected_norm}, actual={actual_norm}, current_verified={current_val}, should_be={should_be}")

    if current_val != should_be and icon:
        try:
            driver.execute_script("arguments[0].scrollIntoView(true);", icon)
            time.sleep(0.2)
            driver.execute_script("arguments[0].click();", icon)
            print(f"✅ verified_{key} diubah dari {current_val} → {should_be} ({field_name})")
        except Exception as e:
            print(f"⚠️ Gagal klik verified_{key}: {e}")
    else:
        print(f"ℹ️ verified_{key} sudah sesuai, skip ({field_name})")

    return should_be

def tulis_comment(driver, wait_short, key, pesan):
    play_sound()
    try:
        print(f"📝 Menulis comment untuk verify_{key}: {pesan}")
        icon = driver.find_element(By.ID, f"verify_{key}")
        div_container = icon.find_element(By.XPATH, "./parent::div[contains(@class,'col-sm')]")
        comment_btn = div_container.find_element(By.CSS_SELECTOR, "button.button_comment")

        driver.execute_script("arguments[0].scrollIntoView(true);", comment_btn)
        time.sleep(0.2)
        comment_btn.click()

        modal = wait_short.until(EC.visibility_of_element_located((By.ID, "myModalAddComment")))
        time.sleep(0.3)

        textarea = modal.find_element(By.ID, "comment_text")
        textarea.clear()
        textarea.send_keys(pesan)

        # print(f"✅ Comment verify_{key} sudah diisi.")
        input("📋 Cek comment di browser, lalu tekan [Enter] untuk lanjut...")

        # Tunggu modal tertutup
        try:
            wait_short.until(EC.invisibility_of_element_located((By.ID, "myModalAddComment")))
            print("✅ Modal sudah tertutup, lanjut script.\n")
            # Cek kalau form ikut tertutup (kembali ke tabel)
            time.sleep(1)
            if len(driver.find_elements(By.ID, "myModalTambahResponden")) == 0:
                print("↩️ Form tertutup — akan ulang dari field 1 nanti.")
                return "form_closed"
            return True
        except TimeoutException:
            print("❌ Modal masih terbuka — kemungkinan user klik Batal.\n")
            return False

    except Exception as e:
        print(f"⚠️ Gagal menulis comment verify_{key}: {e}")
        return False


def sudah_ada_comment(driver, key, debug=True):
    try:
        time.sleep(0.25)
        buttons = driver.find_elements(By.CSS_SELECTOR, "button.button_comment")
        pattern = re.compile(r",\s*{}\)".format(re.escape(str(key))))

        for btn in buttons:
            onclick = (btn.get_attribute("onclick") or "")
            if not pattern.search(onclick):
                continue  # bukan untuk key ini

            try:
                icon = btn.find_element(By.TAG_NAME, "i")
            except:
                if debug:
                    print(f"⚠️ verify_{key}: tombol tanpa ikon.")
                continue

            classes = icon.get_attribute("class") or ""
            style = icon.get_attribute("style") or ""

            if debug:
                print(f"🔍 verify_{key}: class={classes} | style={style}")

            # Deteksi warna merah atau class font_-3
            if "font_-3" in classes or "255, 0, 0" in style:
                if debug:
                    print(f"💬 Comment verify_{key} sudah ADA (warna merah).")
                return True
            else:
                if debug:
                    print(f"💬 Comment verify_{key} belum ada (ikon masih default).")
                return False

        # if debug:
        #     print(f"⚠️ Tidak menemukan tombol comment untuk verify_{key}.")
        # return False

    except Exception as e:
        print(f"⚠️ Gagal deteksi comment verify_{key}: {e}")
        return False

def sudah_ada_comment1(driver, key, debug=True):
    try:
        time.sleep(0.25)
        buttons = driver.find_elements(By.CSS_SELECTOR, "button.button_comment")
        # ✅ Sesuai struktur onclick="commenting(58833, ..."
        pattern = re.compile(r"commenting\(\s*{}".format(re.escape(str(key))))

        for btn in buttons:
            onclick = (btn.get_attribute("onclick") or "")
            if not pattern.search(onclick):
                continue  # bukan untuk key ini

            try:
                icon = btn.find_element(By.TAG_NAME, "i")
            except:
                if debug:
                    print(f"⚠️ verify_{key}: tombol tanpa ikon.")
                continue

            classes = icon.get_attribute("class") or ""
            style = icon.get_attribute("style") or ""

            if debug:
                print(f"🔍 verify_{key}: class={classes} | style={style}")

            # Deteksi warna merah / class font_-3
            if "font_-3" in classes or "255, 0, 0" in style:
                if debug:
                    print(f"💬 Comment verify_{key} sudah ADA (warna merah).")
                return True
            else:
                if debug:
                    print(f"💬 Comment verify_{key} belum ada (ikon masih default).")
                continue  # jangan return langsung

        if debug:
            print(f"⚠️ Tidak menemukan tombol comment untuk verify_{key}.")
        return False

    except Exception as e:
        print(f"⚠️ Gagal deteksi comment verify_{key}: {e}")
        return False
