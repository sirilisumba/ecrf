import sys
import openpyxl
import time, os, winsound
# from datetime import date, datetime
from datetime import datetime, date, time as dtime
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from verif_utils import (
    create_driver, wait_excel_closed, load_excel_data, load_excel_log, 
    EXCEL_PATH, EXCEL_ROW, LOG_PATH, get_mapping, buka_form1, buka_form2,
    wait_for_field_value, normalize_attr, get_label_text, cek_dan_klik_verified_smart,
    parse_date_indonesia, normalize_date_flex, sudah_ada_comment,
    cek_dan_klik_verified, tulis_log, tulis_comment, play_error, tulis_verif
)

def main():
    driver, wait_long, wait_short = create_driver()
    print("🚀 Program mulai. Tekan CTRL+C untuk berhenti.\n")

    while True:
        wait_excel_closed()
        wb, ws, nomor_inklusi = load_excel_data()
        log_wb, log_ws = load_excel_log()

        print(f"\n🔍 Memproses nomor inklusi: {nomor_inklusi}")

        # ===========================================
        # === SCRIPT FILE 1 ===
        # ===========================================
        SECTION = "responden"
        mapping = get_mapping(SECTION)
        try:
            # TODO: isi script file 1

            print("▶️ Script berjalan... tekan CTRL+C untuk berhenti manual.")
            print("⏳ Menunggu tabel '#tbl_responden' & field siap...")

            ###################
            ##### HELPER ######
            ###################

            def compare_web_versus_excel(driver, wait_short, wait_long, key, mapping):
                MIN_DATE = date(2025, 9, 13)
                TODAY = date.today()
                mismatches = []
                global ulang_dari_awal

                try:
                    # --- FIELD KHUSUS 5 & 6 (Tanggal IC dan Enrollment) ---
                    if key in (5, 6):
                        actual5 = wait_for_field_value(driver, wait_short, "verify_5")
                        actual6 = wait_for_field_value(driver, wait_short, "verify_6")
                        tgl5 = parse_date_indonesia(actual5)
                        tgl6 = parse_date_indonesia(actual6)

                        # Validasi range tanggal
                        in_range5 = tgl5 and (MIN_DATE <= tgl5 <= TODAY)
                        in_range6 = tgl6 and (MIN_DATE <= tgl6 <= TODAY)
                        should_be_5, should_be_6 = "1", "1"  # default lolos semua

                        if not in_range5:
                            should_be_5 = "0"

                            if not sudah_ada_comment(driver, 5):
                                if actual5 == "" or actual5 is None:
                                    comment_text = "Wajib diisi."
                                else:
                                    comment_text = f"Tanggal Informed Consent di luar range yang diperbolehkan: 13-Sep-2025 s.d. Today."

                                mismatches.append(f"Tanggal Informed Consent di luar range: {actual5}")
                                tulis_comment(driver, wait_long, 5, comment_text)
                                tulis_log(log_ws, log_wb, nomor_inklusi, SECTION, mismatches)
                                ulang_dari_awal = True
                                buka_form1(driver, wait_long, nomor_inklusi)
                                return None, None

                        if not in_range6:
                            should_be_6 = "0"

                            if not sudah_ada_comment(driver, 6):
                                if actual6 == "" or actual6 is None:
                                    comment_text = "Wajib diisi."
                                else:
                                    comment_text = f"Tanggal Enrollment di luar range yang diperbolehkan 13-Sep-2025 s.d. Today."

                                mismatches.append(f"Tanggal Enrollment di luar range: {actual6}")
                                tulis_comment(driver, wait_long, 6, comment_text)
                                tulis_log(log_ws, log_wb, nomor_inklusi, SECTION, mismatches)
                                ulang_dari_awal = True
                                buka_form1(driver, wait_long, nomor_inklusi)
                                return None, None


                        # --- CASE 2: Dua tanggal beda ---
                        elif tgl5 != tgl6:
                            should_be_5, should_be_6 = "0", "0"

                            if not sudah_ada_comment(driver, 5):
                                if actual5 == "" or actual5 is None:
                                    comment_text = "Wajib diisi."
                                else:
                                    comment_text = "Tanggal Informed Consent tidak sama dengan Tanggal Enrollment."

                                mismatches.append(f"Tanggal Informed Consent ≠ Tanggal Enrollment ({actual5} ≠ {actual6})")
                                tulis_comment(driver, wait_long, 5, comment_text)
                                tulis_log(log_ws, log_wb, nomor_inklusi, SECTION, mismatches)
                                ulang_dari_awal = True
                                buka_form1(driver, wait_long, nomor_inklusi)
                                return None, None

                            if not sudah_ada_comment(driver, 6):
                                if actual6 == "" or actual6 is None:
                                    comment_text = "Wajib diisi."
                                else:
                                    comment_text = "Tanggal Enrollment tidak sama dengan Tanggal Informed Consent."

                                mismatches.append(f"Tanggal Enrollment ≠ Tanggal Informed Consent ({actual6} ≠ {actual5})")
                                tulis_comment(driver, wait_long, 6, comment_text)
                                tulis_log(log_ws, log_wb, nomor_inklusi, SECTION, mismatches)
                                ulang_dari_awal = True
                                buka_form1(driver, wait_long, nomor_inklusi)
                                return None, None

                        # --- CASE 3: Valid & sama ---
                        else:
                            print(f"✅ Tanggal Informed Consent & Enrollment valid dan sama: {actual5}")

                        # Update verified setelah semua cek selesai
                        cek_dan_klik_verified(driver, wait_short, 5, should_be_5, "Tanggal Informed Consent")
                        cek_dan_klik_verified(driver, wait_short, 6, should_be_6, "Tanggal Enrollment")

                        return actual5, actual6

                    # ==========================================================
                    # === FIELD LAIN (Selain 5 & 6)
                    # ==========================================================
                    m = mapping[key]
                    col_index = column_index_from_string(m['col'])
                    cell_value = ws.cell(row=3, column=col_index).value
                    expected = ""

                    if isinstance(cell_value, datetime):
                        expected = f"{cell_value.day} {cell_value.strftime('%B %Y')}"
                    elif cell_value:
                        expected = str(cell_value).strip()

                    verify_elem = driver.find_element(By.ID, f"verify_{key}")
                    field_data = verify_elem.get_attribute("field-data") or f"Field {key}"

                    # --- Ambil nilai actual ---
                    actual = ""
                    if "id" in m:
                        actual = (verify_elem.get_attribute("value-data") or "").strip()
                    elif "radio" in m:
                        for rb in driver.find_elements(By.NAME, m["radio"]):
                            if rb.is_selected():
                                actual = (rb.get_attribute("value") or "").strip()
                                break
                    elif "force" in m:
                        actual = expected

                    # --- Normalisasi tanggal (kalau nama field mengandung "Tanggal") ---
                    if "Tanggal" in field_data or "tgl" in field_data.lower():
                        actual = normalize_date_flex(actual)
                        expected = normalize_date_flex(expected)

                    # --- Case-insensitive comparison ---
                    if isinstance(actual, str):
                        actual = actual.casefold().strip()
                    if isinstance(expected, str):
                        expected = expected.casefold().strip()

                    # --- Tentukan hasil perbandingan ---
                    should_be = "1" if actual == expected else "0"

                    if actual == expected:
                        print(f"[Field {key}] {field_data} ✅ DATA SAMA → Actual: {actual}")
                    else:
                        print(f"[Field {key}] {field_data} ❌ DATA TIDAK SAMA → Actual: {actual}, Expected: {expected}")

                        if sudah_ada_comment(driver, key):
                            print(f"💬 Comment untuk field {field_data} sudah ada, skip buka modal.")
                        else:
                            # 💡 Tentukan isi comment berdasarkan kondisi actual
                            if actual == "" or actual is None:
                                comment_text = "Wajib diisi."
                            else:
                                comment_text = "Data yg di-input tidak sesuai dengan nilai di Excel atau aturan penulisan."
                            mismatches.append(f"{field_data} (expected={expected}, got={actual})")
                            tulis_comment(driver, wait_long, key, comment_text)
                            tulis_log(log_ws, log_wb, nomor_inklusi, SECTION, mismatches)
                            ulang_dari_awal = True
                            buka_form1(driver, wait_long, nomor_inklusi)
                            return None, None

                    # Update verified
                    cek_dan_klik_verified(driver, wait_short, key, should_be, field_data)

                    # Simpan log bila ada mismatch
                    if mismatches:
                        tulis_log(log_ws, log_wb, nomor_inklusi, SECTION, mismatches)

                    return actual, expected

                except Exception as e:
                    play_error()
                    print(f"⚠️ Error di field {key}: {e}")
                    return None, None

            ###################
            ##### END    ######
            ###################

            try:
                print("⏳ Menunggu tabel '#tbl_responden' siap...")
                wait_long.until(EC.visibility_of_element_located((By.ID, "tbl_responden")))
                
                print("⏳ Menunggu input field 'field_cari' siap...")
                field = wait_long.until(EC.element_to_be_clickable((By.ID, "field_cari")))
                
                print("✅ Tabel & field siap, lanjut buka form...")
            except TimeoutException:
                play_error()
                print("⚠️ Timeout: tabel atau field_cari tidak muncul. Periksa halaman atau network.")
                field = None


            # --- Pastikan Excel bisa diakses sebelum mulai ---
            while True:
                try:
                    with open(EXCEL_PATH, 'a'):
                        pass
                    break
                except PermissionError:
                    play_error()
                    print(f"⚠️ File {EXCEL_PATH} sedang terbuka. Tutup Excel dulu.")
                    input("➡️ Tekan Enter setelah menutup Excel...")

            # --- Loop utama untuk memproses semua field 1–9 ---
            try:
                while True:
                    wb = load_workbook(EXCEL_PATH)
                    ws = wb.active

                    # Stop jika tidak ada data
                    if ws.max_row < 3:
                        play_error()
                        print("✅ Semua data selesai diproses")
                        break

                    nomor_inklusi = ws['K3'].value
                    print(f"🔍 Memproses nomor inklusi: {nomor_inklusi} pada FORM {SECTION}")

                    try:
                        # Pastikan spinner hilang dulu sebelum klik field_cari
                        try:
                            wait_long.until_not(
                                EC.presence_of_element_located((By.CSS_SELECTOR, ".loading_spinner[style*='display: block']"))
                            )
                            print("✅ Spinner hilang, aman untuk klik field_cari.")
                        except:
                            play_error()
                            print("⚠️ Spinner tidak muncul atau sudah hilang sebelumnya.")
                        
                        # Tambahan kecil: delay singkat agar halaman benar-benar stabil
                        time.sleep(0.3)

                        # Cari nomor inklusi di web
                        field = wait_long.until(EC.element_to_be_clickable((By.ID, "field_cari")))
                        field.click()
                        field.clear()
                        field.send_keys(nomor_inklusi)
                        field.send_keys(Keys.ENTER)

                        xpath_tr = f"//tr[td[text()='{nomor_inklusi}']]"
                        tr_elem = wait_short.until(EC.presence_of_element_located((By.XPATH, xpath_tr)))
                        print("✅ Baris <TR> ditemukan")

                        # --- Cek status ---
                        status_val = tr_elem.find_element(By.XPATH, "./td[7]").text.strip()
                        print(f"📌 Status kolom 7: {status_val}")

                        # --- Jika Query NOT OK, hapus row 3 dan kembali ke awal loop ---
                        if status_val != "Query - OK":
                            tulis_log(log_ws, log_wb, nomor_inklusi, "", [f"Status not OK (got '{status_val}')"])

                            ws.delete_rows(3)
                            wb.save(EXCEL_PATH)
                            print("🗑️ Baris 3 dihapus karena status not OK.")
                            continue  # lanjut loop, cari baris berikutnya

                            # while True:
                            #     try:
                            #         user_choice = input("⚠️ Status 'not OK' ditemukan di baris 3.\nPilih tindakan: [H]apus & cari lagi, [L]anjut script dan tekan ENTER → ").strip().lower()

                            #         if user_choice == 'h':
                            #             ws.delete_rows(3)
                            #             wb.save(EXCEL_PATH)
                            #             print("🗑️ Baris 3 dihapus karena status not OK.")
                            #             continue  # lanjut loop, cari baris berikutnya

                            #         elif user_choice == 'l':
                            #             print("➡️ Lanjut jalanin script tanpa hapus baris.")
                            #             break  # keluar dari loop dan lanjut script utama

                            #         else:
                            #             print("❓ Pilihan tidak dikenal. Ketik 'H' untuk hapus, atau 'L' untuk lanjut.")
                            #             continue

                            #     except PermissionError:
                            #         play_error()
                            #         print(f"⚠️ Tidak bisa menulis ke {EXCEL_PATH}. Tutup Excel dulu.")
                            #         input("➡️ Tekan Enter setelah menutup Excel...")

                        # Buka form
                        if not buka_form1(driver, wait_long, nomor_inklusi):
                            play_error()
                            print("⛔ Tidak bisa melanjutkan karena form gagal dibuka.")
                            break
                        print("Form berhasil dibuka")

                        while True:
                            ulang_dari_awal = False

                            for key in mapping.keys():
                                actual, expected = compare_web_versus_excel(driver, wait_short, wait_long, key, mapping)
                                if ulang_dari_awal:
                                    print("🔁 Form tertutup — ulang dari field 1...\n")
                                    buka_form1(driver, wait_long, nomor_inklusi)
                                    time.sleep(0.5)
                                    break

                            # jika ulang_dari_awal True, ulangi dari atas; kalau False, selesai
                            if ulang_dari_awal:
                                continue
                            else:
                                break

                        print("✅ Semua field selesai dibandingkan dan verified diupdate.\n")
                        break  # keluar dari while loop untuk row ini

                    except Exception as e:
                        play_error()
                        print(f"❌ Error saat memproses nomor inklusi {nomor_inklusi}: {e}")
                        continue

                # # pause sebelum verify
                # ans = input("👉 Cek data, jika sudah OK, tekan ENTER untuk lanjut, atau ketik 'stop' untuk keluar: ").strip().lower()
                # if ans == "stop":
                #     print("⏹️ Proses dihentikan user.")
                #     exit()
                        
                # --- tunggu tombol VERIFY muncul ---
                verify_btn = wait_short.until(EC.element_to_be_clickable((By.ID, "btn-verify")))
                # scroll ke tombol dan klik
                driver.execute_script("arguments[0].scrollIntoView(true);", verify_btn)
                time.sleep(0.2)
                driver.execute_script("arguments[0].click();", verify_btn)
                print("✅ Tombol VERIFY berhasil diklik")

                # --- tunggu tombol KEMBALI muncul ---
                # back_btn = wait_short.until(EC.element_to_be_clickable((By.ID, "btn-close")))
                # driver.execute_script("arguments[0].scrollIntoView(true);", back_btn)
                # time.sleep(0.2)
                # driver.execute_script("arguments[0].click();", back_btn)
                # print("✅ Tombol KEMBALI berhasil diklik")

            except KeyboardInterrupt:
                print("\n🛑 Script dihentikan manual oleh user (CTRL+C).")
            except Exception as e:
                play_error()
                print(f"❌ Error di loop utama: {e}")

            pass
        except Exception as e:
            play_error()
            print(f"❌ Error di {SECTION}: {e}")

        # ===========================================
        # === SCRIPT FILE 2 ===
        # ===========================================
        SECTION = "Inklusi/Eksklusi"
        mapping = get_mapping(SECTION)
        try:
            # TODO: isi script file 2

            print("▶️ Script berjalan... tekan CTRL+C untuk berhenti manual.")
            print("⏳ Menunggu tabel '#tbl_responden' & field siap...")
            try:
                wait_long.until(EC.visibility_of_element_located((By.ID, "tbl_responden")))
                wait_long.until(EC.presence_of_element_located((By.ID, "field_cari")))

                # tunggu sampai spinner / loading hilang
                wait_long.until(EC.invisibility_of_element_located((
                    By.CSS_SELECTOR,
                    ".loading_spinner[style*='block'], .spinner[style*='block'], .loading[style*='block']"
                )))

                time.sleep(0.5)  # beri waktu agar event JS selesai
                print("✅ Halaman siap, lanjut cari nomor inklusi...")
            except TimeoutException:
                play_error()
                print("⚠️ Timeout: tabel atau field_cari tidak muncul. Periksa halaman atau network.")
                exit()

            # --- Cari nomor inklusi ---
            field = driver.find_element(By.ID, "field_cari")
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", field)
            time.sleep(0.3)
            driver.execute_script("arguments[0].click();", field)  # ✅ pakai JS click biar aman
            field.clear()
            field.send_keys(nomor_inklusi)
            field.send_keys(Keys.ENTER)
            time.sleep(1)  # beri waktu untuk hasil tabel muncul
            wait_long.until(EC.presence_of_element_located(
                (By.XPATH, f"//tr[td[normalize-space(text())='{nomor_inklusi}']]")
            ))


            try:
                print(f"🔍 Memproses nomor inklusi: {nomor_inklusi} pada FORM {SECTION}")
                if not buka_form2(driver, wait_long, nomor_inklusi, nama_form=SECTION):
                    play_error()
                    print("⛔ Tidak bisa melanjutkan karena form gagal dibuka.")
                    exit()

                verified_cache = {}
                for key in mapping:
                    try:
                        verified_cache[key] = driver.find_element(By.ID, f"verified_{key}")
                    except:
                        verified_cache[key] = None
                print("✅ Verified cache done")

                mismatches = []

                # --- Loop utama ---
                for key, m in mapping.items():
                    col_index = column_index_from_string(m['col'])
                    cell_value = ws.cell(row=3, column=col_index).value

                    if cell_value is None:
                        expected = ""
                        field_type = "text"
                    elif isinstance(cell_value, dtime):
                        expected = cell_value.strftime("%H:%M")
                        field_type = "time"
                    elif isinstance(cell_value, datetime):
                        expected = cell_value.strftime("%d-%m-%Y")
                        field_type = "date"
                    else:
                        expected = str(cell_value).strip()
                        field_type = "text"

                    try:
                        val_elem = driver.find_element(By.ID, f"val_verified_{key}")
                        current_val = (val_elem.get_attribute("value") or "").strip()
                    except:
                        current_val = ""

                    if expected == "":
                        print(f"⏭️ Skip verify_{key} ({m.get('id') or m.get('radio')}) karena Excel kosong")
                        continue

                    # --- Ambil actual value ---
                    actual = ""
                    if "id" in m:
                        try:
                            el = driver.find_element(By.ID, m["id"])
                            actual_raw = el.get_attribute("value-data") or el.get_attribute("value") or ""
                            actual = normalize_attr(actual_raw)
                        except:
                            actual = ""
                    elif "radio" in m:
                        try:
                            actual = driver.find_element(
                                By.CSS_SELECTOR, f"#{m['radio']} input:checked"
                            ).get_attribute("value").strip()
                        except:
                            actual = ""

                    # --- Ambil nama field dari cache ---
                    field_name = get_label_text(driver, key, m)

                    print(f"🔎 verify_{key} ({field_name}): expected={expected!r}, got={actual!r}")

                    # --- Klik verified pakai cached element ---
                    icon_elem = verified_cache.get(key)
                    if icon_elem:
                        cek_dan_klik_verified_smart(driver, wait_short, key, field_name, expected, actual, current_val, field_type=field_type)
                    else:
                        print(f"⚠️ verified_{key} tidak ditemukan ({field_name})")

                    if actual != expected:
                        mismatches.append(f"{field_name} (expected={expected}, got={actual})")

                if mismatches:
                    tulis_log(log_ws, log_wb, nomor_inklusi, SECTION, mismatches)

                # --- Pause manual ---
                # ans = input("👉 Cek data, jika sudah OK, tekan ENTER untuk lanjut, atau ketik 'stop' untuk keluar: ").strip().lower()
                # if ans == "stop":
                #     print("⏹️ Proses dihentikan user.")
                #     exit()

                # --- tunggu tombol VERIFY muncul ---
                verify_btn = wait_short.until(EC.element_to_be_clickable((By.ID, "btn-verify")))
                # scroll ke tombol dan klik
                driver.execute_script("arguments[0].scrollIntoView(true);", verify_btn)
                time.sleep(0.2)
                driver.execute_script("arguments[0].click();", verify_btn)
                print("✅ Tombol VERIFY berhasil diklik")

                # --- tunggu tombol KEMBALI muncul ---
                # back_btn = wait_short.until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".btn__cancel")))
                # driver.execute_script("arguments[0].scrollIntoView(true);", back_btn)
                # time.sleep(0.2)
                # driver.execute_script("arguments[0].click();", back_btn)
                # print("✅ Tombol KEMBALI berhasil diklik")


            except KeyboardInterrupt:
                print("\n🛑 Script dihentikan manual oleh user (CTRL+C).")
            except Exception as e:
                play_error()
                print(f"❌ Error di verifinklusi: {e}")

            pass
        except Exception as e:
            play_error()
            print(f"❌ Error di {SECTION}: {e}")

        # ===========================================
        # === SCRIPT FILE 3 ===
        # ===========================================
        SECTION = "Informasi Pelapor"
        mapping = get_mapping(SECTION)
        try:
            # TODO: isi script file 3
            print("▶️ Script berjalan... tekan CTRL+C untuk berhenti manual.")
            print("⏳ Menunggu tabel '#tbl_responden' & field siap...")

            try:
                wait_long.until(EC.visibility_of_element_located((By.ID, "tbl_responden")))
                wait_long.until(EC.presence_of_element_located((By.ID, "field_cari")))

                # 🧩 Tambahan penting: tunggu sampai spinner / loading hilang
                wait_long.until(EC.invisibility_of_element_located((
                    By.CSS_SELECTOR,
                    ".loading_spinner[style*='block'], .spinner[style*='block'], .loading[style*='block']"
                )))

                time.sleep(0.5)  # beri waktu agar event JS selesai
                print("✅ Halaman siap, lanjut cari nomor inklusi...")
            except TimeoutException:
                play_error()
                print("⚠️ Timeout: tabel atau field_cari tidak muncul. Periksa halaman atau network.")
                exit()

            # --- Cari nomor inklusi ---
            field = driver.find_element(By.ID, "field_cari")
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", field)
            time.sleep(0.3)
            driver.execute_script("arguments[0].click();", field)  # ✅ pakai JS click biar aman
            field.clear()
            field.send_keys(nomor_inklusi)
            field.send_keys(Keys.ENTER)
            time.sleep(1)  # beri waktu untuk hasil tabel muncul
            wait_long.until(EC.presence_of_element_located(
                (By.XPATH, f"//tr[td[normalize-space(text())='{nomor_inklusi}']]")
            ))

            try:
                print(f"🔍 Memproses nomor inklusi: {nomor_inklusi} pada FORM {SECTION}")
                if not buka_form2(driver, wait_long, nomor_inklusi, nama_form=SECTION):
                    play_error()
                    print("⛔ Tidak bisa melanjutkan karena form gagal dibuka.")
                    exit()

                verified_cache = {}
                for key in mapping:
                    try:
                        verified_cache[key] = driver.find_element(By.ID, f"verified_{key}")
                    except:
                        verified_cache[key] = None
                print("✅ Verified cache done")

                mismatches = []

                # --- Loop utama ---
                for key, m in mapping.items():
                    col_index = column_index_from_string(m['col'])
                    cell_value = ws.cell(row=3, column=col_index).value

                    if cell_value is None:
                        expected = ""
                        field_type = "text"
                    elif isinstance(cell_value, dtime):
                        expected = cell_value.strftime("%H:%M")
                        field_type = "time"
                    elif isinstance(cell_value, datetime):
                        expected = cell_value.strftime("%d-%m-%Y")
                        field_type = "date"
                    else:
                        expected = str(cell_value).strip()
                        field_type = "text"

                    try:
                        val_elem = driver.find_element(By.ID, f"val_verified_{key}")
                        current_val = (val_elem.get_attribute("value") or "").strip()
                    except:
                        current_val = ""

                    if expected == "":
                        print(f"⏭️ Skip verify_{key} ({m.get('id') or m.get('radio')}) karena Excel kosong")
                        continue

                    # --- Ambil actual value ---
                    actual = ""
                    if "id" in m:
                        try:
                            el = driver.find_element(By.ID, m["id"])
                            actual_raw = el.get_attribute("value-data") or el.get_attribute("value") or ""
                            actual = normalize_attr(actual_raw)
                        except:
                            actual = ""
                    elif "radio" in m:
                        try:
                            actual = driver.find_element(
                                By.CSS_SELECTOR, f"#{m['radio']} input:checked"
                            ).get_attribute("value").strip()
                        except:
                            actual = ""

                    # --- Ambil nama field dari cache ---
                    field_name = get_label_text(driver, key, m)

                    print(f"🔎 verify_{key} ({field_name}): expected={expected!r}, got={actual!r}")

                    # --- Klik verified pakai cached element ---
                    icon_elem = verified_cache.get(key)
                    if icon_elem:
                        cek_dan_klik_verified_smart(driver, wait_short, key, field_name, expected, actual, current_val, field_type=field_type)
                    else:
                        print(f"⚠️ verified_{key} tidak ditemukan ({field_name})")

                    if actual != expected:
                        mismatches.append(f"{field_name} (expected={expected}, got={actual})")


                if mismatches:
                    tulis_log(log_ws, log_wb, nomor_inklusi, SECTION, mismatches)

                # --- Pause manual ---
                # ans = input("👉 Cek data, jika sudah OK, tekan ENTER untuk lanjut, atau ketik 'stop' untuk keluar: ").strip().lower()
                # if ans == "stop":
                #     print("⏹️ Proses dihentikan user.")
                #     exit()

                # --- tunggu tombol VERIFY muncul ---
                verify_btn = wait_short.until(EC.element_to_be_clickable((By.ID, "btn-verify")))
                # scroll ke tombol dan klik
                driver.execute_script("arguments[0].scrollIntoView(true);", verify_btn)
                time.sleep(0.2)
                driver.execute_script("arguments[0].click();", verify_btn)
                print("✅ Tombol VERIFY berhasil diklik")

                # --- tunggu tombol KEMBALI muncul ---
                # back_btn = wait_short.until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".btn__cancel")))
                # driver.execute_script("arguments[0].scrollIntoView(true);", back_btn)
                # time.sleep(0.2)
                # driver.execute_script("arguments[0].click();", back_btn)
                # print("✅ Tombol KEMBALI berhasil diklik")


            except KeyboardInterrupt:
                print("\n🛑 Script dihentikan manual oleh user (CTRL+C).")
            except Exception as e:
                play_error()
                print(f"❌ Error di verifpelapor: {e}")

            pass
        except Exception as e:
            play_error()
            print(f"❌ Error di {SECTION}: {e}")

        # ===========================================
        # === SCRIPT FILE 4 ===
        # ===========================================
        SECTION = "Informasi Pasien"
        mapping = get_mapping(SECTION)
        try:
            # TODO: isi script file 4
            print("▶️ Script berjalan... tekan CTRL+C untuk berhenti manual.")
            print("⏳ Menunggu tabel '#tbl_responden' & field siap...")

            try:
                wait_long.until(EC.visibility_of_element_located((By.ID, "tbl_responden")))
                wait_long.until(EC.presence_of_element_located((By.ID, "field_cari")))

                # 🧩 Tambahan penting: tunggu sampai spinner / loading hilang
                wait_long.until(EC.invisibility_of_element_located((
                    By.CSS_SELECTOR,
                    ".loading_spinner[style*='block'], .spinner[style*='block'], .loading[style*='block']"
                )))

                time.sleep(0.5)  # beri waktu agar event JS selesai
                print("✅ Halaman siap, lanjut cari nomor inklusi...")
            except TimeoutException:
                play_error()
                print("⚠️ Timeout: tabel atau field_cari tidak muncul. Periksa halaman atau network.")
                exit()

            # --- Cari nomor inklusi ---
            field = driver.find_element(By.ID, "field_cari")
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", field)
            time.sleep(0.3)
            driver.execute_script("arguments[0].click();", field)  # ✅ pakai JS click biar aman
            field.clear()
            field.send_keys(nomor_inklusi)
            field.send_keys(Keys.ENTER)
            time.sleep(1)  # beri waktu untuk hasil tabel muncul
            wait_long.until(EC.presence_of_element_located(
                (By.XPATH, f"//tr[td[normalize-space(text())='{nomor_inklusi}']]")
            ))

            try:
                print(f"🔍 Memproses nomor inklusi: {nomor_inklusi} pada FORM {SECTION}")
                if not buka_form2(driver, wait_long, nomor_inklusi, nama_form=SECTION):
                    play_error()
                    print("⛔ Tidak bisa melanjutkan karena form gagal dibuka.")
                    exit()

                verified_cache = {}
                for key in mapping:
                    try:
                        verified_cache[key] = driver.find_element(By.ID, f"verified_{key}")
                    except:
                        verified_cache[key] = None
                print("✅ Verified cache done")

                mismatches = []

                # --- Loop utama ---
                for key, m in mapping.items():
                    col_index = column_index_from_string(m['col'])
                    cell_value = ws.cell(row=3, column=col_index).value

                    if cell_value is None:
                        expected = ""
                        field_type = "text"
                    elif isinstance(cell_value, dtime):
                        expected = cell_value.strftime("%H:%M")
                        field_type = "time"
                    elif isinstance(cell_value, datetime):
                        expected = cell_value.strftime("%d-%m-%Y")
                        field_type = "date"
                    else:
                        expected = str(cell_value).strip()
                        field_type = "text"

                    try:
                        val_elem = driver.find_element(By.ID, f"val_verified_{key}")
                        current_val = (val_elem.get_attribute("value") or "").strip()
                    except:
                        current_val = ""

                    if expected == "":
                        print(f"⏭️ Skip verify_{key} ({m.get('id') or m.get('radio')}) karena Excel kosong")
                        continue

                    # --- Ambil actual value ---
                    actual = ""
                    if "id" in m:
                        try:
                            el = driver.find_element(By.ID, m["id"])
                            actual_raw = el.get_attribute("value-data") or el.get_attribute("value") or ""
                            actual = normalize_attr(actual_raw)
                        except:
                            actual = ""
                    elif "radio" in m:
                        try:
                            actual = driver.find_element(
                                By.CSS_SELECTOR, f"#{m['radio']} input:checked"
                            ).get_attribute("value").strip()
                        except:
                            actual = ""

                    # --- Ambil nama field dari cache ---
                    field_name = get_label_text(driver, key, m)

                    print(f"🔎 verify_{key} ({field_name}): expected={expected!r}, got={actual!r}")

                    # --- Klik verified pakai cached element ---
                    icon_elem = verified_cache.get(key)
                    if icon_elem:
                        cek_dan_klik_verified_smart(driver, wait_short, key, field_name, expected, actual, current_val, field_type=field_type)
                    else:
                        print(f"⚠️ verified_{key} tidak ditemukan ({field_name})")

                    if actual != expected:
                        mismatches.append(f"{field_name} (expected={expected}, got={actual})")


                if mismatches:
                    tulis_log(log_ws, log_wb, nomor_inklusi, SECTION, mismatches)

                # --- Pause manual ---
                # ans = input("👉 Cek data, jika sudah OK, tekan ENTER untuk lanjut, atau ketik 'stop' untuk keluar: ").strip().lower()
                # if ans == "stop":
                #     print("⏹️ Proses dihentikan user.")
                #     exit()

                # --- tunggu tombol VERIFY muncul ---
                verify_btn = wait_short.until(EC.element_to_be_clickable((By.ID, "btn-verify")))
                # scroll ke tombol dan klik
                driver.execute_script("arguments[0].scrollIntoView(true);", verify_btn)
                time.sleep(0.2)
                driver.execute_script("arguments[0].click();", verify_btn)
                print("✅ Tombol VERIFY berhasil diklik")

                # --- tunggu tombol KEMBALI muncul ---
                # back_btn = wait_short.until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".btn__cancel")))
                # driver.execute_script("arguments[0].scrollIntoView(true);", back_btn)
                # time.sleep(0.2)
                # driver.execute_script("arguments[0].click();", back_btn)
                # print("✅ Tombol KEMBALI berhasil diklik")


            except KeyboardInterrupt:
                print("\n🛑 Script dihentikan manual oleh user (CTRL+C).")
            except Exception as e:
                play_error()
                print(f"❌ Error di verifpasien: {e}")

            pass
        except Exception as e:
            play_error()
            print(f"❌ Error di {SECTION}: {e}")

        # ===========================================
        # === SCRIPT FILE 5 ===
        # ===========================================
        SECTION = "Data Vaksinasi"
        mapping = get_mapping(SECTION)
        try:
            # TODO: isi script file 5
            print("▶️ Script berjalan... tekan CTRL+C untuk berhenti manual.")
            print("⏳ Menunggu tabel '#tbl_responden' & field siap...")
            try:
                wait_long.until(EC.visibility_of_element_located((By.ID, "tbl_responden")))
                wait_long.until(EC.presence_of_element_located((By.ID, "field_cari")))

                # 🧩 Tambahan penting: tunggu sampai spinner / loading hilang
                wait_long.until(EC.invisibility_of_element_located((
                    By.CSS_SELECTOR,
                    ".loading_spinner[style*='block'], .spinner[style*='block'], .loading[style*='block']"
                )))

                time.sleep(0.5)  # beri waktu agar event JS selesai
                print("✅ Halaman siap, lanjut cari nomor inklusi...")
            except TimeoutException:
                play_error()
                print("⚠️ Timeout: tabel atau field_cari tidak muncul. Periksa halaman atau network.")
                exit()

            # --- Cari nomor inklusi ---
            field = driver.find_element(By.ID, "field_cari")
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", field)
            time.sleep(0.3)
            driver.execute_script("arguments[0].click();", field)  # ✅ pakai JS click biar aman
            field.clear()
            field.send_keys(nomor_inklusi)
            field.send_keys(Keys.ENTER)
            time.sleep(1)  # beri waktu untuk hasil tabel muncul
            wait_long.until(EC.presence_of_element_located(
                (By.XPATH, f"//tr[td[normalize-space(text())='{nomor_inklusi}']]")
            ))

            try:
                print(f"🔍 Memproses nomor inklusi: {nomor_inklusi} pada FORM {SECTION}")
                if not buka_form2(driver, wait_long, nomor_inklusi, nama_form=SECTION):
                    play_error()
                    print("⛔ Tidak bisa melanjutkan karena form gagal dibuka.")
                    exit()

                verified_cache = {}
                for key in mapping:
                    try:
                        verified_cache[key] = driver.find_element(By.ID, f"verified_{key}")
                    except:
                        verified_cache[key] = None
                print("✅ Verified cache done")

                mismatches = []

                # --- Loop utama ---
                for key, m in mapping.items():
                    col_index = column_index_from_string(m['col'])
                    cell_value = ws.cell(row=3, column=col_index).value

                    if cell_value is None:
                        expected = ""
                        field_type = "text"
                    elif str(cell_value).strip() in ["NA", "N/A"]:
                        expected = "NA"   # atau "N/A"
                        field_type = "text"
                    elif isinstance(cell_value, dtime):
                        expected = cell_value.strftime("%H:%M")
                        field_type = "time"
                    elif isinstance(cell_value, datetime):
                        expected = cell_value.strftime("%d-%m-%Y")
                        field_type = "date"
                    else:
                        expected = str(cell_value).strip()
                        field_type = "text"

                    try:
                        val_elem = driver.find_element(By.ID, f"val_verified_{key}")
                        current_val = (val_elem.get_attribute("value") or "").strip()
                    except:
                        current_val = ""
                    
                    if expected == "":
                        print(f"⏭️ Skip verify_{key} ({m.get('id') or m.get('radio')}) karena Excel kosong")
                        continue

                    # --- Ambil actual value ---
                    actual = ""
                    if "id" in m:
                        try:
                            el = driver.find_element(By.ID, m["id"])
                            actual_raw = el.get_attribute("value-data") or el.get_attribute("value") or ""
                            actual = normalize_attr(actual_raw).strip()
                        except:
                            actual = ""
                    elif "radio" in m:
                        try:
                            actual = driver.find_element(
                                By.CSS_SELECTOR, f"#{m['radio']} input:checked"
                            ).get_attribute("value").strip()
                        except:
                            actual = ""

                    # Normalisasi nilai dari web
                    if actual in ["N/A", "NA"]:
                        actual = "NA"
                    
                    # --- Normalisasi actual agar case-insensitive ---
                    if isinstance(actual, str):
                        actual = actual.casefold().strip()
                    if isinstance(expected, str):
                        expected = expected.casefold().strip()

                    # --- Ambil nama field dari cache ---
                    field_name = get_label_text(driver, key, m)

                    print(f"🔎 verify_{key} ({field_name}): expected={expected!r}, got={actual!r}")

                    # --- Klik verified pakai cached element ---
                    icon_elem = verified_cache.get(key)
                    if icon_elem:
                        cek_dan_klik_verified_smart(driver, wait_short, key, field_name, expected, actual, current_val, field_type=field_type)
                    else:
                        print(f"⚠️ verified_{key} tidak ditemukan ({field_name})")

                    if actual != expected:
                        mismatches.append(f"{field_name} (expected={expected}, got={actual})")


                if mismatches:
                    tulis_log(log_ws, log_wb, nomor_inklusi, SECTION, mismatches)

                # --- Pause manual ---
                # ans = input("👉 Cek data, jika sudah OK, tekan ENTER untuk lanjut, atau ketik 'stop' untuk keluar: ").strip().lower()
                # if ans == "stop":
                #     print("⏹️ Proses dihentikan user.")
                #     exit()

                # --- tunggu tombol VERIFY muncul ---
                verify_btn = wait_short.until(EC.element_to_be_clickable((By.ID, "btn-verify")))
                # scroll ke tombol dan klik
                driver.execute_script("arguments[0].scrollIntoView(true);", verify_btn)
                time.sleep(0.2)
                driver.execute_script("arguments[0].click();", verify_btn)
                print("✅ Tombol VERIFY berhasil diklik")

                # --- tunggu tombol KEMBALI muncul ---
                # back_btn = wait_short.until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".btn__cancel")))
                # driver.execute_script("arguments[0].scrollIntoView(true);", back_btn)
                # time.sleep(0.2)
                # driver.execute_script("arguments[0].click();", back_btn)
                # print("✅ Tombol KEMBALI berhasil diklik")


            except KeyboardInterrupt:
                print("\n🛑 Script dihentikan manual oleh user (CTRL+C).")
            except Exception as e:
                play_error()
                print(f"❌ Error di verifvaksinasi: {e}")

            pass
        except Exception as e:
            play_error()
            print(f"❌ Error di {SECTION}: {e}")

        # ===========================================
        # === SCRIPT FILE 6 ===
        # ===========================================
        SECTION = "KIPI"
        mapping = get_mapping(SECTION)
        try:
            # TODO: isi script file 6
            print("▶️ Script berjalan... tekan CTRL+C untuk berhenti manual.")
            print("⏳ Menunggu tabel '#tbl_responden' & field siap...")
            try:
                wait_long.until(EC.visibility_of_element_located((By.ID, "tbl_responden")))
                wait_long.until(EC.presence_of_element_located((By.ID, "field_cari")))

                # tunggu sampai spinner / loading hilang
                wait_long.until(EC.invisibility_of_element_located((
                    By.CSS_SELECTOR,
                    ".loading_spinner[style*='block'], .spinner[style*='block'], .loading[style*='block']"
                )))

                time.sleep(0.5)  # beri waktu agar event JS selesai
                print("✅ Halaman siap, lanjut cari nomor inklusi...")
            except TimeoutException:
                play_error()
                print("⚠️ Timeout: tabel atau field_cari tidak muncul. Periksa halaman atau network.")
                exit()

            # --- Cari nomor inklusi ---
            field = driver.find_element(By.ID, "field_cari")
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", field)
            time.sleep(0.3)
            driver.execute_script("arguments[0].click();", field)  # ✅ pakai JS click biar aman
            field.clear()
            field.send_keys(nomor_inklusi)
            field.send_keys(Keys.ENTER)
            time.sleep(1)  # beri waktu untuk hasil tabel muncul
            wait_long.until(EC.presence_of_element_located(
                (By.XPATH, f"//tr[td[normalize-space(text())='{nomor_inklusi}']]")
            ))

            try:
                print(f"🔍 Memproses nomor inklusi: {nomor_inklusi} pada FORM {SECTION}")
                if not buka_form2(driver, wait_long, nomor_inklusi, nama_form=SECTION):
                    play_error()
                    print("⛔ Tidak bisa melanjutkan karena form gagal dibuka.")
                    exit()

                try:
                    print("✅ field_cari siap, lanjut buka form...")
                    print("⏳ Menunggu form utama (surveyform) siap...")
                    form_elem = wait_long.until(EC.element_to_be_clickable((By.ID, "surveyform")))
                    wait_short.until(EC.element_to_be_clickable((By.ID, "surveyform")))
                    print("✅ Form 'surveyform' siap, script mulai dijalankan.")
                except TimeoutException:
                    print("⚠️ Timeout: Form 'surveyform' tidak muncul atau belum siap, lanjut dengan hati-hati...")

                verified_cache = {}
                for key in mapping:
                    try:
                        verified_cache[key] = driver.find_element(By.ID, f"verified_{key}")
                    except:
                        verified_cache[key] = None
                print("✅ Verified cache done")

                mismatches = []

                # --- Loop utama ---
                for key, m in mapping.items():
                    col_index = column_index_from_string(m['col'])
                    cell_value = ws.cell(row=3, column=col_index).value

                    if cell_value is None:
                        expected = ""
                        field_type = "text"
                    elif isinstance(cell_value, dtime):
                        expected = cell_value.strftime("%H:%M")
                        field_type = "time"
                    elif isinstance(cell_value, datetime):
                        expected = cell_value.strftime("%d-%m-%Y")
                        field_type = "date"
                    else:
                        expected = str(cell_value).strip()
                        field_type = "text"

                    try:
                        val_elem = driver.find_element(By.ID, f"val_verified_{key}")
                        current_val = (val_elem.get_attribute("value") or "").strip()
                    except:
                        current_val = ""

                    if expected == "":
                        print(f"⏭️ Skip verify_{key} ({m.get('id') or m.get('radio')}) karena Excel kosong")
                        continue

                    # --- Ambil actual value ---
                    actual = ""
                    if "id" in m:
                        try:
                            el = driver.find_element(By.ID, m["id"])
                            actual_raw = el.get_attribute("value-data") or el.get_attribute("value") or ""
                            actual = normalize_attr(actual_raw)
                        except:
                            actual = ""
                    elif "radio" in m:
                        try:
                            actual = driver.find_element(
                                By.CSS_SELECTOR, f"#{m['radio']} input:checked"
                            ).get_attribute("value").strip()
                        except:
                            actual = ""
                    
                    # --- Normalisasi actual agar case-insensitive ---
                    if isinstance(actual, str):
                        actual = actual.casefold().strip()
                    if isinstance(expected, str):
                        expected = expected.casefold().strip()


                    # --- Ambil nama field dari cache ---
                    field_name = get_label_text(driver, key, m)

                    print(f"🔎 verify_{key} ({field_name}): expected={expected!r}, got={actual!r}")

                    # --- Klik verified pakai cached element ---
                    icon_elem = verified_cache.get(key)
                    if icon_elem:
                        cek_dan_klik_verified_smart(driver, wait_short, key, field_name, expected, actual, current_val, field_type=field_type)           
                    else:
                        print(f"⚠️ verified_{key} tidak ditemukan ({field_name})")

                    if actual != expected:
                        mismatches.append(f"{field_name} (expected={expected}, got={actual})")

                if mismatches:
                    tulis_log(log_ws, log_wb, nomor_inklusi, SECTION, mismatches)
                
                tulis_verif(log_ws, log_wb, nomor_inklusi)

                # # --- Pause manual ---
                # ans = input("👉 Cek data, jika sudah OK, tekan ENTER untuk lanjut, atau ketik 'stop' untuk keluar: ").strip().lower()
                # if ans == "stop":
                #     print("⏹️ Proses dihentikan user.")
                #     exit()

                # --- tunggu tombol VERIFY muncul ---
                verify_btn = wait_short.until(EC.element_to_be_clickable((By.ID, "btn-verify")))
                # scroll ke tombol dan klik
                driver.execute_script("arguments[0].scrollIntoView(true);", verify_btn)
                time.sleep(0.2)
                driver.execute_script("arguments[0].click();", verify_btn)
                print("✅ Tombol VERIFY berhasil diklik")

                # --- tunggu tombol KEMBALI muncul ---
                # back_btn = wait_short.until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".btn__cancel")))
                # driver.execute_script("arguments[0].scrollIntoView(true);", back_btn)
                # time.sleep(0.2)
                # driver.execute_script("arguments[0].click();", back_btn)
                # print("✅ Tombol KEMBALI berhasil diklik")
                
                
            except KeyboardInterrupt:
                print("\n🛑 Script dihentikan manual oleh user (CTRL+C).")
            except Exception as e:
                play_error()
                print(f"❌ Error di verifkipi: {e}")

            pass
        except Exception as e:
            play_error()
            print(f"❌ Error di {SECTION}: {e}")

        # === Hapus baris 3 ===
        try:
            ws.delete_rows(EXCEL_ROW)
            wb.save(EXCEL_PATH)
            wb.close()
            log_wb.close()
            print("🗑️ Baris ke-3 dihapus dan disimpan.\n")
        except Exception as e:
            play_error()
            print(f"⚠️ Gagal hapus baris: {e}")

        # jeda antar loop
        print("🕓 Menunggu 0.3 detik sebelum lanjut ke data berikutnya...\n")
        time.sleep(0.3)

    driver.quit()

if __name__ == "__main__":
    main()


