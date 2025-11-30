import subprocess
import sys
import time, os
from openpyxl import load_workbook
import openpyxl
from openpyxl.utils import column_index_from_string
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from datetime import date, datetime
from verif_utils import (
    create_driver, load_excel_data, EXCEL_PATH, get_mapping, buka_form1, wait_for_field_value,
    parse_date_indonesia, normalize_date_flex, sudah_ada_comment, cek_dan_klik_verified,
    tulis_log, tulis_comment, play_error, wait_excel_closed, load_excel_log
)

driver, wait_long, wait_short = create_driver()
wait_excel_closed()
wb, ws, nomor_inklusi = load_excel_data()
log_wb, log_ws = load_excel_log()

SECTION = "responden"
mapping = get_mapping(SECTION)
ulang_dari_awal = False


###################
##### HELPER ######
###################

def compare_web_versus_excel(driver, wait_short, wait_long, key, mapping):
    MIN_DATE = date(2025, 9, 11)
    TODAY = date.today()
    mismatches = []
    global ulang_dari_awal

    try:
        # --- FIELD KHUSUS 5 & 6 (Tanggal IC dan Enrollment) ---
        if key in (5, 6):
            actual5 = wait_for_field_value(driver, wait_short, "verify_5")
            actual6 = wait_for_field_value(driver, wait_short, "verify_6")
            tgl5 = parse_date_indonesia(actual5)
            tgl6 = parse_date_indonesia(actual6)

            # Validasi range tanggal
            in_range5 = tgl5 and (MIN_DATE <= tgl5 <= TODAY)
            in_range6 = tgl6 and (MIN_DATE <= tgl6 <= TODAY)
            should_be_5, should_be_6 = "1", "1"  # default lolos semua

            if not in_range5:
                should_be_5 = "0"

                if not sudah_ada_comment(driver, 5):
                    if actual5 == "" or actual5 is None:
                        comment_text = "Wajib diisi."
                    else:
                        comment_text = f"Tanggal Informed Consent di luar range yang diperbolehkan: 12-Sep-2025 s.d. Today."

                    mismatches.append(f"Tanggal Informed Consent di luar range: {actual5}")
                    tulis_comment(driver, wait_long, 5, comment_text)
                    tulis_log(log_ws, log_wb, nomor_inklusi, SECTION, mismatches)
                    ulang_dari_awal = True
                    buka_form1(driver, wait_long, nomor_inklusi)
                    return None, None

            if not in_range6:
                should_be_6 = "0"

                if not sudah_ada_comment(driver, 6):
                    if actual6 == "" or actual6 is None:
                        comment_text = "Wajib diisi."
                    else:
                        comment_text = f"Tanggal Enrollment di luar range yang diperbolehkan 12-Sep-2025 s.d. Today."

                    mismatches.append(f"Tanggal Enrollment di luar range: {actual6}")
                    tulis_comment(driver, wait_long, 6, comment_text)
                    tulis_log(nomor_inklusi, SECTION, mismatches)
                    ulang_dari_awal = True
                    buka_form1(driver, wait_long, nomor_inklusi)
                    return None, None


            # --- CASE 2: Dua tanggal beda ---
            elif tgl5 != tgl6:
                should_be_5, should_be_6 = "0", "0"

                if not sudah_ada_comment(driver, 5):
                    if actual5 == "" or actual5 is None:
                        comment_text = "Wajib diisi."
                    else:
                        comment_text = "Tanggal Informed Consent tidak sama dengan Tanggal Enrollment."

                    mismatches.append(f"Tanggal Informed Consent ≠ Tanggal Enrollment ({actual5} ≠ {actual6})")
                    tulis_comment(driver, wait_long, 5, comment_text)
                    tulis_log(log_ws, log_wb, nomor_inklusi, SECTION, mismatches)
                    ulang_dari_awal = True
                    buka_form1(driver, wait_long, nomor_inklusi)
                    return None, None

                if not sudah_ada_comment(driver, 6):
                    if actual6 == "" or actual6 is None:
                        comment_text = "Wajib diisi."
                    else:
                        comment_text = "Tanggal Enrollment tidak sama dengan Tanggal Informed Consent."

                    mismatches.append(f"Tanggal Enrollment ≠ Tanggal Informed Consent ({actual6} ≠ {actual5})")
                    tulis_comment(driver, wait_long, 6, comment_text)
                    tulis_log(log_ws, log_wb, nomor_inklusi, SECTION, mismatches)
                    ulang_dari_awal = True
                    buka_form1(driver, wait_long, nomor_inklusi)
                    return None, None

            # --- CASE 3: Valid & sama ---
            else:
                print(f"✅ Tanggal Informed Consent & Enrollment valid dan sama: {actual5}")

            # Update verified setelah semua cek selesai
            cek_dan_klik_verified(driver, wait_short, 5, should_be_5, "Tanggal Informed Consent")
            cek_dan_klik_verified(driver, wait_short, 6, should_be_6, "Tanggal Enrollment")

            return actual5, actual6

        # ==========================================================
        # === FIELD LAIN (Selain 5 & 6)
        # ==========================================================
        m = mapping[key]
        col_index = column_index_from_string(m['col'])
        cell_value = ws.cell(row=3, column=col_index).value
        expected = ""

        if isinstance(cell_value, datetime):
            expected = f"{cell_value.day} {cell_value.strftime('%B %Y')}"
        elif cell_value:
            expected = str(cell_value).strip()

        verify_elem = driver.find_element(By.ID, f"verify_{key}")
        field_data = verify_elem.get_attribute("field-data") or f"Field {key}"

        # --- Ambil nilai actual ---
        actual = ""
        if "id" in m:
            actual = (verify_elem.get_attribute("value-data") or "").strip()
        elif "radio" in m:
            for rb in driver.find_elements(By.NAME, m["radio"]):
                if rb.is_selected():
                    actual = (rb.get_attribute("value") or "").strip()
                    break
        elif "force" in m:
            actual = expected

        # --- Normalisasi tanggal (kalau nama field mengandung "Tanggal") ---
        if "Tanggal" in field_data or "tgl" in field_data.lower():
            actual = normalize_date_flex(actual)
            expected = normalize_date_flex(expected)

        # --- Case-insensitive comparison ---
        if isinstance(actual, str):
            actual = actual.casefold().strip()
        if isinstance(expected, str):
            expected = expected.casefold().strip()

        # --- Tentukan hasil perbandingan ---
        should_be = "1" if actual == expected else "0"

        if actual == expected:
            print(f"[Field {key}] {field_data} ✅ DATA SAMA → Actual: {actual}")
        else:
            print(f"[Field {key}] {field_data} ❌ DATA TIDAK SAMA → Actual: {actual}, Expected: {expected}")

            if sudah_ada_comment(driver, key):
                print(f"💬 Comment untuk field {field_data} sudah ada, skip buka modal.")
            else:
                # 💡 Tentukan isi comment berdasarkan kondisi actual
                if actual == "" or actual is None:
                    comment_text = "Wajib diisi."
                else:
                    comment_text = "Data yg di-input tidak sesuai dengan nilai di Excel atau aturan penulisan."
                mismatches.append(f"{field_data} (expected={expected}, got={actual})")
                tulis_comment(driver, wait_long, key, comment_text)
                tulis_log(log_ws, log_wb, nomor_inklusi, SECTION, mismatches)
                ulang_dari_awal = True
                buka_form1(driver, wait_long, nomor_inklusi)
                return None, None

        # Update verified
        cek_dan_klik_verified(driver, wait_short, key, should_be, field_data)

        # Simpan log bila ada mismatch
        if mismatches:
            tulis_log(log_ws, log_wb, nomor_inklusi, SECTION, mismatches)

        return actual, expected

    except Exception as e:
        play_error()
        print(f"⚠️ Error di field {key}: {e}")
        return None, None


###################
##### END    ######
###################

try:
    print("⏳ Menunggu tabel '#tbl_responden' siap...")
    wait_long.until(EC.visibility_of_element_located((By.ID, "tbl_responden")))
    
    print("⏳ Menunggu input field 'field_cari' siap...")
    field = wait_long.until(EC.element_to_be_clickable((By.ID, "field_cari")))
    
    print("✅ Tabel & field siap, lanjut buka form...")
except TimeoutException:
    play_error()
    print("⚠️ Timeout: tabel atau field_cari tidak muncul. Periksa halaman atau network.")
    field = None


# --- Pastikan Excel bisa diakses sebelum mulai ---
while True:
    try:
        with open(EXCEL_PATH, 'a'):
            pass
        break
    except PermissionError:
        play_error()
        print(f"⚠️ File {EXCEL_PATH} sedang terbuka. Tutup Excel dulu.")
        input("➡️ Tekan Enter setelah menutup Excel...")

# --- Loop utama untuk memproses semua field 1–9 ---
try:
    while True:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active

        # Stop jika tidak ada data
        if ws.max_row < 3:
            play_error()
            print("✅ Semua data selesai diproses")
            break

        nomor_inklusi = ws['K3'].value
        print(f"🔍 Memproses nomor inklusi: {nomor_inklusi} pada FORM {SECTION}")

        try:
            # Pastikan spinner hilang dulu sebelum klik field_cari
            try:
                wait_long.until_not(
                    EC.presence_of_element_located((By.CSS_SELECTOR, ".loading_spinner[style*='display: block']"))
                )
                print("✅ Spinner hilang, aman untuk klik field_cari.")
            except:
                play_error()
                print("⚠️ Spinner tidak muncul atau sudah hilang sebelumnya.")
            
            # Tambahan kecil: delay singkat agar halaman benar-benar stabil
            time.sleep(0.3)

            # Cari nomor inklusi di web
            field = wait_long.until(EC.element_to_be_clickable((By.ID, "field_cari")))
            field.click()
            field.clear()
            field.send_keys(nomor_inklusi)
            field.send_keys(Keys.ENTER)

            xpath_tr = f"//tr[td[text()='{nomor_inklusi}']]"
            tr_elem = wait_short.until(EC.presence_of_element_located((By.XPATH, xpath_tr)))
            print("✅ Baris <TR> ditemukan")

            # --- Cek status ---
            status_val = tr_elem.find_element(By.XPATH, "./td[7]").text.strip()
            print(f"📌 Status kolom 7: {status_val}")

            # --- Jika Query NOT OK, hapus row 3 dan kembali ke awal loop ---
            if status_val != "Query - OK":
                tulis_log(log_ws, log_wb, nomor_inklusi, "", [f"Status not OK (got '{status_val}')"])

                while True:
                    try:
                        user_choice = input("⚠️ Status 'not OK' ditemukan di baris 3.\nPilih tindakan: [H]apus & cari lagi, [L]anjut script dan tekan ENTER → ").strip().lower()

                        if user_choice == 'h':
                            ws.delete_rows(3)
                            wb.save(EXCEL_PATH)
                            print("🗑️ Baris 3 dihapus karena status not OK.")
                            continue  # lanjut loop, cari baris berikutnya

                        elif user_choice == 'l':
                            print("➡️ Lanjut jalanin script tanpa hapus baris.")
                            break  # keluar dari loop dan lanjut script utama

                        else:
                            print("❓ Pilihan tidak dikenal. Ketik 'H' untuk hapus, atau 'L' untuk lanjut.")
                            continue

                    except PermissionError:
                        play_error()
                        print(f"⚠️ Tidak bisa menulis ke {EXCEL_PATH}. Tutup Excel dulu.")
                        input("➡️ Tekan Enter setelah menutup Excel...")

            # Buka form
            if not buka_form1(driver, wait_long, nomor_inklusi):
                play_error()
                print("⛔ Tidak bisa melanjutkan karena form gagal dibuka.")
                break
            print("Form berhasil dibuka")

            while True:
                ulang_dari_awal = False

                for key in mapping.keys():
                    actual, expected = compare_web_versus_excel(driver, wait_short, wait_long, key, mapping)
                    if ulang_dari_awal:
                        print("🔁 Form tertutup — ulang dari field 1...\n")
                        buka_form1(driver, wait_long, nomor_inklusi)
                        time.sleep(0.5)
                        break

                # jika ulang_dari_awal True, ulangi dari atas; kalau False, selesai
                if ulang_dari_awal:
                    continue
                else:
                    break


            print("✅ Semua field selesai dibandingkan dan verified diupdate.\n")
            break  # keluar dari while loop untuk row ini

        except Exception as e:
            play_error()
            print(f"❌ Error saat memproses nomor inklusi {nomor_inklusi}: {e}")
            continue

    # # pause sebelum verify
    ans = input("👉 Cek data, jika sudah OK, tekan ENTER untuk lanjut, atau ketik 'stop' untuk keluar: ").strip().lower()
    if ans == "stop":
        print("⏹️ Proses dihentikan user.")
        exit()
            
    # --- tunggu tombol VERIFY muncul ---
    verify_btn = wait_short.until(EC.element_to_be_clickable((By.ID, "btn-verify")))
    # scroll ke tombol dan klik
    driver.execute_script("arguments[0].scrollIntoView(true);", verify_btn)
    time.sleep(0.2)
    driver.execute_script("arguments[0].click();", verify_btn)
    print("✅ Tombol VERIFY berhasil diklik")

    # --- tunggu tombol KEMBALI muncul ---
    # back_btn = wait_short.until(EC.element_to_be_clickable((By.ID, "btn-close")))
    # driver.execute_script("arguments[0].scrollIntoView(true);", back_btn)
    # time.sleep(0.2)
    # driver.execute_script("arguments[0].click();", back_btn)
    # print("✅ Tombol KEMBALI berhasil diklik")


    # ---  LANJUT KE FORM BERIKUTNYA 
    next_file = "verifinklusi.exe" if getattr(sys, 'frozen', False) else "verifinklusi.py"
    if next_file.endswith(".exe"):
        subprocess.run([next_file])
    else:
        subprocess.run([sys.executable, next_file])

except KeyboardInterrupt:
    print("\n🛑 Script dihentikan manual oleh user (CTRL+C).")
except Exception as e:
    play_error()
    print(f"❌ Error di loop utama: {e}")


