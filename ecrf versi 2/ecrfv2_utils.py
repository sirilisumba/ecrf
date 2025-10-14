import openpyxl
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import time
import datetime
import winsound

EXCEL_PATH = "data.xlsx"
EXCEL_ROW = 3

def create_driver(debugger_address="127.0.0.1:9222", chromedriver_path="chromedriver.exe", wait_long=60, wait_short=10):
    options = Options()
    options.debugger_address = debugger_address
    service = Service(chromedriver_path)
    driver = webdriver.Chrome(service=service, options=options)
    
    wait_long_obj = WebDriverWait(driver, wait_long)
    wait_short_obj = WebDriverWait(driver, wait_short)
    
    return driver, wait_long_obj, wait_short_obj

def is_row_empty(ws, row_num):
    for cell in ws[row_num]:
        if cell.value is not None and str(cell.value).strip() != "":
            return False
    return True

def play_sound():
    duration = 500  # ms
    freq = 1000  # Hz
    for _ in range(3):
        winsound.Beep(freq, duration)
        time.sleep(0.2)

def load_excel_data(path=EXCEL_PATH, sheet_name=None, row_num=EXCEL_ROW):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    if is_row_empty(ws, row_num):
        raise Exception("Baris 3 KOSONG, copy data dulu!")
        play_sound()

    data = {
        "val_puskesmas"             : ws['B3'].value,
        "val_inisial"               : ws['C3'].value,
        "val_tgllahir"              : ws['D3'].value,
        "val_jeniskelamin"          : ws['E3'].value,
        "val_tglscreening"          : ws['F3'].value,
        "val_radio1"                : ws["G3"].value,
        "val_radio2"                : ws["H3"].value,
        "val_radio3"                : ws["I3"].value,
        "val_tgl_inklusi"           : ws["J3"].value,
        "val_no_inklusi"            : ws["K3"].value,
        "val_keterangan"            : ws["L3"].value,
        "val_provinsi"              : ws["M3"].value,
        "val_tgl_lapor"             : ws["N3"].value,
        "val_hasPengobatan"         : ws["O3"].value,
        "val_usia_thn"              : ws["P3"].value,
        "val_usia_bln"              : ws["Q3"].value,
        "val_usia_hr"               : ws["R3"].value,
        "val_jenis_vaksin"          : ws["S3"].value,
        "val_manufaktur"            : ws["T3"].value,
        "val_no_batch"              : ws["U3"].value,
        "val_dosis"                 : ws["V3"].value,
        "val_tgl_vaksin"            : ws["W3"].value,
        "val_wkt_vaksin"            : ws["X3"].value,
        "val_tempat_vaksin"         : ws["Y3"].value,
        "val_vaksin_lain"           : ws["Z3"].value,
        "val_vaksin_lain1"          : ws["AA3"].value,
        "val_vaksin_lain2"          : ws["AB3"].value,
        "val_vaksin_lain3"          : ws["AC3"].value,
        "val_haspengobatan"         : ws["O3"].value,
        "val_kategori"              : ws["AD3"].value,
        "val_lokal"                 : ws["AE3"].value,
        "val_nyeri"                 : ws["AF3"].value,
        "val_lokal_tgl"             : ws["AG3"].value,
        "val_lokal_wkt"             : ws["AH3"].value,
        "val_lokal_hr"              : ws["AI3"].value,
        "val_merah"                 : ws["AJ3"].value,
        "val_merah_tgl"             : ws["AK3"].value,
        "val_merah_wkt"             : ws["AL3"].value,
        "val_merah_hr"              : ws["AM3"].value,
        "val_tebal"                 : ws["AN3"].value,
        "val_tebal_tgl"             : ws["AO3"].value,
        "val_tebal_wkt"             : ws["AP3"].value,
        "val_tebal_hr"              : ws["AQ3"].value,
        "val_bengkak"               : ws["AR3"].value,
        "val_bengkak_tgl"           : ws["AS3"].value,
        "val_bengkak_wkt"           : ws["AT3"].value,
        "val_bengkak_hr"            : ws["AU3"].value,
        "val_lokal_lain"            : ws["AV3"].value,
        "val_lokal_lain1"           : ws["AW3"].value,
        "val_lokal_lain1_nama"      : ws["AX3"].value,
        "val_lokal_lain1_tgl"       : ws["AY3"].value,
        "val_lokal_lain1_wkt"       : ws["AZ3"].value,
        "val_lokal_lain1_hr"        : ws["BA3"].value,
        "val_lokal_lain2"           : ws["BB3"].value,
        "val_lokal_lain2_nama"      : ws["BC3"].value,
        "val_lokal_lain2_tgl"       : ws["BD3"].value,
        "val_lokal_lain2_wkt"       : ws["BE3"].value,
        "val_lokal_lain2_hr"        : ws["BF3"].value,
        "val_lokal_lain3"           : ws["BG3"].value,
        "val_lokal_lain3_nama"      : ws["BH3"].value,
        "val_lokal_lain3_tgl"       : ws["BI3"].value,
        "val_lokal_lain3_wkt"       : ws["BJ3"].value,
        "val_lokal_lain3_hr"        : ws["BK3"].value,
        "val_lokal_lain4"           : ws["BL3"].value,
        "val_lokal_lain4_nama"      : ws["BM3"].value,
        "val_lokal_lain4_tgl"       : ws["BN3"].value,
        "val_lokal_lain4_wkt"       : ws["BO3"].value,
        "val_lokal_lain4_hr"        : ws["BP3"].value,
        "val_sistemik"              : ws["BQ3"].value,
        "val_demam"                 : ws["BR3"].value,
        "val_demam_tgl"             : ws["BS3"].value,
        "val_demam_wkt"             : ws["BT3"].value,
        "val_demam_hr"              : ws["BU3"].value,
        "val_rewel"                 : ws["BV3"].value,
        "val_rewel_tgl"             : ws["BW3"].value,
        "val_rewel_wkt"             : ws["BX3"].value,
        "val_rewel_hr"              : ws["BY3"].value,
        "val_nangis"                : ws["BZ3"].value,
        "val_nangis_tgl"            : ws["CA3"].value,
        "val_nangis_wkt"            : ws["CB3"].value,
        "val_nangis_hr"             : ws["CC3"].value,
        "val_sistemik_lain"         : ws["CD3"].value,
        "val_sistemik_lain_1"       : ws["CE3"].value,
        "val_sistemik_lain_1_nama"  : ws["CF3"].value,
        "val_sistemik_lain_1_tgl"   : ws["CG3"].value,
        "val_sistemik_lain_1_wkt"   : ws["CH3"].value,
        "val_sistemik_lain_1_hr"    : ws["CI3"].value,
        "val_sistemik_lain_2"       : ws["CJ3"].value,
        "val_sistemik_lain_2_nama"  : ws["CK3"].value,
        "val_sistemik_lain_2_tgl"   : ws["CL3"].value,
        "val_sistemik_lain_2_wkt"   : ws["CM3"].value,
        "val_sistemik_lain_2_hr"    : ws["CN3"].value,
        "val_sistemik_lain_3"       : ws["CO3"].value,
        "val_sistemik_lain_3_nama"  : ws["CP3"].value,
        "val_sistemik_lain_3_tgl"   : ws["CQ3"].value,
        "val_sistemik_lain_3_wkt"   : ws["CR3"].value,
        "val_sistemik_lain_3_hr"    : ws["CS3"].value,
        "val_sistemik_lain_4"       : ws["CT3"].value,
        "val_sistemik_lain_4_nama"  : ws["CU3"].value,
        "val_sistemik_lain_4_tgl"   : ws["CV3"].value,
        "val_sistemik_lain_4_wkt"   : ws["CW3"].value,
        "val_sistemik_lain_4_hr"    : ws["CX3"].value,
        "val_kondisi_akhir"         : ws["CY3"].value,
        "val_diagnosis_1"           : ws["CZ3"].value,
        "val_diagnosis_2"           : ws["DA3"].value,
        "val_diagnosis_3"           : ws["DB3"].value,
        "val_kausalitas"            : ws["DC3"].value,
    }
    return data

def isi_dropdown(driver, wait_short, element_id, option_text):
    try:
        # Tunggu sampai dropdown bisa diklik
        dropdown_elem = wait_short.until(EC.element_to_be_clickable((By.ID, element_id)))
        select_obj = Select(dropdown_elem)

        # Pilih opsi berdasarkan teks
        select_obj.select_by_visible_text(option_text)
        print(f"→ Dropdown '{element_id}' diisi dengan '{option_text}'")

        # Verifikasi: pastikan opsi yang aktif sesuai
        # selected_text = select_obj.first_selected_option.text.strip()
        # if selected_text != option_text:
        #     raise Exception(f"❌ Verifikasi GAGL: '{selected_text}' ≠ '{option_text}'")
        
        # print(f"✅ Verifikasi OK: Dropdown '{element_id}' terpilih '{selected_text}'")

    except Exception as e:
        raise Exception(f"❌ GAGAL isi dropdown '{element_id}': {e}")

def save_form1(driver, button_id, toast_success_keyword=None, timeout=10, delay_after_click=1):
    try:
        print(f"⏳ Menunggu tombol '{button_id}' bisa diklik...")
        btn = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.ID, button_id))
        )
        # Scroll ke tombol
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
        time.sleep(0.5)
        try:
            btn.click()
        except Exception:
            print("⚠️ Klik biasa gagal, coba klik lewat JavaScript")
            driver.execute_script("arguments[0].click();", btn)
        print(f"🖱️ Tombol '{button_id}' diklik.")
        time.sleep(delay_after_click)
        if toast_success_keyword:
            toast = WebDriverWait(driver, timeout).until(
                EC.visibility_of_element_located((By.CLASS_NAME, "toast-success"))
            )
            toast_text = toast.text.strip()
            if toast_success_keyword.lower() not in toast_text.lower():
                raise Exception(f"Isi toast-success tidak sesuai: '{toast_text}'")
            print(f"✅ Form disimpan: {toast_text}")
        else:
            print(f"✅ Form disimpan (tanpa validasi toast).")
    except TimeoutException:
        raise Exception(f"❌ Timeout: Tombol '{button_id}' atau toast-success tidak muncul.")
        play_sound()
    except Exception as e:
        raise Exception(f"❌ Gagal menyimpan form '{button_id}': {e}")
        play_sound()

def save_form(driver, button_id, toast_success_keyword=None, timeout=10, delay_after_click=1):
    try:
        print(f"⏳ Menunggu tombol '{button_id}' bisa diklik...")
        btn = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.ID, button_id))
        )
        btn.click()
        print(f"🖱️ Tombol '{button_id}' diklik.")
        time.sleep(delay_after_click)

        if toast_success_keyword:
            toast = WebDriverWait(driver, timeout).until(
                EC.visibility_of_element_located((By.CLASS_NAME, "toast-success"))
            )
            toast_text = toast.text.strip()

            if toast_success_keyword.lower() not in toast_text.lower():
                raise Exception(f"Isi toast-success tidak sesuai: '{toast_text}'")

            print(f"✅ Form disimpan: {toast_text}")
        else:
            print(f"✅ Form disimpan (tanpa validasi toast).")

    except TimeoutException:
        raise Exception(f"❌ Timeout: Tombol '{button_id}' atau toast-success tidak muncul.")
    except Exception as e:
        raise Exception(f"❌ GAGAL menyimpan form '{button_id}': {e}")

def set_text(driver, id_, value, timeout=3):
        if value is None or str(value).strip() == "":
            return False
        try:
            el = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((By.ID, id_))
            )
        except TimeoutException:
            print(f"❌ FAILED element {id_} not found (timeout after {timeout}s)")
            return False
        try:
            el.clear()
            el.send_keys(str(value))
            print(f"→ {id_} = {value}")
            return True
        except Exception as e:
            print(f"❌ FAILED set_text {id_}: {e}")
            return False
        
def format_tanggal_ddmmyyyy(date_obj):
    if not date_obj:
        return ""
    if isinstance(date_obj, datetime.datetime):
        return date_obj.strftime("%d-%m-%Y")
    elif isinstance(date_obj, str):
        return date_obj 
    try:
        return date_obj.strftime("%d-%m-%Y")
    except Exception:
        return str(date_obj)
def isi_datepicker(driver, wait_short, field_id, tanggal_obj, timeout=15, delay=0.2):
    if not tanggal_obj:
        print(f"⚠️  Skip field {field_id} karena tanggal kosong.")
        return
    try:
        formatted = format_tanggal_ddmmyyyy(tanggal_obj)
        field = wait_short.until(EC.element_to_be_clickable((By.ID, field_id)))
        driver.execute_script(
            "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input'));",
            field, formatted
        )
        time.sleep(delay)
        field.send_keys(Keys.TAB)
        print(f"→ {field_id} = {formatted}")
        
    except Exception as e:
        print(f"❌ GAGAL set {field_id}: {e}")


def isi_time(driver, wait_short, field_id, time_obj, timeout=15, delay=0.2):
    if not time_obj:
        print(f"⚠️  Skip field {field_id} karena waktu kosong.")
        return
    try:
        # Format waktu 24 jam: "HH:mm"
        formatted = time_obj.strftime("%H:%M")

        field = wait_short.until(EC.element_to_be_clickable((By.ID, field_id)))
        driver.execute_script(
            "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input'));",
            field, formatted
        )
        time.sleep(delay)
        field.send_keys(Keys.TAB)  # supaya trigger event, tutup timepicker, dsb

        print(f"🕒 Field '{field_id}' diisi dengan waktu: {formatted}")
        
    except Exception as e:
        print(f"❌ GAGAL isi timepicker '{field_id}': {e}")

def wait_clickable(driver, by, locator, timeout=5):
    try:
        element = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((by, locator))
        )
        return element
    except TimeoutException:
        print(f"❌ Element {locator} tidak clickable setelah {timeout} detik")
        return None
def should_check(val):
        if val is None:
            return None
        hv = str(val).strip().lower()
        return hv in ("1", "yes", "ya", "true", "x", "y")
def set_checkbox(driver, element_id, should_check_value):
    cb = wait_clickable(driver, By.ID, element_id)
    if not cb:
        print(f"! Checkbox {element_id} tidak ditemukan atau tidak clickable")
        return False
    want = should_check(should_check_value)
    if want is True and not cb.is_selected():
        driver.execute_script("arguments[0].click();", cb)
        print(f"→ {element_id} checked")
    elif want is False and cb.is_selected():
        driver.execute_script("arguments[0].click();", cb)
        print(f"→ {element_id} unchecked")
    else:
        print(f"→ {element_id} left as-is (matches desired state)")
    return True

def buka_form(driver, wait, nomor_inklusi, spinner_class="spinner", chevron_class="fa-chevron-down", timeout_click=1):
    try:
        print("⏳ ...menunggu data dimuat dan spinner hilang...")
        wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, spinner_class)))
        print("✅ Loading is done.")
    except TimeoutException:
        print("⚠️ Timeout loading selesai. Continue...")

    try:
        print(f"🔍 Mencari nomor_inklusi di tabel: '{nomor_inklusi}'")
        xpath_tr = f"//tr[td[text()='{nomor_inklusi}']]"
        tr_elem = wait.until(EC.presence_of_element_located((By.XPATH, xpath_tr)))
        print("✅ Baris <TR> ditemukan berdasarkan nomor inklusi.")

        # Cari ikon chevron di dalam kolom terakhir
        chevron_xpath = f".//td[last()]//i[contains(@class, '{chevron_class}')]"
        chevron = tr_elem.find_element(By.XPATH, chevron_xpath)
        print(f"🔽 Chevron '{chevron_class}' ditemukan, coba diklik...")

        # Tunggu sampai chevron bisa diklik
        wait.until(EC.element_to_be_clickable((By.XPATH, chevron_xpath)))

        # Scroll ke chevron dan klik
        driver.execute_script("arguments[0].scrollIntoView(true);", chevron)
        time.sleep(0.3)
        driver.execute_script("arguments[0].click();", chevron)
        print("✅ Chevron diklik untuk expand.")

        time.sleep(timeout_click)  # beri waktu agar konten dimuat

    except Exception as e:
        print("❌ GAGAL menemukan atau klik chevron berdasarkan nomor inklusi.")
        print(e)
        return False

    try:
        tombol_isi_xpath = "//tr[.//td[contains(., 'Informasi Pasien')]]//button[contains(., 'ISI')]"
        tombol_isi = wait.until(EC.presence_of_element_located((By.XPATH, tombol_isi_xpath)))
        print("✅ Tombol 'ISI' ketemu")
        print("   OnClick attr:", tombol_isi.get_attribute("onclick"))

        driver.execute_script("arguments[0].scrollIntoView(true);", tombol_isi)
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", tombol_isi)
        print("🖱️ Tombol 'ISI' dklik.")

        return True

    except Exception as e:
        print("❌ GAGAL menemukan atau klik tombol 'ISI'.")
        print(e)
        return False

def set_radio(driver, group_id, value, timeout=10):
    try:
        if value is None or str(value).strip() == "":
            print(f"⚠️  Lewati set_radio: value kosong untuk group {group_id}")
            return False

        css = f"#{group_id} input[type='radio'][value='{value}']"
        el = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, css))
        )

        driver.execute_script("arguments[0].click();", el)
        print(f"✅ {group_id} = {value}")
        return True

    except Exception as e:
        print(f"❌ GAGAL memilih radio '{group_id}' = '{value}': {e}")
        return False

def isi_radio_button(driver, group_name, value):
    radio_buttons = driver.find_elements(By.NAME, group_name)
    found = False
    for radio in radio_buttons:
        if radio.get_attribute("value") == str(value):
            driver.execute_script("arguments[0].click();", radio)
            print(f"→ {group_name} = {value}")
            found = True
            break
    if not found:
        raise Exception(f"❌ Radio button dengan value '{value}' di group '{group_name}' tidak ditemukan.")

# Mapping bulan Indonesia
bulan_id = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
    5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
    9: "September", 10: "Oktober", 11: "November", 12: "Desember"
}
def format_tanggal_indo(tgl_excel):
    if isinstance(tgl_excel, datetime.datetime):
        hari = tgl_excel.day
        bulan = bulan_id[tgl_excel.month]
        tahun = tgl_excel.year
        return f"{hari:02d} {bulan} {tahun}"
    else:
        return str(tgl_excel)
def isi_date_indo(driver, wait_short, element_id, tgl_excel):
    tanggal_str = format_tanggal_indo(tgl_excel)
    try:
        input_elem = wait_short.until(EC.element_to_be_clickable((By.ID, element_id)))
        driver.execute_script("arguments[0].removeAttribute('readonly')", input_elem)
        input_elem.clear()
        input_elem.send_keys(tanggal_str + Keys.TAB)
        time.sleep(1)
        print(f"→ {element_id} = {tanggal_str}")
        return True
    except Exception as e:
        print(f"❌ GAGAL mengisi datepicker {element_id}: {e}")
        return False



