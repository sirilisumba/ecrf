import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, date, time as dtime
from openpyxl import load_workbook
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException
import time
import subprocess

try: 
    # --- file formpasien.py
    # --- Setup Selenium untuk Brave ---
    options = Options()
    options.debugger_address = "127.0.0.1:9222"
    service = Service("chromedriver.exe")
    driver = webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 20)

    # --- Load nomor inklusi dari Excel ---
    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb.active
    val_no_inklusi = str(ws['K2'].value).strip()

    # -------- Fungsi: konversi tanggal ke dd-mm-YYYY (robust) --------
    def format_tanggal_ddmmyyyy(value):
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return value.strftime("%d-%m-%Y")
        s = str(value).strip()
        if not s:
            return ""
        possible = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d %B %Y", "%d %b %Y"]
        for fmt in possible:
            try:
                dt = datetime.strptime(s, fmt)
                return dt.strftime("%d-%m-%Y")
            except Exception:
                pass
        # fallback: kembalikan string apa adanya (mungkin sudah dd-mm-YYYY)
        return s

    def set_text(driver, wait, id_, value):
        if value is None or str(value).strip() == "":
            return False
        try:
            el = wait.until(EC.visibility_of_element_located((By.ID, id_)))
            el.click()
            time.sleep(0.1)
            el.clear()
            time.sleep(0.1)
            el.send_keys(str(value))
            print(f"→ {id_} = {value}")
            return True
        except Exception as e:
            print(f"! failed set_text {id_}: {e}")
            try:
                # fallback pake javascript set value supaya lebih pasti masuk
                el = driver.find_element(By.ID, id_)
                driver.execute_script(
                    "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input'));", el, str(value)
                )
                print(f"→ {id_} set via JS fallback = {value}")
                return True
            except Exception as e2:
                print(f"! JS fallback failed for {id_}: {e2}")
                return False



    # -------- Load data dari Excel --------
    val_no_inklusi = ws["K3"].value   # itemid_58337 (text) >> No Inklusi
    val_inisial = ws["C3"].value   # itemid_58338 (text) >> Inisial
    val_jeniskelamin = ws["E3"].value   # form_group_58340 (radio value) >> Gender
    val_tgllahir = ws["D3"].value   # itemid_59060 (date) >> Tgl Lahir
    val_usia_thn = ws["P3"].value   # itemid_59061 >> Usia th
    val_usia_bln = ws["Q3"].value   # itemid_59062 >> Usia bln
    val_usia_hr = ws["R3"].value   # itemid_59063 >> Usia hari
    val_hasPengobatan = ws["O3"].value  # checkbox

    print("Start to fill-in Form Pasien")
    print(f"📘 No. inklusi from Excel: {val_no_inklusi}")
    print("Data from Excel:", val_no_inklusi, val_inisial, val_jeniskelamin, val_tgllahir, val_usia_thn, val_usia_bln, val_usia_hr, val_usia_hr)

    # -------- Connect to Brave (already opened with remote debugging) --------
    # options = Options()
    # options.debugger_address = "127.0.0.1:9222"  # pastikan Brave dijalankan dengan --remote-debugging-port=9222
    # service = Service(r"chromedriver.exe")       # chromedriver.exe ada di folder project / PATH
    # driver = webdriver.Chrome(service=service, options=options)

    # kecilkan waktu tunggu implicit (kita pakai explicit wait saat butuh)
    driver.implicitly_wait(1)

    # -------- Isi form --------
    # input("Lanjut isi formulir?")
    set_text(driver, wait, "itemid_58337", val_no_inklusi) # no_inklusi
    set_text(driver, wait, "itemid_58338", val_inisial) # inisial

    # Radio form_group_58340
    try:
        if val_jeniskelamin is not None and str(val_jeniskelamin).strip() != "":
            css = f"#form_group_58340 input[type='radio'][value='{val_jeniskelamin}']"
            radio = driver.find_element(By.CSS_SELECTOR, css)
            driver.execute_script("arguments[0].click();", radio)  # JS click to avoid intercepts
            print("→ form_group_58340 =", val_jeniskelamin)
    except Exception as e:
        print("❌ FAILED to select radio form_group_58340:", e)

    # Datepicker itemid_59060 -> dd-mm-YYYY, then close datepicker
    try:
        if val_tgllahir is not None and str(val_tgllahir).strip() != "":
            formatted = format_tanggal_ddmmyyyy(val_tgllahir)
            field = driver.find_element(By.ID, "itemid_59060")
            # set value via JS + dispatch input/change so page frameworks notice
            driver.execute_script(
                "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input')); arguments[0].dispatchEvent(new Event('change'));",
                field, formatted
            )
            time.sleep(0.15)
            # close any datepicker by moving focus (TAB)
            field.send_keys(Keys.TAB)
            print("→ itemid_59060 =", formatted)
    except Exception as e:
        print("❌ FAILED to set itemid_59060:", e)

    # itemid_59061 (text)
    try:
        if val_usia_thn is not None and str(val_usia_thn).strip() != "":
            el = driver.find_element(By.ID, "itemid_59061")
            el.clear()
            el.send_keys(str(val_usia_thn))
            print("→ itemid_59061 =", val_usia_thn)
    except Exception as e:
        print("❌ FAILED to set itemid_59061:", e)

    # itemid_59062 (text)
    try:
        if val_usia_bln is not None and str(val_usia_bln).strip() != "":
            el = driver.find_element(By.ID, "itemid_59062")
            el.clear()
            el.send_keys(str(val_usia_bln))
            print("→ itemid_59062 =", val_usia_bln)
    except Exception as e:
        print("❌ FAILED to set itemid_59062:", e)

    # itemid_59063 (text)
    try:
        if val_usia_hr is not None and str(val_usia_hr).strip() != "":
            el = driver.find_element(By.ID, "itemid_59063")
            el.clear()
            el.send_keys(str(val_usia_hr))
            print("→ itemid_59063 =", val_usia_hr)
    except Exception as e:
        print("❌ FAILED to set itemid_59063:", e)

    # Checkbox hasPengobatan
    try:
        checkbox = driver.find_element(By.ID, "hasPengobatan")
        should_check = False
        if val_hasPengobatan is not None:
            hv = str(val_hasPengobatan).strip().lower()
            if hv in ["1", "ya", "yes", "true", "x", "y"]:
                should_check = True
        if should_check and not checkbox.is_selected():
            driver.execute_script("arguments[0].click();", checkbox)
        elif (not should_check) and checkbox.is_selected():
            driver.execute_script("arguments[0].click();", checkbox)
        print("→ hasPengobatan =", should_check)
    except Exception as e:
        print("❌ FAILED to set hasPengobatan checkbox:", e)

    print("✅ Copy-paste INFORMASI PASIEN: DONE.")

    #
    #
    #
    # >>>> OPSI 1 : otomatis save, lanjut ke script berikutnya <<<<<
    # >>>> DIPAKAI HANYA JIKA BENAR-BENAR YAKIN DATA DI EXCEL BERSIH <<<<
    #
    # try:
    #     WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "btn-submit"))).click()
    #     print("→ Clicked btn-submit")
    # except Exception as e:
    #     print("❌ FAILED to click btn-submit:", e)

    # print("✅ Save FORM: DONE.")

    #
    #
    #
    #
    #
    #
    # >>>> OPSI 2 : tunggu user ENTER di Keyboard <<<<<
    #
    # intentionally do NOT call driver.quit()
    # PAUSE: tunggu Anda klik SAVE manual di browser lalu tekan ENTER di terminal
    input("👉 Klik tombol SAVE di browser, lalu tekan ENTER di terminal untuk lanjut...")
    #
    #
    #
    #
    #
    #
    # >>>> OPSI 3 : tidak perlu tekan ENTER di Keyboard <<<<<
    # >>>> setelah user klik tombol Submit, barulah script lanjut secara otomatis <<<<
    #
    # Tunggu sampai JavaScript var xSaved berisi "Data Berhasil Disimpan"
    # print("⌛ Waiting for SUBMIT button is clicked (dan xSaved = 'Data Berhasil Disimpan')...")
    # while True:
    #     try:
    #         xSaved_value = driver.execute_script("return window.xSaved;")
    #         if xSaved_value == "Data Berhasil Disimpan":
    #             print("✅ Saving succeessfully! Continue automatically...")
    #             break
    #     except:
    #         pass
    #     time.sleep(0.5)
    
        #
    # Loop ke file bukavaksinasi.py
    jawaban = input("➡️  Continue to FORM DATA VAKSINASI? (Y/N): ").strip().lower()

    if jawaban == 'y':
        print("▶️  Continue to next form...")
        # lanjutkan script
    elif jawaban == 'n':
        print("⏹️ Process stop by user.")
        exit()
    else:
        print("⚠️  Unrecognized. Please tap Y or N on your keyboard.")

    # >>>> Lanjut ke file openvaksinasi.py secara otomatis <<<<
    subprocess.run(["python", "openvaksinasi.py"])


except KeyboardInterrupt:
    print("\n⏹️ Kill process by user.")

except Exception as e:
    print(f"\n❌ An unhandled exception occurred: {e}")
