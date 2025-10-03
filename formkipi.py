import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, date, time as dtime
from openpyxl import load_workbook
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
import time
import subprocess

try: 

    # --- file formkipi.py
    # --- Setup Selenium untuk Brave ---
    options = Options()
    options.debugger_address = "127.0.0.1:9222"
    service = Service("chromedriver.exe")
    driver = webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 20)

    # --- Load nomor inklusi dari Excel ---
    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb.active
    val_no_inklusi = str(ws['K3'].value).strip()

    # ---------- Helpers ----------
    def format_tanggal_ddmmyyyy(value):
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return value.strftime("%d-%m-%Y")
        s = str(value).strip()
        if not s:
            return ""
        candidates = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d %B %Y", "%d %b %Y"]
        for fmt in candidates:
            try:
                return datetime.strptime(s, fmt).strftime("%d-%m-%Y")
            except:
                pass
        return s

    def format_time_hhmm(value):
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%H:%M")
        if isinstance(value, dtime):
            return value.strftime("%H:%M")
        s = str(value).strip()
        if not s:
            return ""
        candidates = ["%H:%M", "%H.%M", "%I:%M %p", "%H%M"]
        for fmt in candidates:
            try:
                return datetime.strptime(s, fmt).strftime("%H:%M")
            except:
                pass
        return s

    def should_check(val):
        if val is None:
            return None
        hv = str(val).strip().lower()
        return hv in ("1", "yes", "ya", "true", "x", "y")

    def safe_find(driver, locator, by=By.ID, timeout=3):
        try:
            return WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((by, locator))
            )
        except:
            return None

    def set_text(driver, id_, value, timeout=3):
        if value is None or str(value).strip() == "":
            return False
        try:
            el = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((By.ID, id_))
            )
        except TimeoutException:
            print(f"❌ FAILED element {id_} not found (timeout after {timeout}s)")
            return False
        try:
            el.clear()
            el.send_keys(str(value))
            print(f"→ {id_} = {value}")
            return True
        except Exception as e:
            print(f"❌ FAILED set_text {id_}: {e}")
            return False

    def set_radio(driver, group_id, value):
        if value is None or str(value).strip() == "":
            return False
        css = f"#{group_id} input[type='radio'][value='{value}']"
        el = safe_find(driver, css, By.CSS_SELECTOR, timeout=2)
        if not el:
            print(f"❌ FAILED radio {group_id} value {value} not found")
            return False
        try:
            driver.execute_script("arguments[0].click();", el)
            print(f"→ {group_id} = {value}")
            return True
        except Exception as e:
            print(f"❌ FAILED set_radio {group_id}: {e}")
            return False

    def set_date(driver, id_, value):
        if value is None or str(value).strip() == "":
            return False
        formatted = format_tanggal_ddmmyyyy(value)
        el = safe_find(driver, id_, By.ID, timeout=3)
        if not el:
            print(f"❌ FAILED date element {id_} not found")
            return False
        try:
            # set via JS + dispatch events so frameworks pick it up
            driver.execute_script(
                "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input')); arguments[0].dispatchEvent(new Event('change'));",
                el, formatted
            )
            time.sleep(0.12)
            el.send_keys(Keys.TAB)  # close datepicker
            print(f"→ {id_} = {formatted}")
            return True
        except Exception as e:
            print(f"❌ FAILED set_date {id_}: {e}")
            return False

    def set_time(driver, id_, value):
        if value is None or str(value).strip() == "":
            return False
        formatted = format_time_hhmm(value)
        el = safe_find(driver, id_, By.ID, timeout=3)
        if not el:
            print(f"❌ FAILED time element {id_} not found")
            return False
        try:
            driver.execute_script(
                "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input')); arguments[0].dispatchEvent(new Event('change'));",
                el, formatted
            )
            time.sleep(0.12)
            el.send_keys(Keys.TAB)  # close timepicker
            print(f"→ {id_} = {formatted}")
            return True
        except Exception as e:
            print(f"❌ FAILED set_time {id_}: {e}")
            return False

    # ---------- Load Excel AO2..BB2 ----------
    wb = openpyxl.load_workbook("data.xlsx", data_only=True)
    ws = wb.active

    val_no_inklusi = ws["K3"].value   # itemid_58646 (text)      >> 1. no inklusi
    val_inisial = ws["C3"].value   # itemid_58647 (text)      >> 2. inisial

    val_kategori = ws["AD3"].value   # form_group_58650 (radio) >> 3. Kategori KIPI (serius/non-serius):
    val_lokal = ws["AE3"].value   # form_group_58766 (radio) >> 3A.Reaksi Lokal:
    val_nyeri = ws["AF3"].value   # form_group_58767 (radio) >>  3A1. Nyeri lokal:
    val_lokal_tgl = ws["AG3"].value   # itemid_58772 (date picker)   >> 3A11. Nyeri lokal (tgl)
    val_lokal_wkt = ws["AH3"].value   # itemid_58773 (time picker)   >> 3A12. Nyeri lokal (wkt)
    val_lokal_hr = ws["AI3"].value   # itemid_58846 (text)          >> 3A13. Nyeri lokal (hari)
    val_merah = ws["AJ3"].value   # form_group_58768 (radio) >>  3A2. Kemerahan:
    val_merah_tgl = ws["AK3"].value   # itemid_58775 (date picker)   >> 3A21. Kemerahan (tgl)
    val_merah_wkt = ws["AL3"].value   # itemid_58799 (time picker)   >> 3A22. Kemerahan (wkt)
    val_merah_hr = ws["AM3"].value   # itemid_58847 (text)          >> 3A2. Kemerahan (hari)
    val_tebal = ws["AN3"].value   # form_group_58769 (radio) >>  3A3. Penebalan:
    val_tebal_tgl = ws["AO3"].value   # itemid_58777 (date picker)   >> 3A31. Penebalan (tgl)
    val_tebal_wkt = ws["AP3"].value   # itemid_58801 (time picker)   >> 3A32. Penebalan (wkt)
    val_tebal_hr = ws["AQ3"].value   # itemid_58848 (text)          >> 3A33. Penebalan (hari)
    val_bengkak = ws["AR3"].value   # form_group_58770 (radio) >>  3A4. Pembengkakan:
    val_bengkak_tgl = ws["AS3"].value   # itemid_58779 (date picker)   >> 3A41. Pembengkakan (tgl)
    val_bengkak_wkt = ws["AT3"].value   # itemid_58803 (time picker)   >> 3A42. Pembengkakan (wkt)
    val_bengkak_hr = ws["AU3"].value   # itemid_58849 (text)          >> 3A43. Pembengkakan (hari)
    val_lokal_lain = ws["AV3"].value   # form_group_58782 (radio) >> 3A5. Lain-lain:
    val_lokal_lain1 = ws["AW3"].value   # form_group_59084 (radio) >>  3A51. Lain-lain 1:
    val_lokal_lain1_nama = ws["AX3"].value   # itemid_59088 (text)          >> 3A511. Lain-lain 1 (nama)
    val_lokal_lain1_tgl = ws["AY3"].value   # itemid_59089 (date picker)   >> 3A512. Lain-lain 1 (tgl)
    val_lokal_lain1_wkt = ws["AZ3"].value   # itemid_59090 (time picker)   >> 3A513. Lain-lain 1 (wkt)
    val_lokal_lain1_hr = ws["BA3"].value   # itemid_59091 (text)          >> 3A514. Lain-lain 1 (hari)
    val_lokal_lain2 = ws["BB3"].value   # form_group_59085 (radio) >>  3A52. Lain-lain 2:
    val_lokal_lain2_nama = ws["BC3"].value   # itemid_59092 (text)          >> 3A521. Lain-lain 2 (nama)
    val_lokal_lain2_tgl = ws["BD3"].value   # itemid_59093 (date picker)   >> 3A522. Lain-lain 2 (tgl)
    val_lokal_lain2_wkt = ws["BE3"].value   # itemid_59149 (time picker)   >> 3A523. Lain-lain 2 (wkt)
    val_lokal_lain2_hr = ws["BF3"].value   # itemid_59150 (text)          >> 3A523. Lain-lain 2 (hari)
    val_lokal_lain3 = ws["BG3"].value   # form_group_59086 (radio) >>  3A53. Lain-lain 3:
    val_lokal_lain3_nama = ws["BH3"].value   # itemid_59095 (text)          >> 3A531. Lain-lain 3 (nama)
    val_lokal_lain3_tgl = ws["BI3"].value   # itemid_59096 (date picker)   >> 3A532. Lain-lain 3 (tgl)
    val_lokal_lain3_wkt = ws["BJ3"].value   # itemid_59151 (time picker)   >> 3A533. Lain-lain 3 (wkt)
    val_lokal_lain3_hr = ws["BK3"].value   # itemid_59152 (text)          >> 3A534. Lain-lain 3 (hari)
    val_lokal_lain4 = ws["BL3"].value   # form_group_59087 (radio) >>  3A54. Lain-lain 4:
    val_lokal_lain4_nama = ws["BM3"].value   # itemid_59098 (text)          >> 3A541. Lain-lain 4 (nama)
    val_lokal_lain4_tgl = ws["BN3"].value   # itemid_59099 (date picker)   >> 3A542. Lain-lain 4 (tgl)
    val_lokal_lain4_wkt = ws["BO3"].value   # itemid_59153 (time picker)   >> 3A543. Lain-lain 4 (wkt)
    val_lokal_lain4_hr = ws["BP3"].value   # itemid_59154 (text)          >> 3A454. Lain-lain 4 (hari)

    val_sistemik = ws["BQ3"].value   # form_group_58781 (radio) >> 3B. Reaksi Sistemik:
    val_demam = ws["BR3"].value   # form_group_58795 (radio) >>  3B1. Demam
    val_demam_tgl = ws["BS3"].value   # itemid_58808 (date picker)   >>  3B11. Demam (tgl)
    val_demam_wkt = ws["BT3"].value   # itemid_58809 (time picker)   >>  3B12. Demam (wkt)
    val_demam_hr = ws["BU3"].value   # itemid_58851 (text)          >>  3B13. Demam (hari)
    val_rewel = ws["BV3"].value   # form_group_58796 (radio) >> 3B2. Rewel:
    val_rewel_tgl = ws["BW3"].value   # itemid_58811 (date picker)   >>  3B21. Rewel (tgl)
    val_rewel_wkt = ws["BX3"].value   # itemid_58828 (time picker)   >>  3B22. Rewel (wkt)
    val_rewel_hr = ws["BY3"].value   # itemid_58852 (text)          >>  3B23. Rewel (hari)
    val_nangis = ws["BZ3"].value   # form_group_58797 (radio) >> 3B3. Nangis:
    val_nangis_tgl = ws["CA3"].value   # itemid_58813 (date picker)   >>  3B31. Nangis (tgl)
    val_nangis_wkt = ws["CB3"].value   # itemid_58830 (time picker)   >>  3B32. Nangis (wkt)
    val_nangis_hr = ws["CC3"].value   # itemid_58853 (text)          >>  3B33. Nangis (hari)
    val_sistemik_lain = ws["CD3"].value   # form_group_58807 (radio) >> 3B4. Lain-lain:
    val_sistemik_lain_1 = ws["CE3"].value   # form_group_59129 (radio) >>  3B41. Lain-lain 1:
    val_sistemik_lain_1_nama = ws["CF3"].value   # itemid_59113 (text)          >> 3B411. Lain-lain 1 (nama)
    val_sistemik_lain_1_tgl = ws["CG3"].value   # itemid_59114 (date picker)   >> 3B412. Lain-lain 1 (tgl)
    val_sistemik_lain_1_wkt = ws["CH3"].value   # itemid_59141 (time picker)   >> 3B413. Lain-lain 1 (wkt)
    val_sistemik_lain_1_hr = ws["CI3"].value   # itemid_59142 (text)          >> 3B414. Lain-lain 1 (hari)
    val_sistemik_lain_2 = ws["CJ3"].value   # form_group_59130 (radio) >>  3B42. Lain-lain 2:
    val_sistemik_lain_2_nama = ws["CK3"].value   # itemid_59116 (text)          >> 3B421. Lain-lain 2 (nama)
    val_sistemik_lain_2_tgl = ws["CL3"].value   # itemid_59117 (date picker)   >> 3B422. Lain-lain 2 (tgl)
    val_sistemik_lain_2_wkt = ws["CM3"].value   # itemid_59143 (time picker)   >> 3B423. Lain-lain 2 (wkt)
    val_sistemik_lain_2_hr = ws["CN3"].value   # itemid_59144 (text)          >> 3B424. Lain-lain 2 (hari)
    val_sistemik_lain_3 = ws["CO3"].value   # form_group_59131 (radio) >>  3B43. Lain-lain 3:
    val_sistemik_lain_3_nama = ws["CP3"].value   # itemid_59119 (text)          >> 3B431. Lain-lain 3 (nama)
    val_sistemik_lain_3_tgl = ws["CQ3"].value   # itemid_59120 (date picker)   >> 3B432. Lain-lain 3 (tgl)
    val_sistemik_lain_3_wkt = ws["CR3"].value   # itemid_59145 (time picker)   >> 3B433. Lain-lain 3 (wkt)
    val_sistemik_lain_3_hr = ws["CS3"].value   # itemid_59146 (text)          >> 3B434. Lain-lain 3 (hari)
    val_sistemik_lain_4 = ws["CT3"].value   # form_group_59132 (radio) >>  3B44. Lain-lain 4:
    val_sistemik_lain_4_nama = ws["CU3"].value   # itemid_59122 (text)          >> 3B441. Lain-lain 4 (nama)
    val_sistemik_lain_4_tgl = ws["CV3"].value   # itemid_59123 (date picker)   >> 3B442. Lain-lain 4 (tgl)
    val_sistemik_lain_4_wkt = ws["CW3"].value   # itemid_59147 (time picker)   >> 3B443. Lain-lain 4 (wkt)
    val_sistemik_lain_4_hr = ws["CX3"].value   # itemid_59148 (text)          >> 3B444. Lain-lain 4 (hari)

    val_kondisi_akhir = ws["CY3"].value   # form_group_58827 (radio) >>      3C. Sembuh
    
    val_haspengobatan = ws["O3"].value  # hasPengobatan (checkbox) >> 4. hasPengobatan

    print("Start to fill-in Form KIPI")
    print(f"📘 No. inklusi from Excel: {val_no_inklusi}")
    print("Data from Excel:", val_no_inklusi, val_inisial, val_kategori, val_lokal, val_nyeri, val_lokal_tgl, val_lokal_wkt, val_lokal_hr, val_merah, 
    val_merah_tgl, val_merah_wkt, val_merah_hr, val_tebal, val_tebal_tgl, val_tebal_wkt, val_tebal_hr, val_bengkak, val_bengkak_tgl, 
    val_bengkak_wkt, val_bengkak_hr, val_lokal_lain, val_lokal_lain1, val_lokal_lain1_nama, val_lokal_lain1_tgl, val_lokal_lain1_wkt, val_lokal_lain1_hr, val_lokal_lain2, val_lokal_lain2_nama, val_lokal_lain2_tgl, 
    val_lokal_lain2_wkt, val_lokal_lain2_hr, val_lokal_lain3, val_lokal_lain3_nama, val_lokal_lain3_tgl, val_lokal_lain3_wkt, val_lokal_lain3_hr, val_lokal_lain4, val_lokal_lain4_nama, val_lokal_lain4_tgl, val_lokal_lain4_wkt, val_lokal_lain4_hr, 
    val_sistemik, val_demam, val_demam_tgl, val_demam_wkt, val_demam_hr, val_rewel, val_rewel_tgl, val_rewel_wkt, val_rewel_hr, val_nangis, val_nangis_tgl, val_nangis_wkt, 
    val_nangis_hr, val_sistemik_lain, val_sistemik_lain_1, val_sistemik_lain_1_nama, val_sistemik_lain_1_tgl, val_sistemik_lain_1_wkt, val_sistemik_lain_1_hr, val_sistemik_lain_2, val_sistemik_lain_2_nama, val_sistemik_lain_2_tgl, val_sistemik_lain_2_wkt, val_sistemik_lain_2_hr, 
    val_sistemik_lain_3, val_sistemik_lain_3_nama, val_sistemik_lain_3_tgl, val_sistemik_lain_3_wkt, val_sistemik_lain_3_hr, val_sistemik_lain_4, val_sistemik_lain_4_nama, val_sistemik_lain_4_tgl, val_sistemik_lain_4_wkt, val_sistemik_lain_4_hr, val_kondisi_akhir, val_haspengobatan)      

    # ---------- Connect to existing Brave session ----------
    # options = Options()
    # options.debugger_address = "127.0.0.1:9222"  # pastikan Brave dibuka dgn flag remote debugging
    # service = Service(r"chromedriver.exe")       # chromedriver.exe di folder project / PATH
    # driver = webdriver.Chrome(service=service, options=options)
    # driver.implicitly_wait(1)

    # ---------- Fill form ----------
    # text fields
    set_text(driver, "itemid_58646", val_no_inklusi)  # no_inklusi
    set_text(driver, "itemid_58647", val_inisial)  # inisial

    # main radio group 58650 >>> Kategori KIPI Serius atau Non Serius: 
    if set_radio(driver, "form_group_58650", val_kategori):
        # only handle nested block if value == 2
        try:
            if str(val_kategori).strip() == "2": # 
                time.sleep(0.35)  # wait UI to render

                # AR2 -> form_group_58766 >>> REAKSI LOKAL
                if set_radio(driver, "form_group_58766", val_lokal):
                    if str(val_lokal).strip() == "1":
                        time.sleep(0.25)
                        # AS2 -> form_group_58767 >>> NYERI LOKAL
                        if set_radio(driver, "form_group_58767", val_nyeri):
                            if str(val_nyeri).strip() == "1":
                                set_date(driver, "itemid_58772", val_lokal_tgl) # AT2 (tgl)
                                set_time(driver, "itemid_58773", val_lokal_wkt) # AU2 (wkt)
                                set_text(driver, "itemid_58846", val_lokal_hr) # AV2 (hari)
                        # AW2 -> form_group_58768 >>> KEMERAHAN
                        if set_radio(driver, "form_group_58768", val_merah):
                            if str(val_merah).strip() == "1":
                                set_date(driver, "itemid_58775", val_merah_tgl) # AX2 (tgl)
                                set_time(driver, "itemid_58799", val_merah_wkt) # AY2 (wkt)
                                set_text(driver, "itemid_58847", val_merah_hr) # AZ2 (hari)
                        # BA2 -> form_group_58769 >>> PENEBALAN
                        if set_radio(driver, "form_group_58769", val_tebal):
                            if str(val_tebal).strip() == "1":
                                set_date(driver, "itemid_58777", val_tebal_tgl) # BB2 (tgl)
                                set_time(driver, "itemid_58801", val_tebal_wkt) # BC2 (wkt)
                                set_text(driver, "itemid_58848", val_tebal_hr) # BD2 (hari)
                        # BE2 -> form_group_58770 >>> PEMBENGKAKAN
                        if set_radio(driver, "form_group_58770", val_bengkak):
                            if str(val_bengkak).strip() == "1":
                                set_date(driver, "itemid_58779", val_bengkak_tgl) # BF2 (tgl)
                                set_time(driver, "itemid_58803", val_bengkak_wkt) # BG2 (wkt)
                                set_text(driver, "itemid_58849", val_bengkak_hr) # BH2 (hari)
                        # BI2 -> form_group_58782 >>> LAIN-LAIN
                        if set_radio(driver, "form_group_58782", val_lokal_lain):
                            if str(val_lokal_lain).strip() == "1":
                                time.sleep(0.25)
                                # BJ2 -> form_group_59084 >>> lain-lain 1
                                if set_radio(driver, "form_group_59084", val_lokal_lain1):
                                    if str(val_lokal_lain1).strip() == "1":
                                        set_text(driver, "itemid_59088", val_lokal_lain1_nama) # BK2 (nama)
                                        set_date(driver, "itemid_59089", val_lokal_lain1_tgl) # BL2 (tgl)
                                        set_time(driver, "itemid_59090", val_lokal_lain1_wkt) # BM2 (wkt)
                                        set_text(driver, "itemid_59091", val_lokal_lain1_hr) # BN2 (hari)
                                # BO2 -> form_group_59085 >>> lain-lain 2
                                if set_radio(driver, "form_group_59085", val_lokal_lain2):
                                    if str(val_lokal_lain2).strip() == "1":
                                        set_text(driver, "itemid_59092", val_lokal_lain2_nama) # BP2 (nama)
                                        set_date(driver, "itemid_59093", val_lokal_lain2_tgl) # BQ2 (tgl)
                                        set_time(driver, "itemid_59149", val_lokal_lain2_wkt) # BR2 (wkt)
                                        set_text(driver, "itemid_59150", val_lokal_lain2_hr) # BS2 (hari)
                                # BT2 -> form_group_59086 >>> lain-lain 3
                                if set_radio(driver, "form_group_59086", val_lokal_lain3):
                                    if str(val_lokal_lain3).strip() == "1":
                                        set_text(driver, "itemid_59095", val_lokal_lain3_nama) # BU2 (nama)
                                        set_date(driver, "itemid_59096", val_lokal_lain3_tgl) # BV2 (tgl)
                                        set_time(driver, "itemid_59151", val_lokal_lain3_wkt) # BW2 (wkt)
                                        set_text(driver, "itemid_59152", val_lokal_lain3_hr) # BX2 (hari)
                                # BY2 -> form_group_59087 >>> lain-lain 4
                                if set_radio(driver, "form_group_59087", val_lokal_lain4):
                                    if str(val_lokal_lain4).strip() == "1":
                                        set_text(driver, "itemid_59098", val_lokal_lain4_nama) # BZ2 (nama)
                                        set_date(driver, "itemid_59099", val_lokal_lain4_tgl) # CA2 (tgl)
                                        set_time(driver, "itemid_59153", val_lokal_lain4_wkt) # CB2 (wkt)
                                        set_text(driver, "itemid_59154", val_lokal_lain4_hr) # CC2 (hari)

                # CD2 -> form_group_58781 >>> SISTEMIK
                if set_radio(driver, "form_group_58781", val_sistemik):
                    if str(val_sistemik).strip() == "1":
                        time.sleep(0.25)
                        # CE2 -> form_group_58795 >>> DEMAM
                        if set_radio(driver, "form_group_58795", val_demam):
                            if str(val_demam).strip() == "1":
                                set_date(driver, "itemid_58808", val_demam_tgl) # CF2 (tgl)
                                set_time(driver, "itemid_58809", val_demam_wkt) # CG2 (wkt)
                                set_text(driver, "itemid_58851", val_demam_hr) # CH2 (hari)
                        # CI2 -> form_group_58796 >>> REWEL
                        if set_radio(driver, "form_group_58796", val_rewel):
                            if str(val_rewel).strip() == "1":
                                set_date(driver, "itemid_58811", val_rewel_tgl) # CF2 (tgl)
                                set_time(driver, "itemid_58828", val_rewel_wkt) # CG2 (wkt)
                                set_text(driver, "itemid_58852", val_rewel_hr) # CH2 (hari)
                        # CM2 -> form_group_58797 >>> NANGIS
                        if set_radio(driver, "form_group_58797", val_nangis):
                            if str(val_nangis).strip() == "1":
                                set_date(driver, "itemid_58813", val_nangis_tgl) # CF2 (tgl)
                                set_time(driver, "itemid_58830", val_nangis_wkt) # CG2 (wkt)
                                set_text(driver, "itemid_58853", val_nangis_hr) # CH2 (hari)
                        # CQ2 -> form_group_58807 >>> LAIN-LAIN
                        if set_radio(driver, "form_group_58807", val_sistemik_lain):
                            if str(val_sistemik_lain).strip() == "1":
                                time.sleep(0.25)
                                # CR2 -> form_group_59129 >>> lain-lain 1
                                if set_radio(driver, "form_group_59129", val_sistemik_lain_1):
                                    if str(val_sistemik_lain_1).strip() == "1":
                                        set_text(driver, "itemid_59113", val_sistemik_lain_1_nama) # CS2 (nama)
                                        set_date(driver, "itemid_59114", val_sistemik_lain_1_tgl) # CT2 (tgl)
                                        set_time(driver, "itemid_59141", val_sistemik_lain_1_wkt) # CU2 (wkt)
                                        set_text(driver, "itemid_59142", val_sistemik_lain_1_hr) # CV2 (hari)
                                # CW2 -> form_group_59130 >>> lain-lain 2
                                if set_radio(driver, "form_group_59130", val_sistemik_lain_2):
                                    if str(val_sistemik_lain_2).strip() == "1":
                                        set_text(driver, "itemid_59116", val_sistemik_lain_2_nama) # CX2 (nama)
                                        set_date(driver, "itemid_59117", val_sistemik_lain_2_tgl) # CY2 (tgl)
                                        set_time(driver, "itemid_59143", val_sistemik_lain_2_wkt) # CZ2 (wkt)
                                        set_text(driver, "itemid_59144", val_sistemik_lain_2_hr) # DA2 (hari)
                                # DB2 -> form_group_59131 >>> lain-lain 3
                                if set_radio(driver, "form_group_59131", val_sistemik_lain_3):
                                    if str(val_sistemik_lain_3).strip() == "1":
                                        set_text(driver, "itemid_59119", val_sistemik_lain_3_nama) # DC2 (nama)
                                        set_date(driver, "itemid_59120", val_sistemik_lain_3_tgl) # DD2 (tgl)
                                        set_time(driver, "itemid_59145", val_sistemik_lain_3_wkt) # DE2 (wkt)
                                        set_text(driver, "itemid_59146", val_sistemik_lain_3_hr) # DF2 (hari)
                                # DG2 -> form_group_59132 >>> lain-lain 4
                                if set_radio(driver, "form_group_59132", val_sistemik_lain_4):
                                    if str(val_sistemik_lain_4).strip() == "1":
                                        set_text(driver, "itemid_59122", val_sistemik_lain_4_nama) # DH2 (nama)
                                        set_date(driver, "itemid_59123", val_sistemik_lain_4_tgl) # DI2 (tgl)
                                        set_time(driver, "itemid_59147", val_sistemik_lain_4_wkt) # DJ2 (wkt)
                                        set_text(driver, "itemid_59148", val_sistemik_lain_4_hr) # DK2 (hari)
                # DL2 -> form_group_58827 >> SEMBUH
                set_radio(driver, "form_group_58827", val_kondisi_akhir)
        except Exception as e:
            print("❌ FAILED error handling nested radios:", e)
    else:
        print("→ main radio form_group_58650 not set (or empty)")

    # checkbox hasPengobatan (BB2)
    cb = safe_find(driver, "hasPengobatan", By.ID, timeout=2)
    if cb:
        want = should_check(val_haspengobatan)
        if want is True and not cb.is_selected():
            driver.execute_script("arguments[0].click();", cb)
            print("→ hasPengobatan checked")
        elif want is False and cb.is_selected():
            driver.execute_script("arguments[0].click();", cb)
            print("→ hasPengobatan unchecked")
        else:
            print("→ hasPengobatan left as-is")
    else:
        print("❌ FAILED checkbox hasPengobatan not found")

    print("✅ Copy-paste KIPI: DONE.")


    #
    #
    #
    # >>>> OPSI 1 : otomatis save, lanjut ke script berikutnya <<<<<
    #
    # try:
    #     WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "btn-submit"))).click()
    #     print("→ Clicked btn-submit")
    # except Exception as e:
    #     print("❌ FAILED to click btn-submit:", e)

    # print("✅ Save FORM: DONE.")
    #
    #
    #
    #
    #
    #
    # >>>> OPSI 2 : tunggu user ENTER di Keyboard <<<<<
    #
    # intentionally do NOT call driver.quit()
    # PAUSE: tunggu Anda klik SAVE manual di browser lalu tekan ENTER di terminal
    input("👉 Klik tombol SAVE di browser, lalu tekan ENTER di terminal untuk lanjut...")
    #
    #
    #
    #
    #
    #
    # >>>> OPSI 3 : tidak perlu tekan ENTER di Keyboard <<<<<
    # >>>> setelah user klik tombol Submit, otomatis lanjut <<<<
    #
    # Tunggu sampai JavaScript var xSaved berisi "Data Berhasil Disimpan"
    # print("⌛ Waiting for SUBMIT button is clicked (dan xSaved = 'Data Berhasil Disimpan')...")
    # while True:
    #     try:
    #         xSaved_value = driver.execute_script("return window.xSaved;")
    #         if xSaved_value == "Data Berhasil Disimpan":
    #             print("✅ Saving succeessfully! Continue automatically to deleting row 2 in Excel.")
    #             break
    #     except:
    #         pass
    #     time.sleep(0.5)
    #
    #
    #
    #
    #
    #
    # Loop ke file deleterow.py
    jawaban = input("➡️  Continue to delete row 3? (Y/N): ").strip().lower()

    if jawaban == 'y':
        print("▶️  Continue next process...")
        # lanjutkan script
    elif jawaban == 'n':
        print("⏹️ Process stop by user.")
        exit()
    else:
        print("⚠️  Unrecognized. Please tap Y or N on your keyboard.")

    # >>>> Lanjut ke file deleterow.py secara otomatis <<<<
    print("▶️  Continue deleting row...")
    subprocess.run(["python", "deleterow.py"])


except KeyboardInterrupt:
    print("\n⏹️ Kill process by user.")

except Exception as e:
    print(f"\n❌ An unhandled exception occurred: {e}")
