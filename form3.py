import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, date, time as dtime
from openpyxl import load_workbook
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException
import time
import subprocess

try: 
    # --- file form3.py
    # Setup Brave (attach ke existing session)
    #
    options = Options()
    options.debugger_address = "127.0.0.1:9222"  # koneksi ke Brave yang sudah terbuka
    service = Service(r"chromedriver.exe")
    driver = webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 60)

    # --- Load nomor inklusi dari Excel ---
    wb = load_workbook('data.xlsx')
    ws = wb.active

    val_no_inklusi = ws['K3'].value

    # tunggu sampai form dengan id 'myModalTambahNomorInklusi' muncul, maksimal 30 detik
    # print("⏳ Waiting for button 'ISI' is clicked and modal 'Tambah Nomor Inklusi' muncul...")
    # wait = WebDriverWait(driver, 9999)
    # modal = wait.until(EC.visibility_of_element_located((By.ID, "myModalTambahNomorInklusi")))
    # print("✅ Modal is loaded! Continue to copy-paste form TAMBAH NOMOR INKLUSI...")
    #
    # lanjut isi form atau aksi lainnya
    #
    # --- Fungsi untuk format tanggal ---
    def format_tanggal_1(value):
        if isinstance(value, datetime):
            return f"{value.day} {value.strftime('%B %Y')}"  # 5 September 2025
        return str(value)

    # --- Load data dari Excel ---
    val_tgl_inklusi = ws["J3"].value
    val_no_inklusi   = ws["K3"].value
    val_inisial    = ws["C3"].value
    val_keterangan      = ws["L3"].value

    print(f"📘 No. inklusi from Excel: {val_no_inklusi}")
    print("Data from Excel:", val_tgl_inklusi, val_no_inklusi, val_inisial, val_keterangan)

    # --- Isi form ---

    # Tanggal inklusi (date picker)
    if val_tgl_inklusi:
        field_tgl = driver.find_element(By.ID, "tanggal_inklusi_s")
        field_tgl.clear()
        field_tgl.send_keys(format_tanggal_1(val_tgl_inklusi))
        time.sleep(0.5)  # kasih jeda supaya isi masuk
        field_tgl.send_keys("\t")  # tekan TAB supaya datepicker tertutup
        print("→ tanggal_inklusi_s =", format_tanggal_1(val_tgl_inklusi))

    # Nomor inklusi
    if val_no_inklusi:
        driver.find_element(By.ID, "nomor_inklusi").clear()
        driver.find_element(By.ID, "nomor_inklusi").send_keys(str(val_no_inklusi))
        print("→ nomor_inklusi =", val_no_inklusi)


    # Inisial responden
    if val_inisial:
        driver.find_element(By.ID, "inisial_responden").clear()
        driver.find_element(By.ID, "inisial_responden").send_keys(str(val_inisial))
        print("→ inisial_responden =", val_inisial)

    # Keterangan (boleh kosong)
    if val_keterangan:
        driver.find_element(By.ID, "keterangan").clear()
        driver.find_element(By.ID, "keterangan").send_keys(str(val_keterangan))
        print("→ keterangan =", val_keterangan)

    print("✅ Copy-paste TAMBAH NOMOR INKLUSI: DONE.")
    #
    #
    # >>>> OPSI 1 : otomatis save, lanjut ke script berikutnya <<<<<
    #
    # try:
    #     WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "btn-add-nomor-inklusi"))).click()
    #     print("→ Clicked btn-add-nomor-inklusi")
    # except Exception as e:
    #     print("❌ FAILED to click btn-add-nomor-inklusi:", e)

    # print("✅ Save FORM: DONE.")
    #
    #
    #
    #
    #
    #
    # >>>> OPSI 2 : tunggu user ENTER di Keyboard <<<<<
    #
    # intentionally do NOT call driver.quit()
    # PAUSE: tunggu Anda klik SAVE manual di browser lalu tekan ENTER di terminal
    # input("👉 Klik tombol SAVE di browser, lalu tekan ENTER di terminal untuk lanjut...")
    #
    #
    #
    #
    #
    #
    # >>>> OPSI 3 : tidak perlu tekan ENTER di Keyboard <<<<<
    # >>>> setelah user klik tombol Submit/Save, lalu klik manual tombol Isi <<<<
    #
    btn_simpan = WebDriverWait(driver, 15).until(
        EC.element_to_be_clickable((By.ID, "btn-add-nomor-inklusi"))
    )
    btn_simpan.click()
    time.sleep(3)

    try:
        # Tunggu sampai elemen dengan class toast-success muncul dan mengandung teks tertentu
        WebDriverWait(driver, 15).until(
            EC.text_to_be_present_in_element(
                (By.CLASS_NAME, "toast-success"),
                "Berhasil Menyimpan Nomor Inklusi"
            )
        )
        print("✅ Success notification is loaded: Berhasil Menyimpan Nomor Inklusi.")
    except:
        print("❌ Success notification is not loaded in time frame.")

    #
    #
    # Loop ke file openpelapor.py
    jawaban = input("➡️  Continue to FORM INFORMASI PELAPOR? (Y/N): ").strip().lower()

    if jawaban == 'y':
        print("▶️  Continue to next form...")
        # lanjutkan script
    elif jawaban == 'n':
        print("⏹️ Process stop by user.")
        exit()
    else:
        print("⚠️  Unrecognized. Please tap Y or N on your keyboard.")

    # >>>> Lanjut ke file openpelapor.py secara otomatis <<<<
    subprocess.run(["python", "formpelapor.py"])

except KeyboardInterrupt:
    print("\n⏹️ Kill process by user.")

except Exception as e:
    print(f"\n❌ An unhandled exception occurred: {e}")


